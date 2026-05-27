#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Build monthly customer performance reports from Vertu/Odoo sales detail exports."""

from __future__ import annotations

import argparse
import json
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REQUIRED_COLUMNS = ["销售日期", "实际业绩"]
MONTH_COLUMNS = [f"{i}月" for i in range(1, 13)]

FIELD_ALIASES = {
    "sale_date": ["销售日期", "日期", "订单日期"],
    "customer_name": ["客户名称", "客户", "客户名"],
    "receiver_name": ["收货人名称", "收货人"],
    "actual_performance": ["实际业绩", "实际成交金额(CNY)", "CNY成交金额", "成交金额"],
    "channel": ["渠道"],
    "department_2": ["二级部门"],
    "department": ["部门"],
    "salesperson": ["销售员"],
    "is_refund": ["是否退款"],
    "product_code": ["产品编码"],
    "product_name": ["存货名称"],
    "warehouse": ["出货仓库"],
    "quantity": ["数量"],
    "actual_quantity": ["实际数量"],
}


def find_column(columns: List[str], aliases: List[str]) -> str | None:
    clean = {str(c).strip(): c for c in columns}
    for alias in aliases:
        if alias in clean:
            return clean[alias]
    return None


def normalize_money(value: Any) -> float:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in {"", "-", "--", "nan", "None"}:
        return 0.0
    text = text.replace(",", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", "."}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def normalize_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def read_source_sheets(input_path: Path, sheet: str | None, all_sheets: bool) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    xls = pd.ExcelFile(input_path)
    sheets = [sheet] if sheet else (xls.sheet_names if all_sheets else [xls.sheet_names[0]])
    frames: List[pd.DataFrame] = []
    skipped: List[Dict[str, Any]] = []

    for sheet_name in sheets:
        df = pd.read_excel(input_path, sheet_name=sheet_name)
        df.columns = [normalize_text(c) for c in df.columns]
        missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
        if missing:
            skipped.append({"sheet": sheet_name, "reason": f"missing columns: {', '.join(missing)}"})
            continue
        df["__source_sheet"] = sheet_name
        frames.append(df)

    if not frames:
        raise SystemExit(f"No usable sheets found in {input_path}. Skipped: {skipped}")
    return pd.concat(frames, ignore_index=True), skipped


def build_clean_frame(df: pd.DataFrame, args: argparse.Namespace) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    columns = list(df.columns)
    colmap = {name: find_column(columns, aliases) for name, aliases in FIELD_ALIASES.items()}

    sale_date_col = colmap["sale_date"]
    metric_col = find_column(columns, [args.metric]) or colmap["actual_performance"]
    customer_col = find_column(columns, [args.customer_field]) if args.customer_field else colmap["customer_name"]
    receiver_col = colmap["receiver_name"]

    if not sale_date_col or not metric_col:
        raise SystemExit("Source data must include sale date and performance metric columns.")

    clean = pd.DataFrame()
    clean["sale_date"] = pd.to_datetime(df[sale_date_col], errors="coerce")
    primary_customer = df[customer_col].map(normalize_text) if customer_col else ""
    fallback_customer = df[receiver_col].map(normalize_text) if receiver_col else ""
    clean["customer_name"] = [p if p else (f if f else "(未识别客户)") for p, f in zip(primary_customer, fallback_customer)]
    clean["actual_performance"] = df[metric_col].map(normalize_money)
    clean["source_sheet"] = df.get("__source_sheet", "")

    for target in ["channel", "department_2", "department", "salesperson", "is_refund", "product_code", "product_name", "warehouse"]:
        source = colmap.get(target)
        clean[target] = df[source].map(normalize_text) if source else ""

    for target in ["quantity", "actual_quantity"]:
        source = colmap.get(target)
        clean[target] = df[source].map(normalize_money) if source else 0.0

    before = len(clean)
    clean = clean[clean["sale_date"].notna()].copy()
    clean["year"] = clean["sale_date"].dt.year.astype(int)
    clean["month"] = clean["sale_date"].dt.month.astype(int)

    if args.year:
        clean = clean[clean["year"] == args.year]
    if args.channel:
        clean = clean[clean["channel"] == args.channel]
    if args.department:
        dept_mask = (clean["department_2"] == args.department) | (clean["department"] == args.department)
        clean = clean[dept_mask]
    if args.exclude_refunds:
        refund_text = clean["is_refund"].str.lower()
        clean = clean[~refund_text.isin(["是", "yes", "true", "1", "y"])]

    quality = {
        "source_rows": before,
        "rows_after_filters": int(len(clean)),
        "blank_customer_rows": int((clean["customer_name"] == "(未识别客户)").sum()),
        "negative_performance_rows": int((clean["actual_performance"] < 0).sum()),
        "field_mapping": {k: str(v) if v else None for k, v in colmap.items()},
        "metric_column": str(metric_col),
        "customer_column": str(customer_col) if customer_col else None,
    }
    return clean, quality


def build_customer_monthly(clean: pd.DataFrame) -> pd.DataFrame:
    if clean.empty:
        table = pd.DataFrame(columns=["客户名称", *MONTH_COLUMNS, "总计"])
        return table

    pivot = clean.pivot_table(
        index="customer_name",
        columns="month",
        values="actual_performance",
        aggfunc="sum",
        fill_value=0,
    )
    for month in range(1, 13):
        if month not in pivot.columns:
            pivot[month] = 0
    pivot = pivot[[i for i in range(1, 13)]]
    pivot.columns = MONTH_COLUMNS
    pivot["总计"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("总计", ascending=False)
    pivot.insert(0, "客户名称", pivot.index)
    pivot.reset_index(drop=True, inplace=True)

    total_row = {"客户名称": "总计"}
    for col in MONTH_COLUMNS + ["总计"]:
        total_row[col] = pivot[col].sum()
    pivot = pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)
    return pivot


def write_report(table: pd.DataFrame, clean: pd.DataFrame, output_path: Path, args: argparse.Namespace, quality: Dict[str, Any]) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        table.to_excel(writer, index=False, sheet_name="客户月度业绩")
        summary = pd.DataFrame([
            ["报表主题", args.topic],
            ["源文件", str(args.input)],
            ["年份", args.year or "全部"],
            ["渠道", args.channel or "全部"],
            ["部门", args.department or "全部"],
            ["排除退款", "是" if args.exclude_refunds else "否"],
            ["明细行数", len(clean)],
            ["客户数", max(len(table) - 1, 0)],
            ["实际业绩总计", float(table.iloc[-1]["总计"]) if len(table) else 0],
        ], columns=["项目", "值"])
        summary.to_excel(writer, index=False, sheet_name="说明")
        if args.include_detail:
            clean.to_excel(writer, index=False, sheet_name="清洗明细")

    wb = load_workbook(output_path)
    ws = wb["客户月度业绩"]
    header_fill = PatternFill("solid", fgColor="F79646")
    total_fill = PatternFill("solid", fgColor="F79646")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    max_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.column >= 2:
                cell.number_format = '#,##0'
        if row[0].value == "总计":
            for cell in row:
                cell.fill = total_fill
                cell.font = Font(bold=True)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 42
    for idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 13

    note_ws = wb["说明"]
    note_ws.column_dimensions["A"].width = 18
    note_ws.column_dimensions["B"].width = 90
    for row in note_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    wb.save(output_path)


def write_quality(path: Path, args: argparse.Namespace, skipped: List[Dict[str, Any]], quality: Dict[str, Any], table: pd.DataFrame) -> None:
    total = float(table.iloc[-1]["总计"]) if len(table) else 0.0
    lines = [
        f"# {args.topic} 数据质量检查",
        "",
        f"- 源文件: `{args.input}`",
        f"- 年份: `{args.year or '全部'}`",
        f"- 渠道: `{args.channel or '全部'}`",
        f"- 部门: `{args.department or '全部'}`",
        f"- 排除退款: `{'是' if args.exclude_refunds else '否'}`",
        f"- 源行数: `{quality['source_rows']}`",
        f"- 筛选后行数: `{quality['rows_after_filters']}`",
        f"- 空客户行数: `{quality['blank_customer_rows']}`",
        f"- 负业绩行数: `{quality['negative_performance_rows']}`",
        f"- 汇总业绩: `{total:,.0f}`",
        "",
        "## 字段映射",
    ]
    for key, value in quality["field_mapping"].items():
        lines.append(f"- `{key}` => `{value}`")
    if skipped:
        lines.extend(["", "## 跳过的工作表"])
        for item in skipped:
            lines.append(f"- `{item['sheet']}`: {item['reason']}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build customer monthly performance report.")
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--topic", default="customer-monthly-performance")
    parser.add_argument("--sheet")
    parser.add_argument("--year", type=int)
    parser.add_argument("--channel")
    parser.add_argument("--department")
    parser.add_argument("--customer-field")
    parser.add_argument("--metric", default="实际业绩")
    parser.add_argument("--exclude-refunds", action="store_true")
    parser.add_argument("--all-sheets", action="store_true")
    parser.add_argument("--include-detail", action="store_true")
    args = parser.parse_args()

    args.input = args.input.resolve()
    repo_root = Path(__file__).resolve().parents[1]
    output_dir = args.output_dir.resolve()
    data_clean = repo_root / "data_clean"
    data_metrics = repo_root / "data_metrics"
    data_reports = repo_root / "data_reports"
    data_quality = repo_root / "data_quality"
    for directory in [output_dir, data_clean, data_metrics, data_reports, data_quality]:
        directory.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    topic_slug = re.sub(r"[^A-Za-z0-9_.-]+", "-", args.topic).strip("-") or "performance"

    raw_df, skipped = read_source_sheets(args.input, args.sheet, args.all_sheets)
    clean, quality = build_clean_frame(raw_df, args)
    table = build_customer_monthly(clean)

    clean_path = data_clean / f"{stamp}_{topic_slug}_clean.csv"
    metrics_path = data_metrics / f"{stamp}_{topic_slug}_customer_monthly.csv"
    report_path = data_reports / f"{stamp}_{topic_slug}_customer_monthly_performance.xlsx"
    quality_path = data_quality / f"{stamp}_{topic_slug}_quality.md"

    clean.to_csv(clean_path, index=False, encoding="utf-8-sig")
    table.to_csv(metrics_path, index=False, encoding="utf-8-sig")
    write_report(table, clean, report_path, args, quality)
    write_quality(quality_path, args, skipped, quality, table)

    result = {
        "report": str(report_path),
        "metrics": str(metrics_path),
        "clean": str(clean_path),
        "quality": str(quality_path),
        "rows": int(len(clean)),
        "customers": max(int(len(table) - 1), 0),
        "total_actual_performance": float(table.iloc[-1]["总计"]) if len(table) else 0.0,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()


