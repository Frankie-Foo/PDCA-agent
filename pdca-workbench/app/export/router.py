# -*- coding: utf-8 -*-
"""数据导出 API（Excel / CSV）。"""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Annotated
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select

from app.auth.deps import require_role
from app.auth.models import User
from app.auth.scope import visible_dealer_names, visible_store_ids
from app.database import get_session
from app.models.dealer_sales import DealerSales
from app.models.walkin_daily_report import WalkinDailyReport

router = APIRouter(prefix="/api/export", tags=["export"])

_ROLE_LABEL = {"viewer": "只读", "sales": "销售", "manager": "主管", "admin": "管理员"}


def _wb_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = filename.replace(" ", "_")
    encoded = quote(safe_name, encoding="utf-8")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


def _header_style(ws, row: int, cols: int) -> None:
    """给表头行添加样式。"""
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


# ── 五件套明细导出 ─────────────────────────────────────────────────────────────

@router.get("/walkin-metrics")
async def export_walkin_metrics(
    user: Annotated[User, Depends(require_role("dealer"))],
    session: Annotated[Session, Depends(get_session)],
    month: str = Query(""),
    dealer_id: str = Query(""),
):
    """导出当前账号可见范围内的门店五件套日报 Excel。"""
    stmt = select(WalkinDailyReport).order_by(
        WalkinDailyReport.report_date, WalkinDailyReport.dealer_id
    )
    if month and re.fullmatch(r"\d{4}-\d{2}", month):
        stmt = stmt.where(WalkinDailyReport.report_date.startswith(month))
    allowed = visible_store_ids(user, session)
    if dealer_id and allowed is not None and dealer_id not in allowed:
        raise HTTPException(status_code=403, detail="该门店不在当前账号的数据权限范围内")
    if allowed is not None:
        if not allowed:
            raise HTTPException(status_code=403, detail="账号未绑定可导出的门店")
        stmt = stmt.where(WalkinDailyReport.dealer_id.in_(allowed))
    if dealer_id:
        stmt = stmt.where(WalkinDailyReport.dealer_id == dealer_id)
    rows = session.exec(stmt).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"五件套_{month or '全部'}"

    headers = [
        "日期", "门店ID", "门店名称",
        "walkin直接进店", "异业介绍", "线上", "招聘自带", "存量老客户", "合计进店",
        "触摸产品", "试用体验", "微信添加", "成交组数",
        "成交金额(USD)",
        "提交人", "提交时间",
    ]
    ws.append(headers)
    _header_style(ws, 1, len(headers))

    for r in rows:
        total = r.total_visits
        ws.append([
            r.report_date, r.dealer_id, r.dealer_name,
            r.walkin_visits, r.cross_visits, r.online_visits,
            r.recruit_visits, r.existing_visits, total,
            r.touch_count, r.use_count, r.wechat_add_count, r.deal_count,
            r.deal_amount_yuan,
            r.submitted_by,
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
        ])

    # 自动列宽
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 32)

    fname = f"五件套_{month or '全部'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return _wb_response(wb, fname)


# ── 经销商业绩导出 ─────────────────────────────────────────────────────────────

@router.get("/dealer-sales")
async def export_dealer_sales(
    user: Annotated[User, Depends(require_role("viewer"))],
    session: Annotated[Session, Depends(get_session)],
    month: str = Query(""),
):
    """导出经销商 sell-in/sell-out 业绩 Excel。"""
    stmt = select(DealerSales).order_by(DealerSales.check_date, DealerSales.dealer_name)
    if month and re.fullmatch(r"\d{4}-\d{2}", month):
        stmt = stmt.where(DealerSales.check_date.startswith(month))
    names = visible_dealer_names(user, session)
    if names is not None:
        if not names:
            rows = []
        else:
            rows = session.exec(stmt.where(DealerSales.dealer_name.in_(names))).all()
    else:
        rows = session.exec(stmt).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"业绩_{month or '全部'}"

    headers = ["日期", "经销商", "区域", "国家",
                "Sell-in(万)", "Sell-out(万)", "台量", "同步时间"]
    ws.append(headers)
    _header_style(ws, 1, len(headers))

    for r in rows:
        ws.append([
            r.check_date, r.dealer_name, r.region, r.country,
            r.sell_in_wan, r.sell_out_wan, r.units,
            r.synced_at.strftime("%Y-%m-%d %H:%M") if r.synced_at else "",
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 36)

    fname = f"业绩_{month or '全部'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return _wb_response(wb, fname)
