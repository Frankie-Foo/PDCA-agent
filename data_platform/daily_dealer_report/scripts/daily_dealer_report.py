import argparse
import csv
import json
import os
import sys
import urllib.request
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


REQUIRED_COLUMNS = ["销售日期", "销售员", "实际业绩"]
DEFAULT_GROUP = "经销商数据核对"


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def to_number(value):
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def read_targets(path):
    targets = {}
    if not path or not Path(path).exists():
        return targets
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            sales = row.get("sales", "").strip()
            if not sales:
                continue
            targets[sales] = {
                "department": row.get("department", "").strip() or "经销商",
                "team": row.get("team", "").strip() or "未配置团队",
                "target": to_number(row.get("monthly_target")),
            }
    return targets


def read_pipeline(path, run_date):
    pipeline = defaultdict(float)
    if not path or not Path(path).exists():
        return pipeline
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            if row.get("date", "").strip() != run_date:
                continue
            sales = row.get("sales", "").strip()
            if sales:
                pipeline[sales] += to_number(row.get("pipeline_amount"))
    return pipeline


def inspect_workbook(input_xlsx, sheet_name):
    path = Path(input_xlsx)
    if not path.exists():
        return None, [f"销售明细文件不存在：{input_xlsx}"]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None, [f"工作表不存在：{sheet_name}；可用工作表：{', '.join(wb.sheetnames)}"]
    ws = wb[sheet_name]
    headers = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        return None, [f"缺少必要字段：{', '.join(missing)}"]
    return (wb, ws, headers), []


def build_metrics(input_xlsx, sheet_name, run_date, targets, pipeline):
    inspected, blocking = inspect_workbook(input_xlsx, sheet_name)
    if blocking:
        return [], blocking, []

    _, ws, headers = inspected
    idx = {name: i for i, name in enumerate(headers)}
    run_dt = datetime.strptime(run_date, "%Y-%m-%d")
    month_key = run_dt.strftime("%Y-%m")
    current = defaultdict(float)
    daily_collection = defaultdict(float)
    warnings = []
    negative_rows = 0
    large_rows = 0

    payment_idx = idx.get("付款时间")
    amount_idx = idx["实际业绩"]
    sales_idx = idx["销售员"]
    date_idx = idx["销售日期"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        sales = str(row[sales_idx]).strip() if row[sales_idx] else ""
        if not sales:
            continue
        sales_date = parse_date(row[date_idx])
        amount = to_number(row[amount_idx])
        if amount < 0:
            negative_rows += 1
        if abs(amount) > 5_000_000:
            large_rows += 1
        if sales_date and sales_date.strftime("%Y-%m") == month_key:
            current[sales] += amount
        if payment_idx is not None:
            pay_date = parse_date(row[payment_idx])
            if pay_date and pay_date.strftime("%Y-%m-%d") == run_date:
                daily_collection[sales] += amount

    all_sales = sorted(set(current) | set(targets) | set(daily_collection) | set(pipeline))
    rows = []
    for sales in all_sales:
        info = targets.get(sales, {})
        target = info.get("target", 0.0)
        achievement = current[sales] / target if target else None
        rows.append(
            {
                "department": info.get("department", "经销商"),
                "team": info.get("team", "未配置团队"),
                "sales": sales,
                "current": current[sales],
                "target": target,
                "achievement": achievement,
                "daily_collection": daily_collection[sales],
                "pipeline": pipeline[sales],
            }
        )
        if sales not in targets:
            warnings.append(f"{sales} 未配置团队/目标")
        elif target == 0 and current[sales] != 0:
            warnings.append(f"{sales} 目标为 0 但已有业绩")

    if not rows:
        warnings.append(f"{run_date[:7]} 未读取到当月业绩数据")
    if negative_rows:
        warnings.append(f"发现 {negative_rows} 行负业绩，请核对退款/退货口径")
    if large_rows:
        warnings.append(f"发现 {large_rows} 行超大金额，请核对是否异常")
    if sum(r["daily_collection"] for r in rows) == 0:
        warnings.append(f"{run_date} 当日录单收款为 0 或未读取到付款时间")

    rows.sort(key=lambda r: (r["department"], r["team"], r["sales"]))
    return rows, [], warnings


def wan(value):
    return round(value / 10000, 1)


def pct(value):
    if value is None:
        return ""
    return f"{value * 100:.1f}%"


def make_report(rows, warnings, run_date):
    run_dt = datetime.strptime(run_date, "%Y-%m-%d")
    next_month = datetime(run_dt.year + (run_dt.month // 12), (run_dt.month % 12) + 1, 1)
    month_days = (next_month - datetime(run_dt.year, run_dt.month, 1)).days
    time_progress = run_dt.day / month_days

    total_current = sum(r["current"] for r in rows)
    configured_current = sum(r["current"] for r in rows if r["target"] > 0)
    total_target = sum(r["target"] for r in rows)
    total_collection = sum(r["daily_collection"] for r in rows)
    total_pipeline = sum(r["pipeline"] for r in rows)
    total_rate = configured_current / total_target if total_target else None
    laggards = [r["sales"] for r in rows if r["achievement"] is not None and r["achievement"] < time_progress]
    collectors = [r["sales"] for r in rows if r["daily_collection"] > 0]

    if total_rate is None:
        headline = "目标未配置完整，今日先输出业绩与数据核对提醒。"
    elif total_rate >= time_progress:
        headline = f"整体达成率 {pct(total_rate)}，高于时间进度 {pct(time_progress)}。"
    else:
        headline = f"整体达成率 {pct(total_rate)}，低于时间进度 {pct(time_progress)}，需重点核对落后销售的在途和回款。"

    lines = [
        f"# 海外经销商每日数据汇报 {run_date}",
        "",
        "## 一句话结论",
        f"- {headline}",
        f"- 已配置目标人员当月业绩 {wan(configured_current)} 万，当月目标 {wan(total_target)} 万；全部读取业绩 {wan(total_current)} 万。",
        f"- 当日录单收款 {wan(total_collection)} 万，当日在途业绩 {wan(total_pipeline)} 万。",
    ]
    if laggards:
        lines.append(f"- 达成率低于时间进度：{'、'.join(laggards)}。")
    if collectors:
        lines.append(f"- 今日有录单收款：{'、'.join(collectors)}。")

    lines.extend(
        [
            "",
            "## 明细表",
            "| 部门 | 团队 | 销售 | 当月业绩(万) | 当月目标(万) | 当月达成率 | 当日录单收款(万) | 当日在途业绩(万) |",
            "|---|---|---|---:|---:|---:|---:|---:|",
        ]
    )
    for r in rows:
        lines.append(
            f"| {r['department']} | {r['team']} | {r['sales']} | {wan(r['current'])} | {wan(r['target'])} | {pct(r['achievement'])} | {wan(r['daily_collection'])} | {wan(r['pipeline'])} |"
        )
    lines.append(
        f"| 已配置目标合计 |  |  | {wan(configured_current)} | {wan(total_target)} | {pct(total_rate)} | {wan(total_collection)} | {wan(total_pipeline)} |"
    )
    if configured_current != total_current:
        lines.append(
            f"| 全部读取业绩 |  |  | {wan(total_current)} |  |  | {wan(total_collection)} | {wan(total_pipeline)} |"
        )

    lines.extend(["", "## 数据核对提醒"])
    if warnings:
        lines.extend(f"- {item}" for item in warnings)
    else:
        lines.append("- 暂无阻塞性异常。")

    lines.extend(
        [
            "",
            "## 建议跟进",
            "- 请各组核对达成率低于时间进度的销售是否有高概率在途。",
            "- 请核对当日录单收款为 0 的团队是否存在延迟录入或回款未同步。",
            "- 请数据中台确认未配置团队/目标的销售是否需要补入目标表。",
        ]
    )
    return "\n".join(lines) + "\n"


def post_webhook(webhook_url, group_name, message):
    payload = json.dumps({"group_name": group_name, "text": message}, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        webhook_url,
        data=payload,
        headers={"Content-Type": "application/json; charset=utf-8"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=20) as resp:
        return resp.status, resp.read().decode("utf-8", errors="ignore")


def main():
    parser = argparse.ArgumentParser(description="Generate and optionally push dealer daily report.")
    parser.add_argument("--input-xlsx", required=True)
    parser.add_argument("--sheet", default="26")
    parser.add_argument("--date", required=True)
    parser.add_argument("--targets")
    parser.add_argument("--pipeline")
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--push", action="store_true")
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    report_dir = out_dir / "reports"
    outbox_dir = out_dir / "outbox"
    report_dir.mkdir(parents=True, exist_ok=True)
    outbox_dir.mkdir(parents=True, exist_ok=True)

    targets = read_targets(args.targets)
    pipeline = read_pipeline(args.pipeline, args.date)
    rows, blocking, warnings = build_metrics(args.input_xlsx, args.sheet, args.date, targets, pipeline)
    if blocking:
        message = f"# 海外经销商每日数据汇报 {args.date}\n\n## 数据阻塞\n" + "\n".join(f"- {item}" for item in blocking) + "\n"
    else:
        message = make_report(rows, warnings, args.date)

    report_path = report_dir / f"{args.date}_dealer_daily_report.md"
    report_path.write_text(message, encoding="utf-8")

    group_name = os.environ.get("DEALER_IM_GROUP", DEFAULT_GROUP)
    payload = {"group_name": group_name, "text": message, "report_path": str(report_path)}
    payload_path = outbox_dir / f"{args.date}_im_payload.json"
    payload_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    sent = False
    if args.push:
        webhook = os.environ.get("DEALER_IM_WEBHOOK_URL", "").strip()
        if webhook:
            try:
                status, body = post_webhook(webhook, group_name, message)
                sent = 200 <= status < 300
                print(f"IM push status={status} response={body[:200]}")
            except Exception as exc:
                print(f"IM push failed: {exc}", file=sys.stderr)
        else:
            print("DEALER_IM_WEBHOOK_URL is not set. Payload written to outbox.", file=sys.stderr)

    print(f"Report: {report_path}")
    print(f"IM payload: {payload_path}")
    print(f"IM group: {group_name}")
    print(f"Sent: {sent}")


if __name__ == "__main__":
    main()
