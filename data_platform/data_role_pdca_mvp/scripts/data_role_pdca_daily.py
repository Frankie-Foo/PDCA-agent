import argparse
import csv
import json
import os
import re
import urllib.request
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill


DONE_STATUSES = {"done", "completed", "已完成", "完成"}


def read_csv(path):
    if not path or not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_aliases(path):
    aliases = {}
    for row in read_csv(path):
        raw = text(row.get("raw_sales"))
        canonical = text(row.get("canonical_sales"))
        if raw and canonical:
            aliases[raw] = canonical
    return aliases


def write(path, text):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


def write_csv(path, rows, fieldnames):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def load_json(path):
    if not path or not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8-sig"))


def load_optional_json(path):
    if not path or not path.exists():
        return {}
    try:
        return load_json(path)
    except json.JSONDecodeError:
        return {}


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    return None


def money(value):
    if value is None or value == "":
        return 0.0
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def to_wan(value):
    return round(money(value) / 10000, 2)


def text(value):
    return "" if value is None else str(value).strip()


def first_present(row, names):
    for name in names:
        if name in row and row[name] not in (None, ""):
            return row[name]
    return ""


def normalize_sales_record(row):
    return {
        "date": first_present(row, ["销售日期", "sale_date", "date"]),
        "salesperson": text(first_present(row, ["销售员", "salesperson", "owner"])),
        "customer": text(first_present(row, ["客户名称", "customer_name", "customer", "经销商", "partner_name"])),
        "product": text(first_present(row, ["存货名称", "product_name", "产品名称", "product", "商品名称"])),
        "product_code": text(first_present(row, ["产品编码", "product_code", "sku"])),
        "quantity": money(first_present(row, ["实际数量", "数量", "quantity", "qty"])),
        "performance": money(first_present(row, ["实际业绩", "performance", "performance_cny", "实际成交金额(CNY)"])),
    }


def read_sales_xlsx(path, sheet_name=None):
    if not path or not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    records = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        raw = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}
        records.append(normalize_sales_record(raw))
    return records


def read_sales_json(path):
    payload = load_json(path)
    if isinstance(payload, list):
        rows = payload
    elif isinstance(payload, dict):
        rows = payload.get("rows") or payload.get("data") or []
        if not rows and isinstance(payload.get("result"), dict):
            rows = payload["result"].get("rows") or []
        if not rows and isinstance(payload.get("result"), list):
            rows = payload["result"]
        if not rows and isinstance(payload.get("ai"), dict):
            ai_result = payload["ai"].get("result")
            if isinstance(ai_result, dict):
                rows = ai_result.get("rows") or []
    else:
        rows = []
    return [normalize_sales_record(row) for row in rows]


def read_sales_summary_json(path, aliases=None):
    aliases = aliases or {}
    payload = load_json(path)
    result = None
    if isinstance(payload, dict):
        if isinstance(payload.get("execution"), dict):
            result = payload["execution"].get("result")
        if result is None:
            result = payload.get("result")
        if result is None and isinstance(payload.get("ai"), dict):
            result = payload["ai"].get("result")
    if not isinstance(result, dict) or not result.get("summary_mode"):
        return None

    def convert(rows, key_name):
        converted = []
        for row in rows or []:
            dimension = text(row.get(key_name)) or "(空)"
            if key_name == "salesperson":
                dimension = aliases.get(dimension, dimension)
            converted.append({
                "dimension": dimension,
                "performance": money(row.get("performance")),
                "quantity": money(row.get("quantity")),
                "rows": int(money(row.get("line_count"))),
            })
        return converted

    trend_sales = [
        {
            "day": text(row.get("day_label")),
            "dimension": aliases.get(text(row.get("salesperson")), text(row.get("salesperson"))),
            "performance": money(row.get("performance")),
            "quantity": money(row.get("quantity")),
        }
        for row in result.get("daily_trend_by_salesperson") or []
    ]
    daily_salesperson = convert(result.get("daily_salesperson_summary"), "salesperson")
    daily_fallback_day = ""
    if not daily_salesperson and trend_sales:
        latest_day = max(row.get("day", "") for row in trend_sales)
        daily_fallback_day = latest_day
        daily_salesperson = sorted(
            [
                {
                    "dimension": row.get("dimension", ""),
                    "performance": row.get("performance", 0),
                    "quantity": row.get("quantity", 0),
                    "rows": 0,
                }
                for row in trend_sales
                if row.get("day") == latest_day
            ],
            key=lambda item: item["performance"],
            reverse=True,
        )

    return {
        "source": result.get("source", "VPS/Odoo"),
        "table": result.get("table", ""),
        "salesperson": convert(result.get("salesperson_summary"), "salesperson"),
        "product": convert(result.get("product_summary"), "product_name"),
        "customer": convert(result.get("customer_summary"), "partner_name"),
        "daily_salesperson": daily_salesperson,
        "week_salesperson": convert(result.get("week_salesperson_summary"), "salesperson"),
        "daily_team": convert(result.get("daily_team_summary"), "team"),
        "week_team": convert(result.get("week_team_summary"), "team"),
        "daily_trend_by_salesperson": trend_sales,
        "daily_fallback_day": daily_fallback_day,
    }


def filter_month_to_date(records, date_text):
    run_date = datetime.strptime(date_text, "%Y-%m-%d")
    month_key = run_date.strftime("%Y-%m")
    filtered = []
    for record in records:
        dt = parse_date(record.get("date"))
        if not dt:
            continue
        if dt.strftime("%Y-%m") == month_key and dt <= run_date:
            filtered.append(record)
    return filtered


def aggregate(records, dimension):
    result = defaultdict(lambda: {"dimension": "", "performance": 0.0, "quantity": 0.0, "rows": 0})
    for record in records:
        key = record.get(dimension) or "(空)"
        result[key]["dimension"] = key
        result[key]["performance"] += record.get("performance", 0.0)
        result[key]["quantity"] += record.get("quantity", 0.0)
        result[key]["rows"] += 1
    return sorted(result.values(), key=lambda item: item["performance"], reverse=True)


def add_table_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    headers = ["维度", "业绩", "数量", "明细行数"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    for row in rows:
        ws.append([row["dimension"], row["performance"], row["quantity"], row["rows"]])
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = '#,##0.00'
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12

    if len(rows) >= 1:
        chart = BarChart()
        chart.title = f"{title} TOP 10 业绩"
        chart.y_axis.title = "业绩"
        chart.x_axis.title = title
        max_row = min(len(rows) + 1, 11)
        data = Reference(ws, min_col=2, min_row=1, max_row=max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, "F2")
    return ws


def write_summary_workbook(path, date_text, records):
    wb = Workbook()
    overview = wb.active
    overview.title = "说明"
    overview.append(["项目", "值"])
    overview.append(["日期", date_text])
    overview.append(["明细行数", len(records)])
    overview.append(["业绩合计", sum(r.get("performance", 0.0) for r in records)])
    overview.append(["数量合计", sum(r.get("quantity", 0.0) for r in records)])
    overview["A1"].font = Font(bold=True)
    overview["B1"].font = Font(bold=True)
    overview.column_dimensions["A"].width = 18
    overview.column_dimensions["B"].width = 30

    summaries = {
        "销售员汇总": aggregate(records, "salesperson"),
        "产品汇总": aggregate(records, "product"),
        "客户汇总": aggregate(records, "customer"),
    }
    for sheet, rows in summaries.items():
        add_table_sheet(wb, sheet, rows)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return summaries


def write_summary_workbook_from_vps(path, date_text, summary):
    wb = Workbook()
    overview = wb.active
    overview.title = "说明"
    overview.append(["项目", "值"])
    overview.append(["日期", date_text])
    overview.append(["数据来源", summary.get("source", "VPS/Odoo")])
    overview.append(["数据表", summary.get("table", "")])
    total_perf = sum(r["performance"] for r in summary.get("salesperson", []))
    total_qty = sum(r["quantity"] for r in summary.get("salesperson", []))
    overview.append(["业绩合计", total_perf])
    overview.append(["数量合计", total_qty])
    for cell in overview[1]:
        cell.font = Font(bold=True)
    overview.column_dimensions["A"].width = 18
    overview.column_dimensions["B"].width = 40

    add_table_sheet(wb, "销售员汇总", summary.get("salesperson", []))
    add_table_sheet(wb, "产品汇总", summary.get("product", []))
    add_table_sheet(wb, "客户汇总", summary.get("customer", []))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return {
        "销售员汇总": summary.get("salesperson", []),
        "产品汇总": summary.get("product", []),
        "客户汇总": summary.get("customer", []),
    }


def top_table(rows):
    if not rows:
        return "<p>暂无数据</p>"
    max_perf = max(abs(row.get("performance", 0.0)) for row in rows[:10]) or 1
    parts = ["<table><thead><tr><th>维度</th><th>业绩</th><th>数量</th><th>图示</th></tr></thead><tbody>"]
    for row in rows[:10]:
        perf = row.get("performance", 0.0)
        width = max(2, int(abs(perf) / max_perf * 100))
        parts.append(
            f"<tr><td>{row.get('dimension','')}</td><td>{perf:,.2f}</td><td>{row.get('quantity',0):,.2f}</td><td><div class='bar' style='width:{width}%'></div></td></tr>"
        )
    parts.append("</tbody></table>")
    return "\n".join(parts)


def norm_name(value):
    return re.sub(r"[\s\-'\"“”\.，,()（）]+", "", str(value or "")).upper()


def dashboard_store_data(workspace, chart_data):
    customer_rows = chart_data.get("customer_top", [])
    customer_detail = []
    for row in customer_rows[:20]:
        perf = to_wan(row.get("performance", 0))
        qty = int(money(row.get("quantity", 0)))
        customer_detail.append({
            "name": row.get("dimension", ""),
            "performance": perf,
            "pickupAmount": perf,
            "pickupQty": qty,
            "orderPerformance": perf,
            "vpsUsage": None,
            "inboundQty": None,
            "activationQty": None,
            "walkinQty": None,
            "inTransitPerformance": None,
        })

    perf_map = {norm_name(row.get("dimension")): row for row in customer_rows}
    dealers_cfg = load_optional_json(workspace / "config" / "dealers.json")
    dealers = dealers_cfg.get("dealers", []) if isinstance(dealers_cfg, dict) else []
    region_map = {}
    for dealer in dealers:
        region = dealer.get("region") or "其他"
        country = dealer.get("country") or "未分组"
        row = perf_map.get(norm_name(dealer.get("name")), {})
        perf = to_wan(row.get("performance", 0))
        qty = int(money(row.get("quantity", 0)))
        entry = {
            "name": dealer.get("name", ""),
            "nickname": dealer.get("nickname", ""),
            "salesperson": dealer.get("salesperson", ""),
            "perf": perf,
            "qty": qty,
            "pickupAmount": perf if row else None,
            "pickupQty": qty if row else None,
            "vpsUsage": None,
            "inboundQty": None,
            "activationQty": None,
            "walkinQty": None,
            "orderPerformance": perf if row else None,
            "inTransitPerformance": None,
        }
        region_map.setdefault(region, {}).setdefault(country, []).append(entry)

    dealer_region = []
    for region, countries in region_map.items():
        country_list = []
        region_perf = 0
        region_qty = 0
        for country, entries in countries.items():
            entries = sorted(entries, key=lambda item: item.get("perf", 0), reverse=True)
            country_perf = round(sum(item.get("perf", 0) for item in entries), 2)
            country_qty = sum(item.get("qty", 0) for item in entries)
            region_perf += country_perf
            region_qty += country_qty
            country_list.append({
                "country": country,
                "perf": country_perf,
                "qty": country_qty,
                "dealers": entries,
            })
        dealer_region.append({
            "region": region,
            "perf": round(region_perf, 2),
            "qty": region_qty,
            "countries": country_list,
        })
    dealer_region.sort(key=lambda item: item.get("perf", 0), reverse=True)
    return customer_detail, dealer_region


def inject_dashboard_store_data(html, workspace, chart_data):
    customer_detail, dealer_region = dashboard_store_data(workspace, chart_data)
    customer_rank = [
        {"n": row.get("name", ""), "perf": row.get("performance", 0), "qty": row.get("pickupQty", 0)}
        for row in customer_detail[:10]
    ]
    replacements = {
        "CUSTOMER_RANK": customer_rank,
        "CUSTOMER_DETAIL": customer_detail,
        "DEALER_REGION": dealer_region,
    }
    for const_name, payload in replacements.items():
        js = f"const {const_name}={json.dumps(payload, ensure_ascii=False)};"
        pattern = rf"const {const_name}=.*?;\n"
        if re.search(pattern, html, flags=re.S):
            html = re.sub(pattern, js + "\n", html, count=1, flags=re.S)
        else:
            marker = "// ─── Simple deferred render ───"
            html = html.replace(marker, js + "\n" + marker)
    return html


PERSON_NODE_MAP = {
    "杨晶晶": "p-yangjj",
    "王宇彤": "p-wangyt",
    "何海文": "p-hehai",
    "于冰": "p-yubing",
    "Lydia": "p-lydia",
    "Lina": "p-deh",
    "尤文静": "p-youwj",
}

TEAM_MEMBERS = {
    "tl-yang": ("杨晶晶组", ["杨晶晶", "王宇彤", "何海文"]),
    "tl-yubing": ("于冰组", ["于冰", "Lydia"]),
    "tl-deh": ("Lina组", ["Lina", "尤文静"]),
}


def product_pack(product_rows):
    top_rows = product_rows[:6] or [{"dimension": "暂无", "performance": 0, "quantity": 0}]
    names = [row.get("dimension", "")[:28] for row in top_rows]
    qty = [round(money(row.get("quantity")), 2) for row in top_rows]
    amt = [to_wan(row.get("performance")) for row in top_rows]
    return {
        "phoneBar": {"cats": [name[:14] for name in names], "qty": qty, "amt": amt},
        "matPie": [{"name": name[:18], "value": qty[index]} for index, name in enumerate(names)],
        "phoneTop": names,
        "phoneTopVals": qty,
        "iotTop": ["暂未接入VPS"],
        "iotTopVals": [0],
        "iotCombo": {"cats": ["暂未接入VPS"], "qty": [0], "amt": [0]},
        "iotPie": [{"name": "暂未接入VPS", "value": 0}],
        "stackPhone": {"cats": names, "series": [{"name": "销量", "data": qty}]},
        "stackIot": {"cats": ["暂未接入VPS"], "series": [{"name": "销量", "data": [0]}]},
        "phoneIotTrend": {"phone": [sum(qty)], "iot": [0]},
    }


def period_payload(name, daily_perf, week_perf, month_perf, month_qty, products, rank_rows):
    common = {
        "todayVal": daily_perf,
        "todayTgt": 0,
        "todayPct": 0,
        "todayHb": 0,
        "todayTb": 0,
        "yesterdayVal": 0,
        "yesterdayTgt": 0,
        "yesterdayPct": 0,
        "yesterdayHb": 0,
        "yesterdayTb": 0,
        "weekVal": week_perf,
        "weekTgt": 0,
        "weekAr": 0,
        "weekTp": 0,
        "monthVal": month_perf,
        "monthTgt": 0,
        "monthHb": 0,
        "monthTb": 0,
        "monthAr": 0,
        "monthTp": 0,
        "rkW": rank_rows,
        "rkM": rank_rows,
        "todayRk": [],
        "yesterdayRk": [],
        "trend": [0, daily_perf, week_perf, month_perf],
        "p7": {"dates": [], "t1": [], "t2": [], "t1p": [], "t2p": []},
        "trendSeries": [{"name": name, "data": [0, daily_perf, week_perf, month_perf], "color": "#3b6ef8"}],
        "openRate": [[name], [100 if month_perf else 0]],
        "rcData": [[name], [month_perf]],
        "distData": [[1 if month_perf else 0]],
        "top20Pie": [{"name": name, "value": month_perf}],
        "orgMonth": {"okrTgt": 0, "actual": month_perf, "ar": 0, "tp": 0, "hb": 0, "tb": 0},
        "dailyNote": "",
        "phone": products.get("phoneBar"),
        "iot": products.get("iotCombo"),
        **products,
    }
    daily = dict(common)
    weekly = dict(common)
    monthly = dict(common)
    daily["todayVal"] = daily_perf
    weekly["todayVal"] = week_perf
    monthly["todayVal"] = month_perf
    return {"daily": daily, "weekly": weekly, "monthly": monthly}


def inject_dashboard_org_data(html, chart_data):
    products = product_pack(chart_data.get("product_top", []))
    sales_map = {row.get("dimension"): row for row in chart_data.get("salesperson_top", [])}
    daily_sales_map = {row.get("dimension"): row for row in chart_data.get("salesperson_daily", [])}
    week_sales_map = {row.get("dimension"): row for row in chart_data.get("salesperson_week", [])}
    daily_note = ""
    if chart_data.get("daily_fallback_day"):
        daily_note = f"当天无业绩，显示最近有数据日 {chart_data.get('daily_fallback_day')}。"
    rank_rows = [
        {"n": row.get("dimension", ""), "v": to_wan(row.get("performance", 0)), "q": int(money(row.get("quantity", 0)))}
        for row in chart_data.get("salesperson_top", [])[:10]
    ]

    persons = {}
    for name, node_id in PERSON_NODE_MAP.items():
        row = sales_map.get(name, {})
        daily_row = daily_sales_map.get(name, {})
        week_row = week_sales_map.get(name, {})
        persons[node_id] = period_payload(
            name,
            to_wan(daily_row.get("performance", 0)),
            to_wan(week_row.get("performance", 0)),
            to_wan(row.get("performance", 0)),
            int(money(row.get("quantity", 0))),
            products,
            rank_rows,
        )
        persons[node_id]["daily"]["dailyNote"] = daily_note

    teams = {}
    for team_id, (team_name, members) in TEAM_MEMBERS.items():
        total_perf = round(sum(to_wan(sales_map.get(member, {}).get("performance", 0)) for member in members), 2)
        total_qty = sum(int(money(sales_map.get(member, {}).get("quantity", 0))) for member in members)
        daily_perf = round(sum(to_wan(daily_sales_map.get(member, {}).get("performance", 0)) for member in members), 2)
        week_perf = round(sum(to_wan(week_sales_map.get(member, {}).get("performance", 0)) for member in members), 2)
        payload = period_payload(team_name, daily_perf, week_perf, total_perf, total_qty, products, rank_rows)
        payload["daily"]["dailyNote"] = daily_note
        for period in ("daily", "weekly", "monthly"):
            payload[period]["trendSeries"] = [
                {
                    "name": member,
                    "data": [
                        0,
                        to_wan(daily_sales_map.get(member, {}).get("performance", 0)),
                        to_wan(week_sales_map.get(member, {}).get("performance", 0)),
                        to_wan(sales_map.get(member, {}).get("performance", 0)),
                    ],
                    "color": ["#3b6ef8", "#22c55e", "#f97316"][idx % 3],
                }
                for idx, member in enumerate(members)
            ]
            period_map = daily_sales_map if period == "daily" else week_sales_map if period == "weekly" else sales_map
            payload[period]["openRate"] = [members, [100 if period_map.get(member, {}).get("performance") else 0 for member in members]]
            payload[period]["rcData"] = [members, [to_wan(period_map.get(member, {}).get("performance", 0)) for member in members]]
            payload[period]["top20Pie"] = [{"name": member, "value": to_wan(sales_map.get(member, {}).get("performance", 0))} for member in members]
            payload[period]["distData"] = [[len([member for member in members if period_map.get(member, {}).get("performance")])]]
        teams[team_id] = payload

    total_perf = round(sum(to_wan(row.get("performance", 0)) for row in chart_data.get("salesperson_top", [])), 2)
    total_qty = sum(int(money(row.get("quantity", 0))) for row in chart_data.get("salesperson_top", []))
    daily_total = round(sum(to_wan(row.get("performance", 0)) for row in chart_data.get("salesperson_daily", [])), 2)
    week_total = round(sum(to_wan(row.get("performance", 0)) for row in chart_data.get("salesperson_week", [])), 2)
    director_payload = period_payload("经销商", daily_total, week_total, total_perf, total_qty, products, rank_rows)
    director_payload["daily"]["dailyNote"] = daily_note
    for period in ("daily", "weekly", "monthly"):
        source_rows = chart_data.get("salesperson_daily", []) if period == "daily" else chart_data.get("salesperson_week", []) if period == "weekly" else chart_data.get("salesperson_top", [])
        names = [row.get("dimension", "") for row in source_rows]
        vals = [to_wan(row.get("performance", 0)) for row in source_rows]
        director_payload[period]["trendSeries"] = [{"name": "经销商", "data": [0, daily_total, week_total, total_perf], "color": "#3b6ef8"}]
        director_payload[period]["openRate"] = [names, [100 if val else 0 for val in vals]]
        director_payload[period]["rcData"] = [names, vals]
        director_payload[period]["top20Pie"] = [{"name": name, "value": vals[idx]} for idx, name in enumerate(names)]
        director_payload[period]["distData"] = [[len([val for val in vals if val > 0])]]
    director = {"dir-dealer": director_payload}

    replacements = {
        "PERSONS": persons,
        "TL": teams,
        "DIR": director,
    }
    for const_name, payload in replacements.items():
        js = f"const {const_name}={json.dumps(payload, ensure_ascii=False)};"
        pattern = rf"const {const_name}=.*?;\n"
        html = re.sub(pattern, js + "\n", html, count=1, flags=re.S)
    return html


def write_dashboard(path, workspace, date_text, pending_todos, logistics_rows, chart_data):
    template_path = workspace / "templates" / "dashboard_template.html"
    template = template_path.read_text(encoding="utf-8")
    total_performance = sum(row.get("performance", 0.0) for row in chart_data.get("salesperson_top", []))
    attention = len([row for row in logistics_rows if row.get("judgement") in {"异常", "待关注"}])
    html = template
    html = html.replace("{{date}}", date_text)
    html = html.replace("{{todo_count}}", str(len(pending_todos)))
    html = html.replace("{{logistics_attention_count}}", str(attention))
    html = html.replace("{{total_performance}}", f"{total_performance:,.2f}")
    html = html.replace("{{salesperson_table}}", top_table(chart_data.get("salesperson_top", [])))
    html = html.replace("{{product_table}}", top_table(chart_data.get("product_top", [])))
    html = html.replace("{{customer_table}}", top_table(chart_data.get("customer_top", [])))
    html = inject_dashboard_org_data(html, chart_data)
    html = inject_dashboard_store_data(html, workspace, chart_data)
    write(path, html)


def build_todo_reminder(workspace, date_text):
    today_path = workspace / "inputs" / "todos" / f"{date_text}_todos.csv"
    yesterday = (datetime.strptime(date_text, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    yesterday_path = workspace / "inputs" / "todos" / f"{yesterday}_todos.csv"

    rows = []
    rows.extend(read_csv(yesterday_path))
    rows.extend(read_csv(today_path))

    pending = [r for r in rows if (r.get("status") or "").lower() not in DONE_STATUSES]
    high = [r for r in pending if (r.get("priority") or "").upper() == "HIGH"]
    medium = [r for r in pending if (r.get("priority") or "").upper() == "MEDIUM"]
    other = [r for r in pending if r not in high and r not in medium]

    lines = [f"# 今日代办提醒 {date_text}", ""]
    lines.append("## 今日必须完成")
    lines.extend(f"- [{r.get('source')}] {r.get('title')}（截止：{r.get('due_date') or date_text}）" for r in high)
    if not high:
        lines.append("- 暂无")

    lines.extend(["", "## 今日应完成"])
    lines.extend(f"- [{r.get('source')}] {r.get('title')}" for r in medium)
    if not medium:
        lines.append("- 暂无")

    lines.extend(["", "## 可延后/待确认"])
    lines.extend(f"- [{r.get('source')}] {r.get('title')}" for r in other)
    if not other:
        lines.append("- 暂无")

    return "\n".join(lines), pending


def build_data_summary(workspace, date_text, day_out, sales_xlsx=None, sales_json=None, sales_sheet=None):
    records = []
    source_note = "未提供销售明细，已生成汇总任务框架"
    vps_summary = None
    if sales_json:
        aliases = read_aliases(workspace / "config" / "sales_aliases.csv")
        vps_summary = read_sales_summary_json(Path(sales_json), aliases)
        if not vps_summary:
            records = read_sales_json(Path(sales_json))
        source_note = f"VPS/Odoo JSON：{sales_json}"
    elif sales_xlsx:
        records = read_sales_xlsx(Path(sales_xlsx), sales_sheet)
        source_note = f"离线演示 Excel：{sales_xlsx}"
    records = filter_month_to_date(records, date_text)

    workbook_path = day_out / f"{date_text}_data_summary.xlsx"
    if vps_summary:
        summaries = write_summary_workbook_from_vps(workbook_path, date_text, vps_summary)
        total_perf = sum(r["performance"] for r in summaries["销售员汇总"])
        total_qty = sum(r["quantity"] for r in summaries["销售员汇总"])
        record_count = sum(r["rows"] for r in summaries["销售员汇总"])
    else:
        summaries = write_summary_workbook(workbook_path, date_text, records)
        total_perf = sum(r.get("performance", 0.0) for r in records)
        total_qty = sum(r.get("quantity", 0.0) for r in records)
        record_count = len(records)

    lines = [
        f"# 数据汇总报告 {date_text}",
        "",
        f"- 数据来源：{source_note}",
        f"- 明细行数：{record_count}",
        f"- 业绩合计：{total_perf:,.2f}",
        f"- 数量合计：{total_qty:,.2f}",
        f"- Excel 输出：`{workbook_path}`",
        "",
        "## TOP 销售员",
    ]
    for row in summaries["销售员汇总"][:5]:
        lines.append(f"- {row['dimension']}：业绩 {row['performance']:,.2f}，数量 {row['quantity']:,.2f}")
    lines.extend(["", "## TOP 产品"])
    for row in summaries["产品汇总"][:5]:
        lines.append(f"- {row['dimension']}：业绩 {row['performance']:,.2f}，数量 {row['quantity']:,.2f}")
    lines.extend(["", "## TOP 客户"])
    for row in summaries["客户汇总"][:5]:
        lines.append(f"- {row['dimension']}：业绩 {row['performance']:,.2f}，数量 {row['quantity']:,.2f}")
    if not records and not vps_summary:
        lines.extend(["", "## 数据阻塞", "- 未读取到 VPS/Odoo 正式业绩数据。请先运行 `scripts\\pull_vps_sales_data.ps1`。"])

    chart_data = {
        "date": date_text,
        "source": source_note,
        "workbook": str(workbook_path),
        "salesperson_top": summaries["销售员汇总"][:10],
        "product_top": summaries["产品汇总"][:10],
        "customer_top": summaries["客户汇总"][:10],
        "salesperson_daily": (vps_summary or {}).get("daily_salesperson", [])[:10],
        "salesperson_week": (vps_summary or {}).get("week_salesperson", [])[:10],
        "team_daily": (vps_summary or {}).get("daily_team", [])[:10],
        "team_week": (vps_summary or {}).get("week_team", [])[:10],
        "salesperson_trend": (vps_summary or {}).get("daily_trend_by_salesperson", []),
        "daily_fallback_day": (vps_summary or {}).get("daily_fallback_day", ""),
    }
    return "\n".join(lines), chart_data, workbook_path


def logistics_url(carriers, carrier, tracking_number):
    info = carriers.get(carrier) or carriers.get(carrier.upper()) or {}
    template = info.get("tracking_url", "")
    return template.replace("{tracking_number}", tracking_number) if template else ""


def judge_logistics(row, settings, date_text):
    status = text(row.get("current_status") or row.get("status"))
    status_lower = status.lower()
    for keyword in settings.get("logistics", {}).get("abnormal_keywords", []):
        if keyword.lower() in status_lower:
            return "异常", f"状态包含异常关键词：{keyword}"
    for keyword in settings.get("logistics", {}).get("normal_keywords", []):
        if keyword.lower() in status_lower:
            return "正常", f"状态包含正常关键词：{keyword}"
    ship_date = parse_date(row.get("ship_date"))
    run_date = datetime.strptime(date_text, "%Y-%m-%d")
    if ship_date and (run_date - ship_date).days >= 7 and (row.get("expected_status") or "") != "delivered":
        return "待关注", "发货超过 7 天，且未标记签收"
    return "待核查", "未接入承运商 API 或未填写当前状态"


def build_logistics_report(workspace, date_text, day_out, logistics_csv=None):
    carriers = load_json(workspace / "config" / "carriers.json")
    settings = load_json(workspace / "config" / "settings.json")
    path = Path(logistics_csv) if logistics_csv else workspace / "inputs" / "logistics" / f"{date_text}_tracking.csv"
    rows = read_csv(path)
    output_rows = []

    lines = [f"# 物流核查报告 {date_text}", "", "## 结论"]
    if not rows:
        lines.append("- 今日没有物流单号输入。")
    else:
        lines.append(f"- 读取到 {len(rows)} 个物流单号。已按当前状态/发货天数做初步判断。")

    lines.extend(["", "## 核查明细", "| 单号 | 承运商 | 客户 | 销售 | 发货日期 | 当前状态 | 查询链接 | 判断 | 原因 |", "|---|---|---|---|---|---|---|---|---|"])
    for row in rows:
        tracking_number = row.get("tracking_number", "").strip()
        carrier = row.get("carrier", "").strip()
        url = logistics_url(carriers, carrier, tracking_number)
        judgement, reason = judge_logistics(row, settings, date_text)
        output = dict(row)
        output.update({"tracking_url": url, "judgement": judgement, "reason": reason})
        output_rows.append(output)
        lines.append(
            f"| {tracking_number} | {carrier} | {row.get('customer','')} | {row.get('salesperson','')} | {row.get('ship_date','')} | {row.get('current_status','')} | {url} | {judgement} | {reason} |"
        )

    if output_rows:
        fieldnames = list(output_rows[0].keys())
        write_csv(day_out / f"{date_text}_logistics_results.csv", output_rows, fieldnames)
    return "\n".join(lines), output_rows


def ensure_questionnaire(workspace, date_text):
    template = (workspace / "templates" / "daily_questionnaire.md").read_text(encoding="utf-8")
    text_value = template.replace("YYYY-MM-DD", date_text)
    path = workspace / "inputs" / "questionnaires" / f"{date_text}_questionnaire.md"
    if not path.exists():
        write(path, text_value)
    return path


def parse_questionnaire(path):
    if not path.exists():
        return {}
    sections = {}
    current = None
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        if line.startswith("## "):
            current = line.lstrip("#").strip()
            sections[current] = []
        elif current:
            stripped = line.strip()
            if stripped and stripped != "-":
                sections[current].append(stripped)
    return sections


def build_next_todos(workspace, date_text, pending_todos, questionnaire):
    next_day = (datetime.strptime(date_text, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    rows = []
    for item in pending_todos:
        rows.append({
            "date": next_day,
            "source": "rollover",
            "title": item.get("title", ""),
            "priority": item.get("priority", "MEDIUM"),
            "status": "pending",
            "owner": item.get("owner", "frank"),
            "due_date": next_day,
            "notes": f"from {date_text}",
        })
    for section, values in questionnaire.items():
        if "明天要完成" in section:
            for value in values:
                rows.append({
                    "date": next_day,
                    "source": "questionnaire",
                    "title": value.lstrip("- ").strip(),
                    "priority": "MEDIUM",
                    "status": "pending",
                    "owner": "frank",
                    "due_date": next_day,
                    "notes": f"from {date_text} questionnaire",
                })
    if rows:
        write_csv(workspace / "inputs" / "todos" / f"{next_day}_todos.csv", rows, ["date", "source", "title", "priority", "status", "owner", "due_date", "notes"])
    return rows


def build_pdca_check(date_text, pending_todos, logistics_rows, questionnaire_path, questionnaire, workbook_path, next_todos):
    unfinished = [r for r in pending_todos if (r.get("status") or "").lower() not in DONE_STATUSES]
    abnormal_logistics = [r for r in logistics_rows if r.get("judgement") in {"异常", "待关注"}]
    completed_answers = questionnaire.get("1. 今天完成了什么？", [])
    tomorrow_answers = questionnaire.get("2. 明天要完成什么？", [])

    lines = [f"# 数据岗位 PDCA 日结 {date_text}", "", "## Plan", f"- 今日代办 {len(pending_todos)} 项，其中未完成 {len(unfinished)} 项。"]
    lines.extend(["", "## Do", f"- 已生成数据汇总 Excel：`{workbook_path}`", f"- 已读取物流单号 {len(logistics_rows)} 个。", f"- 每日问卷：`{questionnaire_path}`"])
    if completed_answers:
        lines.append(f"- 问卷记录今日完成 {len(completed_answers)} 项。")
    else:
        lines.append("- 问卷尚未填写今日完成事项。")

    lines.extend(["", "## Check"])
    if unfinished:
        lines.extend(f"- 未完成：{r.get('title')}" for r in unfinished)
    else:
        lines.append("- 今日代办均已完成。")
    if abnormal_logistics:
        lines.extend(f"- 物流需关注：{r.get('tracking_number')}（{r.get('judgement')}：{r.get('reason')}）" for r in abnormal_logistics)
    if not completed_answers:
        lines.append("- 晚间问卷未完整填写，PDCA 闭环不完整。")

    lines.extend(["", "## Act"])
    if next_todos:
        lines.append(f"- 已滚动生成明日待办 {len(next_todos)} 项。")
    if tomorrow_answers:
        lines.extend(f"- 明日计划：{item}" for item in tomorrow_answers)
    lines.append("- 明早继续拉取未完成、问卷明日计划和上级交办。")
    return "\n".join(lines)


def push_webhook(webhook_url, message):
    payload = json.dumps({"text": message}, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(webhook_url, data=payload, headers={"Content-Type": "application/json; charset=utf-8"}, method="POST")
    with urllib.request.urlopen(req, timeout=20) as resp:
        return resp.status, resp.read().decode("utf-8", errors="ignore")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", required=True)
    parser.add_argument("--workspace", required=True)
    parser.add_argument("--sales-xlsx")
    parser.add_argument("--sales-sheet")
    parser.add_argument("--sales-json")
    parser.add_argument("--start-date")
    parser.add_argument("--allow-excel-demo", action="store_true")
    parser.add_argument("--logistics-csv")
    parser.add_argument("--push", action="store_true")
    args = parser.parse_args()

    workspace = Path(args.workspace)
    date_text = args.date
    day_out = workspace / "outputs" / date_text

    todo_text, pending = build_todo_reminder(workspace, date_text)
    if args.sales_xlsx and not args.allow_excel_demo and not args.sales_json:
        raise SystemExit("正式业绩数据必须来自 VPS/Odoo JSON。Excel 仅允许加 --allow-excel-demo 用于离线演示。")

    data_text, chart_data, workbook_path = build_data_summary(workspace, date_text, day_out, args.sales_xlsx, args.sales_json, args.sales_sheet)
    logistics_text, logistics_rows = build_logistics_report(workspace, date_text, day_out, args.logistics_csv)
    questionnaire_path = ensure_questionnaire(workspace, date_text)
    questionnaire = parse_questionnaire(questionnaire_path)
    next_todos = build_next_todos(workspace, date_text, pending, questionnaire)
    pdca_text = build_pdca_check(date_text, pending, logistics_rows, questionnaire_path, questionnaire, workbook_path, next_todos)

    write(day_out / "todo_reminder.md", todo_text)
    write(day_out / "data_summary_report.md", data_text)
    write(day_out / "logistics_check_report.md", logistics_text)
    write(day_out / "pdca_daily_check.md", pdca_text)
    write(day_out / "chart_data.json", json.dumps(chart_data, ensure_ascii=False, indent=2))
    write_dashboard(day_out / "dashboard.html", workspace, date_text, pending, logistics_rows, chart_data)

    im_message = "\n\n".join([todo_text, data_text, logistics_text, pdca_text])
    outbox = workspace / "outbox" / f"{date_text}_im_message.md"
    write(outbox, im_message)

    sent = False
    if args.push:
        webhook = os.environ.get("DATA_PDCA_IM_WEBHOOK_URL", "").strip()
        if webhook:
            status, body = push_webhook(webhook, im_message)
            sent = 200 <= status < 300
            print(f"IM push status={status} response={body[:200]}")
        else:
            print("DATA_PDCA_IM_WEBHOOK_URL is not set. IM message written to outbox.")

    print(f"Generated outputs: {day_out}")
    print(f"Summary workbook: {workbook_path}")
    print(f"Questionnaire: {questionnaire_path}")
    print(f"IM outbox: {outbox}")
    print(f"Sent: {sent}")


if __name__ == "__main__":
    main()
