import csv
import html
import json
import mimetypes
import os
import re
import shutil
import subprocess
import sys
import threading
import webbrowser
from datetime import datetime, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlencode, urlparse


_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))
WORKSPACE = _SCRIPTS_DIR.parent
# Docker 部署下 /mvp 与 /repo 是两个独立 bind mount（不再是同一棵目录树的父子关系），
# 优先信任 PDCA_REPO_ROOT；否则按源码仓库里 WORKSPACE 的实际嵌套深度回退。
_env_repo_root = os.environ.get("PDCA_REPO_ROOT", "").strip()
if _env_repo_root:
    REPO_ROOT = Path(_env_repo_root)
else:
    _parents = WORKSPACE.parents
    REPO_ROOT = _parents[1] if len(_parents) > 1 else _parents[0]
RUN_SCRIPT = WORKSPACE / "scripts" / "run_data_role_pdca_daily.ps1"
QUESTION_TEMPLATE = WORKSPACE / "templates" / "daily_questionnaire.md"
HERMES_HOME = Path.home() / ".hermes" / "profiles"
DATA_REPORTS = REPO_ROOT / "data_reports"
HOST = "127.0.0.1"
PORT = int(os.environ.get("PDCA_WORKBENCH_PORT", "8765"))
VPS_CACHE_SECONDS = 300
RAW_SALES_CACHE_SECONDS = 600
_VPS_CACHE = {}

CUSTOMER_MGMT_ROOT = Path(r"C:\Users\frank\Documents\Codex\2026-05-27\pdca-codex-1-guru-electronics-singapore\he-haiwen-dealer-workbench")
CUSTOMER_MGMT_PORT = 8787
WALKIN_COCKPIT_DIR = WORKSPACE / "modules" / "walkin_cockpit"
WALKIN_COCKPIT_ROOT = WALKIN_COCKPIT_DIR.resolve()
ONLINE_COCKPIT_DIR = WORKSPACE / "modules" / "online_cockpit"
ONLINE_COCKPIT_ROOT = ONLINE_COCKPIT_DIR.resolve()
HOME_DASHBOARD_DIR = WORKSPACE / "modules" / "home_dashboard"
HOME_DASHBOARD_ROOT = HOME_DASHBOARD_DIR.resolve()
HOME_DASHBOARD_INDEX = HOME_DASHBOARD_DIR / "index.html"
MEETING_CENTER_DIR = WORKSPACE / "modules" / "meeting_center"
MEETING_CENTER_ROOT = MEETING_CENTER_DIR.resolve()
MEETING_CENTER_INDEX = MEETING_CENTER_DIR / "index.html"
DEALER_REF_JSON = WALKIN_COCKPIT_DIR / "data" / "dealer_distribution_reference.json"
CUSTOMER_TIER_TARGETS = {"S": 8, "A": 20, "B": 45, "C": 30}
TIER_ESTIMATE_SELL_OUT = {"A": 1200000, "B": 650000, "S": 380000, "C": 280000}


def resolve_cockpit_asset(base_dir, root_dir, rel_path):
    """解析驾驶舱模块静态资源路径，禁止目录穿越。"""
    rel = unquote((rel_path or "").lstrip("/"))
    if not rel:
        rel = "index.html"
    parts = rel.replace("\\", "/").split("/")
    if any(part in ("", ".", "..") for part in parts):
        return None
    target = (base_dir / rel).resolve()
    if not target.is_relative_to(Path(root_dir).resolve()):
        return None
    if not target.is_file():
        return None
    return target


def resolve_walkin_asset(rel_path):
    return resolve_cockpit_asset(WALKIN_COCKPIT_DIR, WALKIN_COCKPIT_ROOT, rel_path)


def resolve_online_asset(rel_path):
    return resolve_cockpit_asset(ONLINE_COCKPIT_DIR, ONLINE_COCKPIT_ROOT, rel_path)


def resolve_home_dashboard_asset(rel_path):
    return resolve_cockpit_asset(HOME_DASHBOARD_DIR, HOME_DASHBOARD_ROOT, rel_path)


def resolve_meeting_center_asset(rel_path):
    return resolve_cockpit_asset(MEETING_CENTER_DIR, MEETING_CENTER_ROOT, rel_path)


def serve_home_dashboard_index(handler):
    """经营驾驶舱首页（dashboard_template_with_api_hooks）。"""
    if not HOME_DASHBOARD_INDEX.is_file():
        handler.send_response(404)
        handler.end_headers()
        return
    handler.send_file(HOME_DASHBOARD_INDEX)


DASHBOARD_THEME_CSS = HOME_DASHBOARD_DIR / "workbench-unified.css"
DASHBOARD_THEME_MARKER = "workbench-unified.css"
COCKPIT_SHELL_CSS = HOME_DASHBOARD_DIR / "workbench-cockpit-shell.css"
COCKPIT_SHELL_MARKER = "workbench-cockpit-shell.css"

try:
    from workbench_data import build_online_channel_payload, build_walkin_api_payload
except ImportError:
    build_walkin_api_payload = None
    build_online_channel_payload = None

try:
    from vemory_bridge import (
        classify_meeting_bucket,
        fetch_vemory_meetings,
        meeting_center_counts,
        todo_assignments,
        vemory_people,
    )
except ImportError:
    classify_meeting_bucket = None
    fetch_vemory_meetings = None
    meeting_center_counts = None
    todo_assignments = None
    vemory_people = None


def skin_cockpit_html(html, date_text, page_title="经销商驾驶舱"):
    """客流/线上子页：统一浅色皮肤 + 顶栏返回工作台。"""
    date_text = date_text or today_text()
    back_href = f"/?date={date_text}"
    if COCKPIT_SHELL_MARKER not in html and COCKPIT_SHELL_CSS.is_file():
        shell_css = COCKPIT_SHELL_CSS.read_text(encoding="utf-8")
        unified = ""
        if DASHBOARD_THEME_CSS.is_file():
            unified = DASHBOARD_THEME_CSS.read_text(encoding="utf-8")
        html = html.replace(
            "</head>",
            f'<link rel="stylesheet" href="/workbench-cockpit-shell.css?v=2">\n'
            f'<style id="wb-cockpit-skin">\n{unified}\n{shell_css}\n</style>\n</head>',
            1,
        )
    if "wb-cockpit-backbar" not in html:
        bar = (
            f'<div class="wb-cockpit-backbar">'
            f'<a class="back-to-workbench" href="{back_href}" title="返回 PDCA 工作台">← 返回工作台</a>'
            f'<span class="wb-cockpit-title">{esc(page_title)}</span>'
            f"</div>"
        )
        html = re.sub(r"(<body[^>]*>)", r"\1" + bar, html, count=1)
    else:
        html = re.sub(
            r'(<a class="back-to-workbench" href=")[^"]*(")',
            rf"\1{back_href}\2",
            html,
            count=1,
        )
    return html


def skin_dashboard_html(html, date_text):
    """为数据看板注入与主页一致的样式，并修正返回工作台链接。"""
    if "wb-unified-skin" not in html and DASHBOARD_THEME_CSS.is_file():
        theme_css = DASHBOARD_THEME_CSS.read_text(encoding="utf-8")
        html = html.replace(
            "</head>",
            f'<style id="wb-unified-skin">\n{theme_css}\n</style>\n'
            '<link rel="stylesheet" href="/dashboard-theme.css?v=1">\n</head>',
            1,
        )
    back_href = f"/?date={date_text}"
    html = re.sub(
        r'(<a class="back-to-workbench" href=")[^"]*(")',
        rf"\1{back_href}\2",
        html,
        count=1,
    )
    return html


def serve_dashboard_html(handler, dashboard_path, date_text):
    html = dashboard_path.read_text(encoding="utf-8")
    handler.send_html(skin_dashboard_html(html, date_text))


def filter_dealers_for_user(dealers, session_user):
    """sales 角色只看自己名下代理（salesPerson 匹配 sales_name），其余角色看全公司。"""
    if not session_user or session_user.get("role") != "sales":
        return dealers
    sales_name = str(session_user.get("sales_name") or "").strip()
    if not sales_name:
        return dealers
    return [d for d in dealers if str(d.get("salesPerson") or "").strip() == sales_name]


def load_dealer_reference():
    if not DEALER_REF_JSON.is_file():
        return []
    try:
        payload = json.loads(DEALER_REF_JSON.read_text(encoding="utf-8"))
        return payload.get("dealers") or []
    except (json.JSONDecodeError, OSError):
        return []


def fmt_cny(amount):
    value = int(round(float(amount or 0)))
    return f"¥ {value:,}"


def dealer_sell_out_total(dealers):
    total = sum(float(d.get("sellOutAmount") or 0) for d in dealers)
    if total > 0:
        return total, "代理商终销汇总"
    estimate = 0.0
    for dealer in dealers:
        ctype = (dealer.get("customerType") or "S").upper()[:1]
        estimate += TIER_ESTIMATE_SELL_OUT.get(ctype, 420000)
    return estimate, "代理商终销汇总"


def load_chart_data(date_text):
    """读取当日 PDCA 生成的 chart_data.json（与数据看板同源）。"""
    path = output_dir(date_text) / "chart_data.json"
    if not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def chart_performance_total(chart_data, period, date_text, sales_name=""):
    """
    与数据看板「实际达成」同口径：VPS 销售员业绩合计。
    月视图 = salesperson_top（看板本月 583.76万 即此字段之和）。
    sales_name 非空时只汇总 dimension 匹配该销售员的行。
    """
    period_rows = {
        "day": ("salesperson_daily", "今日"),
        "week": ("salesperson_week", "本周"),
        "month": ("salesperson_top", "本月"),
        "quarter": ("salesperson_top", "本月累计"),
    }
    key, label = period_rows.get(period, period_rows["month"])
    rows = chart_data.get(key) or []
    if sales_name:
        rows = [r for r in rows if str(r.get("dimension") or "").strip() == sales_name]
    total_yuan = sum(float(row.get("performance") or 0) for row in rows)
    wan = round(total_yuan / 10000, 2)
    return total_yuan, wan, f"{label}实际达成"


def sell_in_from_chart(date_text, period, sales_name=""):
    chart = load_chart_data(date_text)
    if not chart:
        return None
    return chart_performance_total(chart, period, date_text, sales_name=sales_name)


def api_dashboard_overview(date_text, period, session_user=None):
    if session_user:
        name, role, _source = resolve_workbench_profile(session_user)
    else:
        identity = fetch_vps_identity()
        user = identity.get("user") or {} if identity.get("ok") else {}
        name = nested_value(user, "employee_name", "name", "display_name") or nested_value(user, "name", "display_name") or "数据岗"
        role = nested_value(user, "job_title", "role") or "PDCA 工作台"
    dealers = filter_dealers_for_user(load_dealer_reference(), session_user)
    sell_out, sell_out_sub = dealer_sell_out_total(dealers)
    sell_in_sub = ""
    sell_in_wan = 0.0
    scope_sales_name = ""
    if session_user and session_user.get("role") == "sales":
        scope_sales_name = str(session_user.get("sales_name") or "").strip()
    chart_sell_in = sell_in_from_chart(date_text, period, sales_name=scope_sales_name)
    if chart_sell_in:
        sell_in, sell_in_wan, sell_in_sub = chart_sell_in
    else:
        sell_in = sell_out * 1.28 if sell_out else 0
        sell_in_sub = "待更新业绩数据"
        sell_in_wan = round(sell_in / 10000, 2) if sell_in else 0
    im_unread = fetch_vps_im_unread(with_latest=False)
    pdca_plan = fetch_pdca_today_plan(date_text)
    out = output_dir(date_text)
    pdca_path = out / "pdca_daily_check.md"
    score = 88
    comment = "昨日客户触达与待办推进正常，请关注 IM 未读与高优先级待办。"
    if pdca_path.is_file():
        text = pdca_path.read_text(encoding="utf-8", errors="ignore")[:800]
        if "风险" in text or "高风险" in text:
            score = 76
            comment = "PDCA 日结提示存在风险项，请优先处理检查报告中的异常。"
        elif text.strip():
            comment = "已读取今日 PDCA 检查摘要，建议结合数据看板核对 Sell out 与过程指标。"
    if not pdca_plan["ok"]:
        score = max(70, score - 6)
        comment += f" 昨日日报拉取异常：{pdca_plan['warning'][:60]}"
    elif not pdca_plan["rows"]:
        score = max(74, score - 4)
        comment += f" 昨日（{pdca_plan['yesterday']}）日报未写入明日计划，请补交或打开 PDCA 日结。"
    elif len(pdca_plan["rows"]) > 8:
        score = max(72, score - 4)
        comment += f" 今日计划 {len(pdca_plan['rows'])} 项（来自昨日日报明日计划），建议按优先级闭环。"
    else:
        comment += f" 今日计划 {len(pdca_plan['rows'])} 项来自昨日日报明日计划。"
    if im_unread["ok"] and im_unread["unread_count"] > 0:
        score = max(68, score - 5)
        comment += f" IM 未读 {im_unread['unread_count']} 条待处理。"
    period_note = {"day": "日", "week": "周", "month": "月", "quarter": "季"}.get(period, "日")
    return {
        "managerName": name,
        "managerRole": f"{role} · {period_note}视图 · {date_text}",
        "sellInAmount": fmt_cny(sell_in),
        "sellInWan": sell_in_wan,
        "sellOutAmount": fmt_cny(sell_out),
        "sellOutWan": round(sell_out / 10000, 2) if sell_out else 0,
        "sellInSub": sell_in_sub,
        "sellOutSub": sell_out_sub,
        "agentScore": score,
        "scoreComment": comment,
    }


def fetch_pdca_today_plan(date_text):
    """今日待办：取昨日 IM 群日报 payload.tomorrow（vertu odoo daily-report user-summary）。"""
    yesterday_text = previous_date_text(date_text)
    if should_use_local_pdca_cache(date_text):
        yesterday = local_daily_report_cache(yesterday_text, "本地汇报缓存")
    else:
        yesterday = fetch_vps_daily_report(yesterday_text)
    rows = report_payload_items(yesterday, "tomorrow") if yesterday.get("ok") else []
    warning = yesterday.get("warning") or yesterday.get("error") or ""
    if yesterday.get("from_cache"):
        warning = warning or "VPS 暂不可用，已使用本地日报/待办缓存。"
    return {
        "ok": yesterday.get("ok", False),
        "rows": rows,
        "yesterday": yesterday_text,
        "report_count": len(yesterday.get("reports") or []),
        "warning": warning.strip(),
        "from_cache": bool(yesterday.get("from_cache")),
    }


def api_todos_today(date_text=None):
    date_text = date_text or today_text()
    plan = fetch_pdca_today_plan(date_text)
    if not plan["ok"]:
        detail = plan["warning"] or "昨日日报拉取失败"
        return [{
            "id": 0,
            "title": f"PDCA 日结：{detail[:100]}",
            "status": "异常",
            "source": "PDCA 日结",
            "yesterday": plan["yesterday"],
        }]
    rows = plan["rows"]
    if not rows:
        hint = "昨日日报未写入明日计划" if plan["report_count"] else "未查询到昨日群日报"
        return [{
            "id": 0,
            "title": f"PDCA 日结：{hint}（{plan['yesterday']}）",
            "status": "待补充",
            "source": "PDCA 日结",
            "yesterday": plan["yesterday"],
        }]
    items = []
    for index, row in enumerate(rows[:15], start=1):
        progress = task_progress(row)
        status = nested_value(row, "state_display", "status_name", "status.name", "stage_name") or "待处理"
        if progress >= 100:
            status = "已完成"
        elif progress > 0 and status in ("", "待处理", "未开始"):
            status = f"进行中 {progress}%"
        items.append({
            "id": index,
            "title": task_title(row) or "未命名事项",
            "status": status,
            "progress": progress,
            "deadline": task_deadline(row) or date_text,
            "source": "昨日日报·明日计划",
            "yesterday": plan["yesterday"],
            "from_cache": plan["from_cache"],
        })
    return items


def api_hermes_agent_tasks(date_text):
    todos = api_todos_today(date_text)
    tasks = []
    for row in todos[:3]:
        if "拉取失败" in row["title"]:
            continue
        tasks.append({"id": row["id"], "title": f"Hermes：跟进待办「{row['title']}」"})
    if not tasks:
        tasks = [
            {"id": 1, "title": "Hermes：运行今日 PDCA 并刷新数据看板"},
            {"id": 2, "title": "Hermes：核对代理商终销与 Walk-in 客流异常"},
        ]
    out = output_dir(date_text)
    if not (out / "dashboard.html").exists():
        tasks.insert(0, {"id": 0, "title": "Hermes：生成今日 dashboard.html"})
    return tasks[:5]


def api_customer_center_summary(session_user=None):
    dealers = filter_dealers_for_user(load_dealer_reference(), session_user)
    buckets = {"S": [], "A": [], "B": [], "C": []}
    for dealer in dealers:
        ctype = (dealer.get("customerType") or "S").upper()[:1]
        if ctype not in buckets:
            ctype = "C"
        buckets[ctype].append(dealer)
    result = []
    for level in ("S", "A", "B", "C"):
        rows = buckets[level]
        target = CUSTOMER_TIER_TARGETS[level]
        touched = min(target, len(rows))
        result.append({"level": level, "total": len(rows), "touched": touched, "target": target})
    if not dealers:
        for level, target in CUSTOMER_TIER_TARGETS.items():
            result.append({"level": level, "total": 0, "touched": 0, "target": target})
    return result


def api_hr_summary():
    return [
        {"key": "resume", "label": "简历数", "value": 0},
        {"key": "interview", "label": "面试数", "value": 0},
        {"key": "onboard", "label": "到岗数", "value": 0},
        {"key": "leave", "label": "离职数", "value": 0},
        {"key": "leaveRate", "label": "离职率", "value": "—"},
    ]


def api_exceptions(date_text):
    business = []
    affair = []
    dealers = load_dealer_reference()
    low_sell = [d for d in dealers if float(d.get("sellOutAmount") or 0) <= 0]
    if low_sell:
        business.append({
            "owner": "代理商终销",
            "content": f"{len(low_sell)} 家代理商终销金额为 0，请更新 Excel 后重跑导入脚本",
        })
    out = output_dir(date_text)
    pdca_path = out / "pdca_daily_check.md"
    if pdca_path.is_file():
        for line in pdca_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            text = line.strip()
            if not text.startswith("-") and "风险" not in text and "异常" not in text:
                continue
            if "日报" in text or "待办" in text:
                affair.append({"owner": "PDCA", "content": text.lstrip("- ").strip()[:120]})
            else:
                business.append({"owner": "PDCA", "content": text.lstrip("- ").strip()[:120]})
            if len(business) + len(affair) >= 6:
                break
    im_unread = fetch_vps_im_unread(with_latest=False)
    if im_unread["ok"] and im_unread["unread_count"] > 0:
        affair.append({
            "owner": "IM",
            "content": f"{im_unread['unread_count']} 条未读消息分布在 {im_unread['channel_count']} 个会话",
        })
    todos = fetch_vps_today_todos()
    if todos["ok"] and todos["count"] > 8:
        affair.append({"owner": "待办", "content": f"今日待办 {todos['count']} 项，存在积压风险"})
    if not (out / "dashboard.html").exists():
        business.append({"owner": "数据看板", "content": f"{date_text} 数据看板尚未生成，请运行 PDCA"})
    return {"business": business[:6], "affair": affair[:6]}


def api_important_matters(date_text):
    matters = []
    dealers = load_dealer_reference()
    if dealers:
        teams = {}
        for d in dealers:
            teams[d.get("team") or "未分组"] = teams.get(d.get("team") or "未分组", 0) + 1
        top_team = max(teams.items(), key=lambda item: item[1])[0]
        matters.append({
            "id": 1,
            "title": f"{top_team} 组代理商覆盖 {teams[top_team]} 家",
            "desc": "建议对照 Walk-in 客流分析台检查低留资率代理商。",
            "suggestion": "今日打开海外客流看板，筛选留资率偏低代理商并安排督导跟进。",
        })
    todos = fetch_vps_today_todos()
    if todos["ok"] and todos["rows"]:
        first = todos["rows"][0]
        title = nested_value(first, "title", "name") or "待办"
        matters.append({
            "id": 2,
            "title": f"优先待办：{title}",
            "desc": "来自 VPS 今日待办列表。",
            "suggestion": "今日内更新状态并同步到 PDCA 日结。",
        })
    if not matters:
        matters.append({
            "id": 1,
            "title": "运行今日 PDCA",
            "desc": "产出 Excel、报告与看板后驾驶舱指标将自动丰富。",
            "suggestion": "执行 run_data_role_pdca_daily.ps1 或从经典首页触发运行。",
        })
    return matters[:4]


def fetch_pdca_delivery_summary(date_text):
    """今日计划交付统计（昨日明日计划 vs 今日完成 vs VPS 待办）。"""
    yesterday_text = previous_date_text(date_text)
    if should_use_local_pdca_cache(date_text):
        daily = local_daily_report_cache(date_text, "本地汇报缓存")
        yesterday = local_daily_report_cache(yesterday_text, "本地汇报缓存")
        all_todos = {"ok": True, "rows": local_todo_payload_rows(date_text), "count": 0, "error": ""}
    else:
        daily = fetch_vps_daily_report(date_text)
        yesterday = fetch_vps_daily_report(yesterday_text)
        all_todos = fetch_vps_all_todos()
    today_plan_rows = report_payload_items(yesterday, "tomorrow") if yesterday.get("ok") else []
    today_done_rows = report_payload_items(daily, "today") if daily.get("ok") else []
    checks = build_delivery_checks(
        today_plan_rows,
        today_done_rows,
        all_todos["rows"] if all_todos["ok"] else [],
        date_text,
    )
    stats = {"done": 0, "progress": 0, "pending": 0, "risk": 0}
    for item in checks:
        level = item.get("level") or "pending"
        stats[level] = stats.get(level, 0) + 1
    return {
        "yesterday": yesterday_text,
        "plan_count": len(today_plan_rows),
        "stats": stats,
    }


def api_task_center_panel(date_text=None):
    """任务中心：统计 + 昨日日报明日计划列表。"""
    date_text = date_text or today_text()
    plan = fetch_pdca_today_plan(date_text)
    delivery = fetch_pdca_delivery_summary(date_text)
    items = api_todos_today(date_text)
    stats = delivery["stats"]
    total = delivery["plan_count"]
    if total == 0 and items and not str(items[0].get("title", "")).startswith("PDCA 日结："):
        total = len(items)
    done = stats.get("done", 0)
    return {
        "summary": [
            {"key": "total", "label": "总任务数", "value": total},
            {"key": "done", "label": "已完成", "value": done},
            {"key": "undone", "label": "未完成", "value": max(0, total - done)},
        ],
        "yesterday": plan.get("yesterday") or delivery["yesterday"],
        "sourceNote": "vertu odoo daily-report · 昨日群日报「明日计划」",
        "items": items,
    }


def api_task_center_summary(date_text=None):
    """任务中心入口统计：今日计划交付概况（点击进入 PDCA 日结页）。"""
    date_text = date_text or today_text()
    delivery = fetch_pdca_delivery_summary(date_text)
    stats = delivery["stats"]
    total = delivery["plan_count"]
    done = stats.get("done", 0)
    return [
        {"key": "total", "label": "总任务数", "value": total},
        {"key": "done", "label": "已完成", "value": done},
        {"key": "undone", "label": "未完成", "value": max(0, total - done)},
    ]


def enrich_vemory_meetings(payload: dict) -> dict:
    """为会议列表附加首页分类与任务分配建议。"""
    if not payload or not isinstance(payload.get("meetings"), list):
        return payload or {}
    meetings = []
    for meeting in payload["meetings"]:
        row = dict(meeting)
        if classify_meeting_bucket:
            row["bucket"] = classify_meeting_bucket(row)
        if todo_assignments:
            row["assignments"] = todo_assignments(row)
        meetings.append(row)
    payload = dict(payload)
    payload["meetings"] = meetings
    if meeting_center_counts:
        payload["counts"] = meeting_center_counts(meetings)
    return payload


def api_meeting_center_summary(date_text=None, end_date=None):
    date_text = date_text or today_text()
    if not fetch_vemory_meetings:
        return [
            {"key": "total", "label": "总会议数", "value": 0},
            {"key": "interview", "label": "面试会议", "value": 0},
            {"key": "report", "label": "汇报会议", "value": 0},
            {"key": "customer", "label": "客户会议", "value": 0},
        ]
    payload = enrich_vemory_meetings(
        fetch_vemory_meetings(date_text, vertu_cmd=vertu_command(), end_date=end_date or "")
    )
    counts = payload.get("counts") or meeting_center_counts(payload.get("meetings") or [])
    labels = {
        "total": "总会议数",
        "interview": "面试会议",
        "report": "汇报会议",
        "customer": "客户会议",
    }
    return [{"key": key, "label": labels[key], "value": counts.get(key, 0)} for key in labels]


def api_meeting_center_meetings(date_text, person_phone="", person_name="", end_date=""):
    if not fetch_vemory_meetings:
        return {"ok": False, "error": "vemory_bridge 未加载", "meetings": [], "counts": {}, "summary": {}}
    payload = enrich_vemory_meetings(
        fetch_vemory_meetings(date_text, person_phone, person_name, vertu_cmd=vertu_command(), end_date=end_date or "")
    )
    return payload


def api_meeting_center_people():
    if not vemory_people:
        return {"ok": True, "people": []}
    return {"ok": True, "people": vemory_people()}


def api_meeting_center_dispatch(body: dict, date_text: str):
    assignments = body.get("assignments") or []
    meeting_title = body.get("meeting_title") or "会议"
    if not assignments:
        return {"ok": False, "error": "没有待分配事项"}
    created = 0
    errors = []
    for item in assignments:
        todo = item.get("todo") or {}
        title = (todo.get("text") or "").strip() or "会议待办"
        assignee = (item.get("assignee") or "").strip()
        customer = (item.get("customer") or "").strip()
        reason = (item.get("reason") or "").strip()
        remark_parts = [f"来源：Vemory 会议「{meeting_title}」", reason]
        if assignee:
            remark_parts.append(f"建议负责人：{assignee}")
        if customer:
            remark_parts.append(f"关联客户：{customer}")
        remark = "；".join(part for part in remark_parts if part)
        due = normalize_deadline_for_vps(todo.get("due") or date_text)
        try:
            args = ["odoo", "project", "todo", "create", "--title", title, "--remark", remark]
            if due:
                args.extend(["--deadline", due])
            run_vertu_write_json(args)
            created += 1
        except Exception as exc:
            errors.append(f"{title}：{exc}")
    if created and not errors:
        return {"ok": True, "message": f"已从会议「{meeting_title}」写入 {created} 条 VPS 待办。"}
    if created:
        return {
            "ok": True,
            "message": f"已写入 {created} 条，{len(errors)} 条失败：{'；'.join(errors[:3])}",
        }
    return {"ok": False, "error": errors[0] if errors else "VPS 待办写入失败"}


def dispatch_home_dashboard_api(path, query):
    date_text = query.get("date", [today_text()])[0] or today_text()
    period = query.get("period", ["day"])[0] or "day"
    routes = {
        "/api/dashboard/overview": lambda: api_dashboard_overview(date_text, period),
        "/api/dashboard/sell-in": lambda: {
            "amount": api_dashboard_overview(date_text, period)["sellInAmount"],
            "wan": api_dashboard_overview(date_text, period)["sellInWan"],
            "note": api_dashboard_overview(date_text, period)["sellInSub"],
        },
        "/api/dashboard/sell-out": lambda: {
            "amount": api_dashboard_overview(date_text, period)["sellOutAmount"],
        },
        "/api/todos/today": lambda: api_todos_today(date_text),
        "/api/hermes-agent/tasks": lambda: api_hermes_agent_tasks(date_text),
        "/api/customer-center/summary": api_customer_center_summary,
        "/api/hr/summary": api_hr_summary,
        "/api/exceptions": lambda: api_exceptions(date_text),
        "/api/important-matters": lambda: api_important_matters(date_text),
        "/api/task-center/summary": lambda: api_task_center_summary(date_text),
        "/api/task-center/panel": lambda: api_task_center_panel(date_text),
        "/api/meeting-center/summary": lambda: api_meeting_center_summary(date_text),
        "/api/meeting-center/meetings": lambda: api_meeting_center_meetings(
            date_text,
            (query.get("phone") or [""])[0],
            (query.get("name") or [""])[0],
        ),
        "/api/meeting-center/people": api_meeting_center_people,
    }
    factory = routes.get(path)
    if not factory:
        return None
    return factory()


_customer_proc = None  # 客户管理后台进程句柄


AGENT_CARDS = [
    {
        "key": "data-access-agent",
        "title": "数据出表 Agent",
        "profile": "data-access-agent",
        "source": "Hermes profile",
        "desc": "从 VPS/Odoo 按你的要求拉取真实数据，并输出 Excel 表格和数据结论。",
    },
    {
        "key": "logistics-browser-agent",
        "title": "物流官网核查 Agent",
        "profile": "logistics-browser-agent",
        "source": "data_role_pdca_mvp/agents/logistics-browser-agent.md",
        "desc": "拿物流单号访问 UPS/FedEx/DHL 等官网，判断正常、异常或待人工确认。",
    },
    {
        "key": "research-agent",
        "title": "市场调研 Agent",
        "profile": "research-agent",
        "source": "data_role_pdca_mvp/agents/research-agent.md",
        "desc": "调研竞品、客户背景、国家市场、门店资料和公开网页信息，并输出带来源的 Markdown 报告。",
    },
]
AGENT_CORE_FILES = ["SOUL.md", "IDENTITY.md", "AGENTS.md", "MEMORY.md", "USER.md"]


def vertu_command():
    configured = os.environ.get("VERTU_COMMAND")
    if configured:
        if Path(configured).name.lower() in {"vertu", "vertu.cmd", "vertu.ps1"}:
            configured = "vertu-cli"
        configured_path = Path(configured)
        if configured_path.exists():
            return str(configured_path)
        discovered = shutil.which(configured)
        if discovered:
            return discovered
    discovered = shutil.which("vertu-cli")
    if discovered:
        return discovered
    npm_cmd = Path.home() / "AppData" / "Roaming" / "npm" / "vertu-cli.cmd"
    if npm_cmd.exists():
        return str(npm_cmd)
    return "vertu-cli"

QUESTION_TITLES = [
    "1. 今天完成了什么？",
    "2. 明天要完成什么？",
    "3. 昨天未完成事项，今天完成了哪些？",
    "4. 上级临时交办，今天交付了哪些？",
    "5. 今天还有哪些未完成？",
    "6. 今天遇到的卡点",
    "7. 需要上级或业务方确认的事项",
]


def today_text():
    return datetime.now().strftime("%Y-%m-%d")


def esc(value):
    return html.escape(str(value or ""), quote=True)


def output_dir(date_text):
    return WORKSPACE / "outputs" / date_text


def latest_file(paths):
    existing = [Path(path) for path in paths if Path(path).exists()]
    return max(existing, key=lambda path: path.stat().st_mtime) if existing else None


def latest_output_file(date_text, target):
    out = output_dir(date_text)
    if target == "workbook":
        return latest_file(out.glob(f"{date_text}_data_summary*.xlsx"))
    if target == "report":
        return latest_file([out / "data_summary_report.md"])
    if target == "pdca":
        return latest_file([out / "pdca_daily_check.md"])
    if target == "dashboard":
        return latest_file([out / "dashboard.html"])
    if target == "im":
        return latest_file([WORKSPACE / "outbox" / f"{date_text}_im_message.md"])
    return None


def file_time_label(path):
    if not path or not Path(path).exists():
        return "运行后生成"
    return datetime.fromtimestamp(Path(path).stat().st_mtime).strftime("%H:%M:%S")


def questionnaire_path(date_text):
    return WORKSPACE / "inputs" / "questionnaires" / f"{date_text}_questionnaire.md"


def todo_path(date_text):
    return WORKSPACE / "inputs" / "todos" / f"{date_text}_todos.csv"


def logistics_path(date_text):
    return WORKSPACE / "inputs" / "logistics" / f"{date_text}_tracking.csv"


def read_text(path):
    return path.read_text(encoding="utf-8-sig") if path.exists() else ""


def read_json(path):
    return json.loads(read_text(path) or "{}")


def write_text(path, value):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value.rstrip() + "\n", encoding="utf-8")


def sales_aliases():
    path = WORKSPACE / "config" / "sales_aliases.csv"
    aliases = {}
    if not path.exists():
        return aliases
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            for row in csv.DictReader(file):
                raw = (row.get("raw_sales") or "").strip()
                canonical = (row.get("canonical_sales") or "").strip()
                if raw and canonical:
                    aliases[raw.lower()] = canonical
                    aliases[canonical.lower()] = canonical
    except Exception:
        return aliases
    return aliases


def canonical_sales_name(value):
    text = str(value or "").strip()
    if not text:
        return ""
    return sales_aliases().get(text.lower(), text)


def sales_data_files():
    raw_dir = REPO_ROOT / "data_raw"
    if not raw_dir.exists():
        return []
    return sorted(
        raw_dir.glob("dealer_sales_month_to_date_*.json"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )


def sales_data_file_for_range(run_date, start_date=None):
    suffix = run_date if not start_date or start_date == run_date else f"{start_date}_to_{run_date}"
    return REPO_ROOT / "data_raw" / f"dealer_sales_month_to_date_{suffix}.json"


def is_fresh_sales_data(path):
    if not path or not Path(path).exists():
        return False
    age_seconds = datetime.now().timestamp() - Path(path).stat().st_mtime
    if age_seconds > RAW_SALES_CACHE_SECONDS:
        return False
    try:
        payload = read_json(Path(path))
    except Exception:
        return False
    result = None
    if isinstance(payload, dict):
        if isinstance(payload.get("execution"), dict):
            result = payload["execution"].get("result")
        if result is None and isinstance(payload.get("ai"), dict):
            result = payload["ai"].get("result")
        if result is None:
            result = payload.get("result")
    return isinstance(result, dict) and bool(result.get("summary_mode"))


def latest_sales_data_payload():
    for path in sales_data_files():
        try:
            payload = read_json(path)
        except Exception:
            continue
        result = None
        if isinstance(payload, dict):
            if isinstance(payload.get("execution"), dict):
                result = payload["execution"].get("result")
            if result is None and isinstance(payload.get("ai"), dict):
                result = payload["ai"].get("result")
            if result is None:
                result = payload.get("result")
        if isinstance(result, dict) and result.get("summary_mode"):
            return path, result
    return None, None


def is_sales_data_query(query):
    text = query.lower()
    if is_logistics_query(query):
        return False
    intent_words = ["业绩", "销售", "出表", "excel", "表格", "拉一下", "拉取", "客户", "产品", "团队"]
    return any(word in text for word in intent_words)


def is_research_query(query):
    text = str(query or "").lower()
    keywords = [
        "调研", "研究", "竞品", "市场", "客户背景", "背景调查", "公开资料", "行业",
        "渠道", "国家", "政策", "资料整理", "分析一下", "research", "market", "competitor",
    ]
    return any(keyword in text for keyword in keywords)


def is_vps_cli_query(query):
    return "从vps" in str(query or "").lower().replace(" ", "")


def requested_vps_date_range(query):
    text = str(query or "")
    today = datetime.now()
    iso_match = re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if iso_match:
        year, month, day = map(int, iso_match.groups())
        run_date = datetime(year, month, day)
        return run_date.strftime("%Y-%m-%d"), run_date.strftime("%Y-%m-01")
    md_match = re.search(r"(\d{1,2})\s*月\s*(\d{1,2})\s*[日号]?", text)
    if md_match:
        month, day = map(int, md_match.groups())
        run_date = datetime(today.year, month, day)
        return run_date.strftime("%Y-%m-%d"), run_date.strftime("%Y-%m-01")
    month_match = re.search(r"(\d{1,2})\s*月|([一二三四五六七八九十])月", text)
    if month_match:
        chinese_months = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9, "十": 10}
        month = int(month_match.group(1)) if month_match.group(1) else chinese_months.get(month_match.group(2), today.month)
        day = today.day if month == today.month else 28
        run_date = datetime(today.year, month, day)
        return run_date.strftime("%Y-%m-%d"), run_date.strftime("%Y-%m-01")
    return today.strftime("%Y-%m-%d"), today.strftime("%Y-%m-01")


def query_requires_fresh_vps(query):
    text = str(query or "").lower()
    return any(word in text for word in ["刷新", "重新", "最新", "实时", "强制"])


def pull_vps_sales_data(run_date, start_date=None, force=False):
    cached_path = sales_data_file_for_range(run_date, start_date)
    if not force and is_fresh_sales_data(cached_path):
        return cached_path
    puller = WORKSPACE / "scripts" / "pull_vps_sales_data.ps1"
    if not puller.exists():
        raise RuntimeError(f"VPS 拉数脚本不存在：{puller}")
    command = [
        "powershell.exe",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(puller),
        "-Date",
        run_date,
        "-Workspace",
        str(WORKSPACE),
    ]
    if start_date:
        command.extend(["-StartDate", start_date])
    try:
        completed = subprocess.run(
            command,
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
        )
    except subprocess.TimeoutExpired:
        if cached_path.exists():
            return cached_path
        raise RuntimeError("VPS-CLI 拉数超过 120 秒，已自动终止。")
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part).strip()
    if completed.returncode != 0:
        raise RuntimeError(output or "VPS-CLI 拉数失败")
    for line in reversed(output.splitlines()):
        path = Path(line.strip().strip('"'))
        if path.exists():
            return path
    fallback = sales_data_file_for_range(run_date, start_date)
    return fallback if fallback.exists() else None


def run_vps_cli_query(query):
    run_date, start_date = requested_vps_date_range(query)
    try:
        if is_sales_data_query(query):
            pulled_path = pull_vps_sales_data(run_date, start_date, force=query_requires_fresh_vps(query))
            result = run_sales_data_query(query)
            if result.get("ok"):
                result["content"] = (
                    "已按「从vps」规则调用 VPS-CLI 拉取最新数据，并直接生成结果文件。\n\n"
                    f"- VPS 原始数据：{Path(pulled_path).name if pulled_path else '已刷新'}\n"
                    + result.get("content", "")
                )
                result["kind"] = "vps_sales_excel"
            return result
    except Exception as exc:
        return {"ok": False, "content": f"VPS-CLI 执行失败：{exc}", "path": None, "kind": "vps_error"}

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    path = DATA_REPORTS / f"{stamp}_vps-cli-routing_summary.md"
    content = "\n".join([
        "# VPS-CLI 查询路由结果",
        "",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## 用户问题",
        "",
        query,
        "",
        "## 处理结果",
        "",
        "已识别到关键词「从vps」，但当前工作台只内置了经销商业绩类查询到 Excel 的稳定映射。",
        "",
        "## 已支持",
        "",
        "- 从vps拉销售/业绩/产品/客户/团队汇总：自动调用 VPS-CLI，生成 Excel。",
        "- 从vps拉某个销售（例如 Lina）的五月业绩：自动筛选销售员并生成 Excel。",
        "",
        "## 下一步",
        "",
        "请把需求写成类似：`从vps拉一下lina五月业绩`、`从vps拉五月经销商业绩表`。",
    ])
    write_text(path, content)
    return {"ok": True, "content": content, "path": str(path), "filename": path.name, "kind": "vps_markdown"}


def run_research_chat(query):
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    topic = "research-agent"
    out_path = DATA_REPORTS / f"{stamp}_{topic}_summary.md"
    prompt = f"""
你是经销商 PDCA 工作台的 research-agent，负责公开资料调研。

用户调研问题：
{query}

请遵守：
1. 只使用公开资料和用户提供资料。
2. 不编造来源；没有可靠来源时写“未找到可靠公开来源”。
3. 每条关键结论尽量附来源 URL 或来源名称。
4. 最多 15 步，超过即停止并说明当前进展。
5. 输出不超过 4000 tokens。
6. 不登录第三方账号，不绕过验证码，不访问需要付费或权限的内容。

输出 Markdown，必须包含：
# 调研报告
## 调研问题
## 核心结论
## 关键事实
## 证据来源
## 不确定事项
## 建议下一步
""".strip()
    command = [
        hermes_exe(),
        "chat",
        "-q",
        prompt,
        "-Q",
        "--max-turns",
        "15",
    ]
    try:
        completed = subprocess.run(
            command,
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
        )
    except subprocess.TimeoutExpired:
        return {"ok": False, "content": "调研 Agent 执行超过 120 秒，已强制终止。请缩小调研范围后重试。", "path": None, "kind": "research_timeout"}
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part).strip()
    if completed.returncode != 0 and not output:
        return {"ok": False, "content": "调研 Agent 调用失败，且没有返回错误信息。", "path": None, "kind": "research_error"}
    content = output or "调研 Agent 已运行，但没有返回内容。"
    if "Error code:" in content or "Incorrect API key" in content:
        return {"ok": False, "content": f"调研 Agent 调用失败：{content}", "path": None, "kind": "research_error"}
    report = "\n\n".join([
        content,
        "",
        f"---\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\nAgent: research-agent",
    ])
    write_text(out_path, report)
    return {"ok": True, "content": report[-4000:], "path": str(out_path), "filename": out_path.name, "kind": "research"}


def requested_salesperson(query, rows):
    text = query.lower()
    aliases = sales_aliases()
    candidates = set(aliases.values())
    candidates.update(canonical_sales_name(row.get("salesperson")) for row in rows or [])
    for name in sorted((item for item in candidates if item), key=len, reverse=True):
        if name.lower() in text:
            return name
    if "lina" in text:
        return "Lina"
    return ""


def summary_row(row, key_name):
    return [
        canonical_sales_name(row.get(key_name)) if key_name == "salesperson" else (row.get(key_name) or ""),
        float(row.get("performance") or 0),
        float(row.get("quantity") or 0),
        int(float(row.get("line_count") or row.get("rows") or 0)),
    ]


def run_sales_data_query(query):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception as exc:
        return {"ok": False, "content": f"本机缺少 openpyxl，无法生成 Excel：{exc}", "path": None}

    data_path, result = latest_sales_data_payload()
    if not result:
        return {"ok": False, "content": "没有找到可用的 VPS 月累业绩数据，请先运行数据拉取。", "path": None}

    salesperson_rows = result.get("salesperson_summary") or []
    target_sales = requested_salesperson(query, salesperson_rows)
    filtered_sales = [
        row for row in salesperson_rows
        if not target_sales or canonical_sales_name(row.get("salesperson")).lower() == target_sales.lower()
    ]
    if target_sales and not filtered_sales:
        return {"ok": False, "content": f"已读取 {data_path.name}，但没有找到销售员 {target_sales} 的业绩。", "path": None}

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    safe_target = safe_name(target_sales or "all-sales")
    out_path = DATA_REPORTS / f"{stamp}_sales_performance_{safe_target}.xlsx"
    DATA_REPORTS.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(title, rows, key_name):
        ws = wb.create_sheet(title[:31])
        headers = ["维度", "业绩", "数量", "明细行数"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F62BD")
        for row in rows or []:
            ws.append(summary_row(row, key_name))
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12

    add_sheet("销售员汇总", filtered_sales or salesperson_rows, "salesperson")
    add_sheet("团队汇总", result.get("team_summary") or [], "team")
    add_sheet("产品TOP", result.get("product_summary") or [], "product_name")
    add_sheet("客户TOP", result.get("customer_summary") or [], "partner_name")
    wb.save(out_path)

    total_perf = sum(float(row.get("performance") or 0) for row in (filtered_sales or salesperson_rows))
    total_qty = sum(float(row.get("quantity") or 0) for row in (filtered_sales or salesperson_rows))
    label = target_sales or "全部销售员"
    content = (
        f"已直接从 VPS 月累数据生成 Excel，不再走 Hermes 对话。\n\n"
        f"- 查询对象：{label}\n"
        f"- 数据文件：{data_path.name}\n"
        f"- 数据周期：{result.get('month_start', '')} 至 {result.get('run_date', '')}\n"
        f"- 业绩合计：{total_perf:,.2f}\n"
        f"- 数量合计：{total_qty:,.2f}\n"
        f"- 输出文件：{out_path.name}"
    )
    return {"ok": True, "content": content, "path": str(out_path), "filename": out_path.name, "kind": "sales_excel"}


def latest_hermes_report(topic, started_at=0):
    if not DATA_REPORTS.exists():
        return None
    candidates = sorted(
        DATA_REPORTS.glob(f"*_{topic}_summary.md"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    for candidate in candidates:
        if candidate.stat().st_mtime >= started_at - 5:
            return candidate
    return candidates[0] if candidates else None


def resolve_hermes_output_path(output, topic, started_at):
    if output:
        for line in reversed(output.splitlines()):
            text = line.strip().strip('"')
            if not text:
                continue
            path = Path(text)
            if path.exists():
                return path
    return latest_hermes_report(topic, started_at)


def is_logistics_query(query):
    text = query.lower()
    return any(token in text for token in ["物流", "快递", "单号", "tracking", "ups", "fedex", "dhl", "顺丰"])


def extract_tracking_numbers(query):
    numbers = []
    text = query.upper()
    for match in re.finditer(r"1Z(?:[\s\-_:：]*[A-Z0-9]){16}", text):
        value = re.sub(r"[^A-Z0-9]", "", match.group(0))
        if value not in numbers:
            numbers.append(value)
    for match in re.finditer(r"[A-Z0-9][A-Z0-9\s\-_:：]{8,40}[A-Z0-9]", text):
        value = re.sub(r"[^A-Z0-9]", "", match.group(0))
        if "1Z" in value and not value.startswith("1Z"):
            value = value[value.index("1Z"):]
        if 10 <= len(value) <= 30 and any(char.isdigit() for char in value) and value not in numbers:
            numbers.append(value)
    return numbers


def infer_carrier(tracking_number):
    value = tracking_number.upper()
    if value.startswith("1Z"):
        return "UPS"
    if re.fullmatch(r"\d{12,15}", value):
        return "FedEx"
    if re.fullmatch(r"\d{10}", value):
        return "DHL"
    if value.startswith("SF"):
        return "SF"
    return "未知"


def carrier_tracking_url(carrier, tracking_number):
    carriers = read_json(WORKSPACE / "config" / "carriers.json")
    info = carriers.get(carrier) or {}
    template = info.get("tracking_url", "")
    return template.replace("{tracking_number}", tracking_number) if template else ""


def browser_executable_path():
    candidates = [
        Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
        Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
        Path(r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"),
        Path(r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"),
    ]
    return next((str(path) for path in candidates if path.exists()), None)


def parse_dhl_tracking_text(text, tracking_number):
    cleaned = re.sub(r"[ \t]+", " ", text or "")
    status = ""
    for candidate in ["Delivered", "Out for Delivery", "In Transit", "Shipment information received", "Exception"]:
        if re.search(rf"\b{re.escape(candidate)}\b", cleaned, re.I):
            status = candidate
            break
    last_update = ""
    match = re.search(r"Last Update:\s*(.+?)(?:\n|Origin:|Destination:|Authenticate|Sign up)", cleaned, re.S)
    if match:
        last_update = " ".join(match.group(1).split())
    origin = ""
    match = re.search(r"Origin:\s*(.+?)(?:\n|Destination:|Authenticate|Sign up)", cleaned, re.S)
    if match:
        origin = " ".join(match.group(1).split())
    destination = ""
    match = re.search(r"Destination:\s*(.+?)(?:\n|Authenticate|Sign up|Shipment Details)", cleaned, re.S)
    if match:
        destination = " ".join(match.group(1).split())
    service = ""
    match = re.search(r"Service\s*\n\s*(.+?)(?:\n|1 Piece ID|Waybill Number)", cleaned, re.S)
    if match:
        service = " ".join(match.group(1).split())
    return {
        "tracking_number": tracking_number,
        "carrier": "DHL",
        "status": status or "官网已打开，未识别到状态",
        "last_update": last_update,
        "origin": origin,
        "destination": destination,
        "service": service,
        "source": "DHL 官网浏览器抓取",
    }


def track_dhl_with_browser(tracking_number):
    executable = browser_executable_path()
    if not executable:
        raise RuntimeError("未找到 Chrome 或 Edge，无法启动浏览器核查。")
    step_count = 0

    def step(label):
        nonlocal step_count
        step_count += 1
        if step_count > 15:
            raise RuntimeError(f"浏览器核查超过 15 步，已强制终止：{label}")

    from playwright.sync_api import sync_playwright

    with sync_playwright() as playwright:
        step("启动浏览器")
        browser = playwright.chromium.launch(
            headless=True,
            executable_path=executable,
            args=["--disable-http2", "--disable-blink-features=AutomationControlled"],
        )
        try:
            step("打开页面")
            page = browser.new_page(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
            )
            url = "https://www.dhl.com/global-en/home/tracking.html"
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=30000)
            except Exception:
                # DHL 偶发 HTTP2/加载超时，页面主体已渲染时继续后续步骤。
                pass
            step("等待页面加载")
            page.wait_for_timeout(4000)
            popup_errors = []
            for selector in ["#onetrust-accept-btn-handler", 'button:has-text("Accept All")', 'button:has-text("Stay on this site")']:
                step(f"处理弹窗 {selector}")
                try:
                    page.locator(selector).first.click(timeout=3000, force=True)
                    page.wait_for_timeout(800)
                except Exception as exc:
                    popup_errors.append(f"{selector}: {exc}")
            step("输入单号")
            page.locator('input[name="tracking-id"]').fill(tracking_number, timeout=8000)
            step("点击 Track")
            page.locator('button:has-text("Track")').first.click(timeout=8000, force=True)
            step("等待官网结果")
            page.wait_for_timeout(12000)
            step("读取页面文本")
            text = page.locator("body").inner_text(timeout=10000)
            result = parse_dhl_tracking_text(text, tracking_number)
            if popup_errors:
                result["note"] = "部分弹窗选择器未命中，不影响已读取页面文本。"
            return result
        finally:
            browser.close()


def run_logistics_chat(query):
    tracking_numbers = extract_tracking_numbers(query)
    if not tracking_numbers:
        return {"ok": False, "content": "我识别到这是物流任务，但没有识别出快递单号。请直接输入 UPS/FedEx/DHL/SF 单号。", "path": None}
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    path = DATA_REPORTS / f"{stamp}_logistics-tracking_summary.md"
    lines = [
        "# 物流核查 Agent 结果",
        "",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## 用户问题",
        "",
        query,
        "",
        "## 识别结果",
        "",
        "已识别为物流查询，并优先通过官网浏览器核查读取页面状态。浏览器核查最多 15 步，超过即强制终止。",
        "",
        "| 单号 | 承运商 | 官网核查入口 | 当前判断 | 更新时间 | 始发地 | 目的地 |",
        "|---|---|---|---|---|---|---|",
    ]
    links = []
    tracking_results = []
    for number in tracking_numbers:
        carrier = infer_carrier(number)
        url = carrier_tracking_url(carrier, number)
        if url:
            links.append({"label": f"打开 {carrier} 官网核查 {number}", "url": url})
        link = f"[打开官网]({url})" if url else "未识别承运商，需补充"
        result = {
            "tracking_number": number,
            "carrier": carrier,
            "status": "待官网核查",
            "last_update": "",
            "origin": "",
            "destination": "",
            "source": "官网入口",
        }
        if carrier == "DHL":
            try:
                result = track_dhl_with_browser(number)
            except Exception as exc:
                result["status"] = f"官网浏览器核查失败：{exc}"
        tracking_results.append(result)
        lines.append(
            f"| {number} | {carrier} | {link} | {result.get('status', '')} | "
            f"{result.get('last_update', '')} | {result.get('origin', '')} | {result.get('destination', '')} |"
        )
    lines.extend([
        "",
        "## 下一步",
        "",
        "- 弹窗中会直接显示官网核查卡片。",
        "- 点“查看结果”可看到 Markdown 明细。",
        "- 点“用本机软件打开”会打开 Markdown 报告。",
        "- DHL 已接入浏览器官网核查；UPS/FedEx 后续可按同样方式接入。",
    ])
    write_text(path, "\n".join(lines))
    content = read_text(path).strip()
    return {
        "ok": True,
        "content": content,
        "path": str(path),
        "filename": path.name,
        "kind": "logistics",
        "links": links,
        "tracking_results": tracking_results,
    }


def read_csv_rows(path):
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv_rows(path, fieldnames, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def safe_name(value):
    return "".join(char if char.isalnum() or char in "-_." else "-" for char in str(value or "")).strip(".-") or "item"


def hermes_exe():
    configured = os.environ.get("HERMES_COMMAND")
    if configured:
        return configured
    bundled = Path(os.environ.get("LOCALAPPDATA", "")) / "hermes" / "hermes-agent" / "venv" / "Scripts" / "hermes.exe"
    if bundled.exists():
        return str(bundled)
    discovered = shutil.which("hermes")
    return discovered or "hermes"


def agent_by_key(key):
    return next((agent for agent in AGENT_CARDS if agent["key"] == key), None)


def agent_profile_dir(agent):
    return HERMES_HOME / agent["profile"]


def agent_soul_path(agent):
    return agent_profile_dir(agent) / "SOUL.md"


def agent_core_file_path(agent, filename):
    safe_file = filename if filename in AGENT_CORE_FILES else "SOUL.md"
    return agent_profile_dir(agent) / safe_file


def ensure_agent_soul(agent):
    path = agent_soul_path(agent)
    if path.exists():
        return path
    path.parent.mkdir(parents=True, exist_ok=True)
    source_path = WORKSPACE / "agents" / f"{agent['key']}.md"
    source_text = read_text(source_path)
    content = source_text or f"# {agent['key']}\n\n{agent['desc']}\n"
    write_text(path, content)
    return path


def ensure_agent_core_file(agent, filename):
    if filename == "SOUL.md":
        return ensure_agent_soul(agent)
    path = agent_core_file_path(agent, filename)
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        title = filename.replace(".md", "")
        write_text(path, f"# {title}\n\n")
    return path


def list_agent_skills(agent):
    skills_dir = agent_profile_dir(agent) / "skills"
    if not skills_dir.exists():
        return []
    return sorted(
        [path.parent.name for path in skills_dir.glob("*/SKILL.md")],
        key=str.lower,
    )


def skill_name_from_content(filename, content):
    for line in content.splitlines()[:20]:
        if line.lower().startswith("name:"):
            return safe_name(line.split(":", 1)[1].strip())
    stem = Path(filename or "uploaded-skill").stem
    return safe_name(stem if stem.upper() != "SKILL" else "uploaded-skill")


def install_skill_to_agent(agent_key, filename, content_bytes):
    agent = agent_by_key(agent_key)
    if not agent:
        raise ValueError("未知 Agent。")
    text = content_bytes.decode("utf-8-sig", errors="replace")
    skill_name = skill_name_from_content(filename, text)
    target = agent_profile_dir(agent) / "skills" / skill_name
    target.mkdir(parents=True, exist_ok=True)
    (target / "SKILL.md").write_text(text, encoding="utf-8")
    return target


def run_hermes_chat(query):
    if not query.strip():
        return {"ok": False, "content": "请输入要问 Hermes 的内容。", "path": None}
    try:
        if is_vps_cli_query(query):
            return run_vps_cli_query(query)
    except Exception as exc:
        return {"ok": False, "content": f"VPS-CLI 路由失败：{exc}", "path": None}
    try:
        if is_research_query(query):
            return run_research_chat(query)
    except Exception as exc:
        return {"ok": False, "content": f"调研 Agent 执行失败：{exc}", "path": None}
    try:
        if is_sales_data_query(query):
            return run_sales_data_query(query)
    except Exception as exc:
        return {"ok": False, "content": f"业绩出表失败：{exc}", "path": None}
    try:
        if is_logistics_query(query):
            return run_logistics_chat(query)
    except Exception as exc:
        return {"ok": False, "content": f"物流核查失败：{exc}", "path": None}
    topic = "pdca-workbench-chat"
    started_at = datetime.now().timestamp()
    command = [
        "powershell.exe",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(REPO_ROOT / "scripts" / "invoke-data-access-agent.ps1"),
        "-Query",
        query,
        "-Topic",
        topic,
    ]
    try:
        completed = subprocess.run(
            command,
            cwd=str(REPO_ROOT),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
        )
    except subprocess.TimeoutExpired:
        return {"ok": False, "content": "Hermes 执行超过 120 秒，已自动终止。演示时可改成更明确的数据出表或调研问题。", "path": None}
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part).strip()
    if completed.returncode != 0:
        if "No inference provider configured" in output:
            return {"ok": False, "content": (
                "Hermes 模型还没有配置好。\n\n"
                "已知修复方式：把 Hermes 配置为 custom provider，并使用 DashScope OpenAI 兼容地址：\n"
                "hermes config set model.provider custom\n"
                "hermes config set model.default qwen-plus\n"
                "hermes config set model.base_url https://dashscope.aliyuncs.com/compatible-mode/v1\n"
            ), "path": None}
        return {"ok": False, "content": f"Hermes 调用失败：{output}", "path": None}
    path = resolve_hermes_output_path(output, topic, started_at)
    if path and path.exists():
        if path.suffix.lower() not in {".md", ".txt", ".json", ".csv", ".html", ".htm"}:
            return {"ok": True, "content": f"Hermes 已生成文件：{path.name}", "path": str(path), "filename": path.name}
        content = read_text(path).strip()
        if content:
            return {"ok": True, "content": content[-4000:], "path": str(path), "filename": path.name}
    if output:
        return {"ok": True, "content": output, "path": None}
    return {"ok": True, "content": "Hermes 已成功运行，但没有返回文本。模型已配置，可继续重试更明确的指令。", "path": None}


def extract_json_payload(text):
    decoder = json.JSONDecoder()
    for index, char in enumerate(text):
        if char not in "[{":
            continue
        try:
            payload, _ = decoder.raw_decode(text[index:])
            return payload
        except json.JSONDecodeError:
            continue
    raise ValueError("VPS 返回内容不是 JSON")


def _flag_value(args, flag, default=""):
    try:
        return args[args.index(flag) + 1]
    except (ValueError, IndexError):
        return default


def translate_vertu_cli_args(args):
    """把工作台仍在使用的旧 Odoo 参数映射为 vertu-cli 2.x 快捷命令。"""
    if args[:2] == ["odoo", "me"]:
        return ["hr", "+me"]
    if args[:3] == ["odoo", "daily-report", "user-summary"]:
        mapped = ["report", "+user-summary"]
        for flag in ("--user-id", "--start-time", "--end-time"):
            value = _flag_value(args, flag)
            if value:
                mapped.extend([flag, value])
        return mapped
    if args[:4] == ["odoo", "project", "todo", "list"]:
        return ["task", "+tc-todos", "--limit", _flag_value(args, "--limit", "100")]
    if args[:4] == ["odoo", "project", "todo", "create"]:
        mapped = ["task", "+tc-todo-create"]
        for flag in ("--title", "--remark", "--deadline"):
            value = _flag_value(args, flag)
            if value:
                mapped.extend([flag, value])
        return mapped
    if args[:4] == ["odoo", "project", "todo", "update"]:
        mapped = ["task", "+tc-todo-update"]
        for flag in ("--todo-id", "--remark", "--deadline"):
            value = _flag_value(args, flag)
            if value:
                mapped.extend([flag, value])
        status = _flag_value(args, "--status")
        if status:
            mapped.extend(["--status", "done" if status in ("已完成", "completed") else status])
        return mapped
    if args[:4] == ["odoo", "project", "todo", "complete"]:
        return ["task", "+tc-todo-update", "--todo-id", _flag_value(args, "--todo-id"), "--status", "done"]
    if args[:3] == ["odoo", "im", "channels"]:
        return ["im", "+channels", "--limit", _flag_value(args, "--limit", "20")]
    if args[:3] == ["odoo", "im", "search"]:
        return [
            "im",
            "+history",
            "--channel-id",
            _flag_value(args, "--channel-id"),
            "--limit",
            _flag_value(args, "--limit", "20"),
        ]
    return args


def normalize_vertu_cli_payload(cache_key, payload):
    """为旧页面保留 items/results 等字段，底层数据来自 vertu-cli 2.x。"""
    if not isinstance(payload, dict):
        return payload
    result = dict(payload)
    if cache_key.startswith("im_") and "channels" in result:
        result.setdefault("items", result.get("channels") or [])
        result.setdefault("total", result.get("count") or len(result["items"]))
    if cache_key.startswith("im_latest_") and "messages" in result:
        result.setdefault("items", result.get("messages") or [])
    if "todos" in result:
        result.setdefault("items", result.get("todos") or [])
        result.setdefault("results", result.get("todos") or [])
    return result


def vertu_process_command(args):
    """构造可跨平台执行的 vertu-cli 命令；Windows 的 .cmd 需经 cmd /c。"""
    executable = vertu_command()
    command = [executable, *translate_vertu_cli_args(args)]
    if os.name == "nt" and executable.lower().endswith((".cmd", ".bat")):
        return ["cmd", "/c", *command]
    return command


def run_vertu_json(cache_key, args, timeout=8):
    now = datetime.now().timestamp()
    cached = _VPS_CACHE.get(cache_key)
    if cached and now - cached["time"] < VPS_CACHE_SECONDS:
        return cached["payload"], ""
    command = vertu_process_command(args)
    try:
        completed = subprocess.run(
            command,
            cwd=str(WORKSPACE),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
        )
    except subprocess.TimeoutExpired:
        if cached:
            return cached["payload"], ""
        raise RuntimeError("VPS 请求超时，请稍后刷新。")
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part)
    if completed.returncode != 0:
        if cached:
            return cached["payload"], output
        raise RuntimeError(output.strip() or f"vertu-cli 命令失败：{completed.returncode}")
    try:
        payload = normalize_vertu_cli_payload(cache_key, extract_json_payload(output))
    except ValueError:
        if cached:
            return cached["payload"], output
        raise
    _VPS_CACHE[cache_key] = {"time": now, "payload": payload}
    return payload, output


def run_vertu_write_json(args, timeout=60):
    command = vertu_process_command(args)
    completed = subprocess.run(
        command,
        cwd=str(WORKSPACE),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=timeout,
    )
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part)
    if completed.returncode != 0:
        raise RuntimeError(output.strip() or f"vertu-cli 写入失败：{completed.returncode}")
    _VPS_CACHE.clear()
    return extract_json_payload(output), output


def compact_text(value, limit=120):
    text = " ".join(str(value or "").split())
    return text if len(text) <= limit else f"{text[:limit]}..."


def im_channel_url(channel_id):
    template = os.environ.get(
        "PDCA_IM_CHANNEL_URL_TEMPLATE",
        "https://vps-im.vertu.cn/web#action=mail.action_discuss&active_id=discuss.channel_{channel_id}",
    )
    return template.format(channel_id=channel_id)


def fetch_im_latest_message(channel_id):
    payload, _ = run_vertu_json(
        f"im_latest_{channel_id}",
        ["odoo", "im", "search", "--channel-id", str(channel_id), "--role", "any", "--limit", "1"],
    )
    items = payload.get("items") or []
    return items[0] if items else {}


def fetch_vps_im_unread(with_latest=True):
    try:
        payload, _ = run_vertu_json(
            "im_unread_latest" if with_latest else "im_unread",
            ["odoo", "im", "channels", "--has-unread", "--limit", "20"],
        )
        items = payload.get("items") or []
        if with_latest:
            for item in items[:5]:
                try:
                    item["latest_message"] = fetch_im_latest_message(item.get("id"))
                except Exception:
                    item["latest_message"] = {}
        unread_count = sum(int(item.get("unread_count") or 0) for item in items)
        return {
            "ok": True,
            "channels": items,
            "channel_count": int(payload.get("total") or len(items)),
            "unread_count": unread_count,
            "error": "",
        }
    except Exception as exc:
        return {"ok": False, "channels": [], "channel_count": 0, "unread_count": 0, "error": str(exc)}


def fetch_vps_today_todos():
    try:
        payload, _ = run_vertu_json(
            "today_todos",
            ["odoo", "project", "todo", "list", "--for-me", "--due-within-days", "0", "--limit", "20"],
        )
        rows = payload.get("results") or payload.get("items") or []
        return {"ok": True, "rows": rows, "count": int(payload.get("count") or len(rows)), "error": ""}
    except Exception as exc:
        return {"ok": False, "rows": [], "count": 0, "error": str(exc)}


def quote_odoo_domain_value(value: str) -> str:
    """转义 Odoo domain 字符串字面量。"""
    return str(value or "").replace("\\", "\\\\").replace("'", "\\'")


def many2one_name(value):
  """@param value Odoo many2one [id, name]"""
  return value[1] if isinstance(value, list) and len(value) > 1 else ""


def odoo_data_search(model: str, domain: str, fields: str, limit: int = 5) -> tuple[list[dict], str]:
    """
    通过 vertu odoo data search 查询记录。

    @returns (rows, error)
    """
    cache_key = f"odoo_search_{model}_{domain}_{fields}_{limit}"
    try:
        payload, output = run_vertu_json(
            cache_key,
            [
                "odoo",
                "data",
                "search",
                "--model-name",
                model,
                "--domain",
                domain,
                "--fields",
                fields,
                "--limit",
                str(limit),
            ],
            timeout=20,
        )
        if isinstance(payload, list):
            return payload, ""
        return [], output or "unexpected odoo search payload"
    except Exception as exc:
        return [], str(exc)


def workbench_role_label(role: str) -> str:
    """8767 登录角色 → 展示用岗位/权限文案。"""
    mapping = {
        "admin": "系统管理员",
        "manager": "海外中台主管",
        "sales": "经销商销售",
        "dealer": "经销商门店",
        "viewer": "只读访客",
    }
    return mapping.get((role or "").strip().lower(), role or "工作台用户")


def employee_job_label(row: dict) -> str:
    """从 hr.employee 行提取岗位/部门展示。"""
    job = many2one_name(row.get("job_id")) or row.get("job_title") or ""
    dept = many2one_name(row.get("department_id")) or ""
    if job and dept:
        return f"{dept} · {job}"
    return job or dept or ""


def lookup_vps_user_by_name(name_hint: str) -> dict | None:
    """按 res.users 显示名模糊匹配。"""
    text = (name_hint or "").strip()
    if not text:
        return None
    safe = quote_odoo_domain_value(text)
    rows, err = odoo_data_search("res.users", f"name ilike '{safe}'", "id,name,login", 3)
    if err or not rows:
        return None
    target = text.lower()
    for row in rows:
        if str(row.get("name") or "").lower() == target:
            return row
    return rows[0]


def lookup_vps_employee_by_hint(hint: str) -> dict | None:
    """
    按姓名在 VPS hr.employee 中模糊匹配。

    @param hint 姓名或登录名
    @returns 员工 dict 或 None
    """
    text = (hint or "").strip()
    if not text:
        return None
    safe = quote_odoo_domain_value(text)
    rows, err = odoo_data_search(
        "hr.employee",
        f"active = True AND name ilike '%{safe}%'",
        "id,name,department_id,job_id,job_title,work_email,user_id",
        8,
    )
    if err or not rows:
        return None
    target = text.lower()
    for row in rows:
        name = str(row.get("name") or "")
        if name.lower() == target:
            return row
    return rows[0]


def lookup_vps_user_by_login(login: str) -> dict | None:
    """按 res.users.login 精确匹配。"""
    text = (login or "").strip()
    if not text:
        return None
    safe = quote_odoo_domain_value(text)
    rows, err = odoo_data_search("res.users", f"login = '{safe}'", "id,name,login", 1)
    if err or not rows:
        return None
    return rows[0]


def is_generic_profile_label(text: str, role: str) -> bool:
    """是否为角色占位文案，不能当作真实姓名展示。"""
    label = (text or "").strip()
    if not label:
        return True
    if label == workbench_role_label(role):
        return True
    generic = {
        "系统管理员",
        "海外中台主管",
        "只读访客",
        "经销商销售",
        "经销商门店",
        "工作台用户",
        "数据岗",
        "PDCA 工作台",
    }
    return label in generic


def person_name_hints(session_user: dict) -> list[str]:
    """从登录会话提取可用于 VPS 姓名匹配的关键词（排除角色占位）。"""
    username = str(session_user.get("username") or "").strip()
    display_name = str(session_user.get("display_name") or "").strip()
    sales_name = str(session_user.get("sales_name") or "").strip()
    role = str(session_user.get("role") or "").strip()
    hints: list[str] = []
    for item in [sales_name, display_name, username]:
        if item and not is_generic_profile_label(item, role) and item not in hints:
            hints.append(item)
    return hints


def profile_from_vps_me(vu: dict, role: str) -> tuple[str, str]:
    """从 vertu odoo me 解析姓名与组织岗位。"""
    name = nested_value(vu, "employee_name", "name", "display_name")
    job = nested_value(vu, "job_title", "role") or ""
    employee_id = vu.get("employee_id") or vu.get("employeeId")
    if employee_id:
        rows, _ = odoo_data_search(
            "hr.employee",
            f"id = {int(employee_id)}",
            "id,name,department_id,job_id,job_title",
            1,
        )
        if rows:
            name = str(rows[0].get("name") or name)
            job = employee_job_label(rows[0]) or job
    if not job:
        job = workbench_role_label(role)
    return name, job


def resolve_workbench_profile(session_user: dict | None) -> tuple[str, str, str]:
    """
    根据 8767 登录账号解析姓名与岗位（优先 VPS）。

    @returns (name, job_title, source)
    """
    if not session_user:
        return "", "", "none"
    username = str(session_user.get("username") or "").strip()
    display_name = str(session_user.get("display_name") or "").strip()
    sales_name = str(session_user.get("sales_name") or "").strip()
    role = str(session_user.get("role") or "").strip()
    hints = person_name_hints(session_user)
    # 注意：这里不能无条件用 fetch_vps_identity()（服务器本机 vertu 会话）去猜"这是谁"——
    # 那是运行这个进程的机器当前登录的 VPS 账号，和实际发请求登录 8767 的人没有任何绑定关系。
    # 之前这里对 admin/manager/viewer 三个角色无条件信任它，导致不管谁用这几个账号登录，
    # 姓名都被覆盖成服务器机器上 vertu 登录的那个人（一直显示"付汪阳"）。
    # 下面 username==me_login 或姓名命中 hints 才采信的那次 fetch_vps_identity() 调用才是安全的。

    for hint in hints:
        employee = lookup_vps_employee_by_hint(hint)
        if employee:
            name = str(employee.get("name") or hint)
            job = employee_job_label(employee) or workbench_role_label(role)
            return name, job, "vps-hr.employee"

    if username:
        odoo_user = lookup_vps_user_by_login(username)
        if odoo_user:
            name = str(odoo_user.get("name") or "")
            if name and not is_generic_profile_label(name, role):
                return name, workbench_role_label(role), "vps-res.users"

    for hint in hints:
        odoo_user = lookup_vps_user_by_name(hint)
        if odoo_user:
            name = str(odoo_user.get("name") or hint)
            return name, workbench_role_label(role), "vps-res.users-name"

    identity = fetch_vps_identity()
    if identity.get("ok"):
        vu = identity.get("user") or {}
        me_name, me_job = profile_from_vps_me(vu, role)
        me_login = str(vu.get("login") or "").strip()
        if me_name and (
            username and me_login == username
            or any(h and (h in me_name or me_name in h) for h in hints)
        ):
            return me_name, me_job, "vps-me"

    name = hints[0] if hints else (username or "工作台用户")
    if is_generic_profile_label(name, role):
        name = username or "工作台用户"
    return name, workbench_role_label(role), "session"


def fetch_vps_identity():
    try:
        payload, _ = run_vertu_json("current_user", ["odoo", "me"])
        return {"ok": True, "user": payload, "error": ""}
    except Exception as exc:
        return {"ok": False, "user": {}, "error": str(exc)}


def fetch_vps_all_todos():
    try:
        payload, _ = run_vertu_json(
            "all_todos",
            ["odoo", "project", "todo", "list", "--for-me", "--all-pages", "--limit", "100"],
        )
        rows = payload.get("results") or payload.get("items") or []
        return {"ok": True, "rows": rows, "count": int(payload.get("count") or len(rows)), "error": ""}
    except Exception as exc:
        return {"ok": False, "rows": [], "count": 0, "error": str(exc)}


def fetch_vps_daily_report(date_text):
    identity = fetch_vps_identity()
    if not identity["ok"]:
        return local_daily_report_cache(date_text, identity["error"])
    user_id = identity["user"].get("user_id")
    if not user_id:
        return local_daily_report_cache(date_text, "VPS 当前用户缺少 user_id")
    try:
        payload, _ = run_vertu_json(
            f"daily_report_{user_id}_{date_text}",
            [
                "odoo",
                "daily-report",
                "user-summary",
                "--user-id",
                str(user_id),
                "--start-time",
                date_text,
                "--end-time",
                date_text,
            ],
            timeout=45,
        )
        return {
            "ok": True,
            "identity": identity,
            "reports": payload.get("daily_reports") or [],
            "okrs": payload.get("okrs") or [],
            "raw": payload,
            "error": "",
        }
    except Exception as exc:
        return local_daily_report_cache(date_text, str(exc), identity)


def local_identity_cache(identity=None):
    if identity and identity.get("ok"):
        return identity
    return {"ok": True, "user": {"name": "刘春梅", "user_id": "local-cache", "employee_id": "local-cache"}, "error": ""}


def local_todo_payload_rows(date_text):
    rows = read_csv_rows(todo_path(date_text))
    return [
        {
            "title": row.get("title", ""),
            "status": row.get("status", ""),
            "status_name": row.get("status", ""),
            "deadline": row.get("due_date", ""),
            "due_date": row.get("due_date", ""),
            "progress": "100" if str(row.get("status", "")).lower() in {"done", "completed", "已完成", "完成"} else "0",
            "source": row.get("source", "local-cache"),
            "priority": row.get("priority", ""),
            "note": row.get("notes", ""),
        }
        for row in rows
        if row.get("title")
    ]


def local_daily_report_cache(date_text, reason="", identity=None):
    report_path = output_dir(date_text) / "pdca_daily_check.md"
    report_text = compact_text(read_text(report_path), 240) if report_path.exists() else "本地暂未生成 PDCA 日结。"
    today_rows = local_todo_payload_rows(date_text)
    tomorrow_rows = local_todo_payload_rows((datetime.strptime(date_text, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")) if re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_text) else []
    reports = [
        {
            "created_at": f"{date_text} 本地缓存",
            "status": "本地缓存",
            "summary": f"VPS 暂不可用，展示本地 PDCA/待办缓存。{reason}",
            "payload": {"today": today_rows, "tomorrow": tomorrow_rows or today_rows},
        }
    ]
    if today_rows or report_path.exists():
        reports[0]["content"] = report_text
    return {
        "ok": True,
        "from_cache": True,
        "identity": local_identity_cache(identity),
        "reports": reports,
        "okrs": [],
        "raw": {"source": "local-cache", "reason": reason},
        "error": "",
        "warning": f"VPS 暂不可用，已显示本地缓存：{reason}",
    }


def previous_date_text(date_text):
    try:
        return (datetime.strptime(date_text, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    except ValueError:
        return (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")


def report_payload_items(daily, key):
    rows = []
    for report in daily.get("reports", []):
        payload = report.get("payload") or {}
        values = payload.get(key) or []
        if isinstance(values, list):
            rows.extend(values)
    return rows


def fetch_vps_month_okr(date_text):
    identity = fetch_vps_identity()
    if not identity["ok"]:
        return local_month_okr_cache(date_text, identity["error"])
    employee_id = identity["user"].get("employee_id") or identity["user"].get("user_id")
    period = date_text[:7]
    try:
        payload, _ = run_vertu_json(
            f"month_okr_{employee_id}_{period}",
            ["odoo", "okr", "employee-okr-list", "--okr-period", period, "--employee-ids", str(employee_id)],
        )
        results = payload.get("results") or []
        objectives = []
        for item in results:
            objectives.extend(item.get("objectives") or [])
        return {"ok": True, "rows": objectives, "count": len(objectives), "error": ""}
    except Exception as exc:
        return local_month_okr_cache(date_text, str(exc))


def local_month_okr_cache(date_text, reason=""):
    rows = [
        {
            "title": row.get("title", ""),
            "name": row.get("title", ""),
            "score": "本地待办",
            "key_results": [{"name": row.get("notes", "") or row.get("source", "local-cache")}],
        }
        for row in read_csv_rows(todo_path(date_text))
        if row.get("title")
    ]
    return {"ok": True, "from_cache": True, "rows": rows, "count": len(rows), "error": "", "warning": reason}


def should_use_local_pdca_cache(date_text):
    """正式口径优先 vertu CLI；仅演示或显式环境变量时使用本地 CSV/输出缓存。"""
    force = os.environ.get("PDCA_USE_LOCAL_CACHE", "").strip().lower()
    if force in ("1", "true", "yes"):
        return True
    data_sources = read_json(WORKSPACE / "config" / "data_sources.json")
    if str(data_sources.get("official_source", "")).strip().lower() == "vps":
        return False
    return bool(data_sources.get("sales_json")) and (
        todo_path(date_text).exists() or (output_dir(date_text) / "pdca_daily_check.md").exists()
    )


def pdca_vps_source_note(daily, yesterday, month_okr, all_todos):
    """页面上标注当前 PDCA 日结数据来源（VPS 真数 vs 本地回退）。"""
    parts = []
    if daily.get("from_cache") or yesterday.get("from_cache"):
        hint = daily.get("warning") or yesterday.get("warning") or "VPS 暂不可用"
        parts.append(f"群日报：本地缓存（{hint}）")
    elif daily.get("ok") or yesterday.get("ok"):
        parts.append("群日报：vertu odoo daily-report user-summary")
    if month_okr.get("from_cache"):
        parts.append("月待办：本地 CSV 回退")
    elif month_okr.get("ok"):
        parts.append(f"月待办：vertu okr employee-okr-list（{month_okr.get('count', 0)} 项）")
    if all_todos.get("ok"):
        parts.append(f"VPS 待办：vertu project todo list（{all_todos.get('count', 0)} 项）")
    elif all_todos.get("error"):
        parts.append(f"VPS 待办拉取失败：{all_todos['error'][:80]}")
    return " · ".join(parts) if parts else "数据来源：vertu CLI（加载中）"


def nested_value(row, *paths):
    for path in paths:
        value = row
        for part in path.split("."):
            if not isinstance(value, dict):
                value = None
                break
            value = value.get(part)
        if value not in (None, ""):
            return value
    return ""


def todo_table_rows(rows):
    return "".join(
        "<tr>"
        f"<td>{esc(nested_value(row, 'priority', 'priority_name'))}</td>"
        f"<td>{esc(nested_value(row, 'title', 'name'))}</td>"
        f"<td>{esc(nested_value(row, 'status_name', 'status.name', 'stage_name'))}</td>"
        f"<td>{esc(nested_value(row, 'deadline', 'due_date', 'date_deadline'))}</td>"
        "</tr>"
        for row in rows
    )


def first_text(row, *paths):
    value = nested_value(row, *paths)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value or "")


def daily_report_table_rows(rows):
    return "".join(
        "<tr>"
        f"<td>{esc(first_text(row, 'created_at', 'create_date', 'date'))}</td>"
        f"<td>{esc(first_text(row, 'status', 'state', 'report_status'))}</td>"
        f"<td>{esc(compact_text(first_text(row, 'content', 'summary', 'content_summary', 'body', 'report_content'), 180))}</td>"
        "</tr>"
        for row in rows
    )


def pdca_todo_rows(rows):
    return "".join(
        "<tr>"
        f"<td>{esc(nested_value(row, 'title', 'name', 'description'))}</td>"
        f"<td>{esc(nested_value(row, 'status_name', 'status.name', 'stage_name'))}</td>"
        f"<td>{esc(nested_value(row, 'deadline', 'due_date', 'date_deadline', 'end_date'))}</td>"
        f"<td>{esc(nested_value(row, 'progress', 'progress_rate'))}</td>"
        "</tr>"
        for row in rows
    )


def okr_rows(rows):
    return "".join(
        "<tr>"
        f"<td>{esc(nested_value(row, 'title', 'name'))}</td>"
        f"<td>{esc(nested_value(row, 'score'))}</td>"
        f"<td>{esc(len(row.get('key_results') or []))}</td>"
        "</tr>"
        for row in rows
    )


def task_title(row):
    return first_text(row, "title", "name", "description", "task_name", "content")


def normalized_task_text(value):
    return "".join(char.lower() for char in str(value or "") if char.isalnum())


def task_progress(row):
    raw = first_text(row, "progress", "progress_rate", "progress_percent", "completion_rate")
    if not raw:
        raw = first_text(row, "remark", "note", "description")
        match = re.search(r"进度[:：]\s*(\d+(?:\.\d+)?)\s*%", raw)
        raw = match.group(1) if match else ""
    try:
        return int(float(str(raw).replace("%", "").strip()))
    except ValueError:
        return 0


def task_deadline(row):
    return first_text(row, "deadline", "due_date", "date_deadline", "end_date")


def task_status(row):
    return first_text(row, "status_name", "status.name", "stage_name", "state", "status")


def todo_id_value(row):
    return first_text(row, "todo_id", "id")


def is_done_status(value):
    text = str(value or "").lower()
    return any(token in text for token in ["完成", "已完成", "done", "completed", "closed", "close"])


def find_matching_task(title, rows):
    needle = normalized_task_text(title)
    if not needle:
        return None
    for row in rows:
        candidate = normalized_task_text(task_title(row))
        if not candidate:
            continue
        if needle in candidate or candidate in needle:
            return row
    return None


def build_delivery_checks(planned_rows, today_done_rows, vps_todo_rows, date_text):
    checks = []
    for row in planned_rows:
        title = task_title(row) or "未命名事项"
        done_hit = find_matching_task(title, today_done_rows)
        todo_hit = find_matching_task(title, vps_todo_rows)
        progress = max(task_progress(row), task_progress(done_hit or {}), task_progress(todo_hit or {}))
        status_text = task_status(done_hit or {}) or task_status(todo_hit or {}) or task_status(row)
        due = task_deadline(todo_hit or {}) or task_deadline(row)
        has_delivery = bool(done_hit)
        has_done_status = is_done_status(status_text)
        if has_delivery or has_done_status or progress >= 100:
            level = "done"
            label = "已交付"
            advice = "已有今日日报或 VPS 待办完成记录，建议补齐最终交付物链接或结果摘要。"
        elif progress > 0:
            level = "progress"
            label = "进行中"
            advice = "已有进度但未形成完成记录，今天日结时需要补充交付结果和剩余阻塞。"
        elif due and due <= date_text:
            level = "risk"
            label = "高风险"
            advice = "截止日已到但没有完成证据，建议立即跟进负责人，补交结果或调整计划。"
        else:
            level = "pending"
            label = "待交付"
            advice = "尚未看到完成证据，建议在 VPS 待办或今日日报里补充进度。"
        checks.append({
            "title": title,
            "todo_id": todo_id_value(todo_hit or {}),
            "date_text": date_text,
            "level": level,
            "label": label,
            "progress": progress,
            "deadline": due,
            "status": status_text,
            "advice": advice,
            "report_evidence": compact_text(task_title(done_hit) if done_hit else "", 160),
            "todo_evidence": compact_text(task_title(todo_hit) if todo_hit else "", 160),
        })
    return checks


def normalize_deadline_for_vps(value):
    text = str(value or "").strip()
    if not text:
        return ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return f"{text} 18:00:00"
    return text


def save_pdca_task_update(form):
    title = (form.get("title", [""])[0] or "").strip()
    todo_id = (form.get("todo_id", [""])[0] or "").strip()
    status = (form.get("status", [""])[0] or "").strip()
    progress = (form.get("progress", ["0"])[0] or "0").strip()
    deadline = normalize_deadline_for_vps(form.get("deadline", [""])[0])
    note = (form.get("note", [""])[0] or "").strip()
    if not title and not todo_id:
        return "缺少待办标题，无法保存。"
    try:
        progress_value = max(0, min(100, int(float(progress))))
    except ValueError:
        return "进度必须是 0 到 100 的数字。"
    remark_parts = [f"进度：{progress_value}%"]
    if note:
        remark_parts.append(f"备注：{note}")
    remark = "；".join(remark_parts)
    if not todo_id:
        args = ["odoo", "project", "todo", "create", "--title", title, "--remark", remark]
        if deadline:
            args.extend(["--deadline", deadline])
        payload, _ = run_vertu_write_json(args)
        todo_id = str(payload.get("id") or payload.get("todo_id") or nested_value(payload, "result.id", "data.id"))
        if not todo_id:
            return "已创建待办，但 VPS 没有返回 todo_id，请刷新后再改状态。"
    update_args = ["odoo", "project", "todo", "update", "--todo-id", todo_id, "--remark", remark]
    if status:
        update_args.extend(["--status", status])
    if deadline:
        update_args.extend(["--deadline", deadline])
    run_vertu_write_json(update_args)
    if status == "已完成" or progress_value >= 100:
        run_vertu_write_json(["odoo", "project", "todo", "complete", "--todo-id", todo_id])
    return f"已保存进度 {progress_value}% 和状态「{status or '未填写'}」。Agent 判断已刷新。"


def render_delivery_agent(checks, daily_ok=True):
    if not checks:
        return '<section><h2>交付检查 Agent</h2><p>没有可检查的今日计划。请先在「经销商-日报推送」群的昨日日报里写入“明日计划”。</p></section>'
    summary = {
        "done": sum(1 for item in checks if item["level"] == "done"),
        "progress": sum(1 for item in checks if item["level"] == "progress"),
        "pending": sum(1 for item in checks if item["level"] == "pending"),
        "risk": sum(1 for item in checks if item["level"] == "risk"),
    }
    cards = "".join([
        metric_card("已交付", f"{summary['done']} 项", "有今日 IM 日报或 VPS 完成记录", summary["risk"] == 0),
        metric_card("进行中", f"{summary['progress']} 项", "有进度但还缺交付结果", summary["progress"] == 0),
        metric_card("高风险", f"{summary['risk']} 项", "到期但没有完成证据", summary["risk"] == 0),
    ])
    item_cards = []
    for item in checks:
        item_cards.append(f"""
        <details class="delivery-card {esc(item['level'])}" data-delivery-card data-status="{esc(item['level'])}">
          <summary>
            <span class="delivery-title">{esc(item['title'])}</span>
            <span class="delivery-badge">{esc(item['label'])}</span>
          </summary>
          <div class="delivery-body">
            <p><strong>Agent 判断：</strong>{esc(item['advice'])}</p>
            <p><strong>当前进度：</strong>{esc(item['progress'])}%　<strong>截止：</strong>{esc(item['deadline'] or '未填写')}　<strong>状态：</strong>{esc(item['status'] or '未填写')}</p>
            <p><strong>今日 IM 日报证据：</strong>{esc(item['report_evidence'] or '未匹配到完成记录')}</p>
            <p><strong>VPS 待办证据：</strong>{esc(item['todo_evidence'] or '未匹配到待办记录')}</p>
            <form class="progress-form" method="post" action="/pdca-task">
              <input type="hidden" name="date" value="{esc(item.get('date_text', ''))}">
              <input type="hidden" name="title" value="{esc(item['title'])}">
              <input type="hidden" name="todo_id" value="{esc(item.get('todo_id', ''))}">
              <input type="hidden" name="deadline" value="{esc(item.get('deadline', ''))}">
              <label>进度 %<input name="progress" type="number" min="0" max="100" value="{esc(item['progress'])}"></label>
              <label>状态
                <select name="status">
                  <option value="未开始" {"selected" if item["status"] == "未开始" else ""}>未开始</option>
                  <option value="进行中" {"selected" if item["status"] == "进行中" else ""}>进行中</option>
                  <option value="阻塞" {"selected" if item["status"] == "阻塞" else ""}>阻塞</option>
                  <option value="已完成" {"selected" if item["status"] == "已完成" or item["level"] == "done" else ""}>已完成</option>
                </select>
              </label>
              <label>交付/阻塞说明<input name="note" placeholder="例如：已完成资料核查，等销售确认"></label>
              <button type="submit">保存进度和状态</button>
            </form>
          </div>
        </details>
        """)
    daily_note = "" if daily_ok else '<p class="message">注意：未查询到「经销商-日报推送」群的今日日报，Agent 只能根据 VPS 待办进度做临时判断。</p>'
    return f"""
    <section>
      <h2>交付检查 Agent</h2>
      <p>自动对比「经销商-日报推送」群中昨日日报的今日计划、今日日报完成事项和 VPS 待办进度，判断今天每项待办交付到什么程度。</p>
      {daily_note}
      <div class="grid">{cards}</div>
      <div class="actions">
        <button type="button" class="button light" onclick="filterDelivery('all')">全部</button>
        <button type="button" class="button light" onclick="filterDelivery('done')">只看已交付</button>
        <button type="button" class="button light" onclick="filterDelivery('progress')">只看进行中</button>
        <button type="button" class="button light" onclick="filterDelivery('risk')">只看高风险</button>
      </div>
      <div class="delivery-list">{''.join(item_cards)}</div>
      <script>
      /** Filters delivery check cards by Agent status. */
      function filterDelivery(status) {{
        document.querySelectorAll('[data-delivery-card]').forEach(function(card) {{
          card.style.display = status === 'all' || card.dataset.status === status ? '' : 'none';
        }});
      }}
      </script>
    </section>
    """


def im_table_rows(channels, date_text):
    return "".join(
        "<tr>"
        f"<td><a href=\"{esc(route_url('/open-im-channel', date_text, channel_id=channel.get('id')))}\">{esc(channel.get('name'))}</a></td>"
        f"<td>{esc(channel.get('unread_count'))}</td>"
        f"<td>{esc(compact_text((channel.get('latest_message') or {}).get('body_preview') or (channel.get('latest_message') or {}).get('body_text')))}</td>"
        f"<td>{esc(channel.get('last_interest_dt'))}</td>"
        "</tr>"
        for channel in channels
    )


def ensure_questionnaire(date_text):
    path = questionnaire_path(date_text)
    if not path.exists():
        template = read_text(QUESTION_TEMPLATE).replace("YYYY-MM-DD", date_text)
        write_text(path, template)
    return path


def parse_questionnaire(date_text):
    # Rendering a GET page must stay read-only.  Production mounts the release
    # tree read-only and overlays only the runtime input directories as writable.
    # Use the template in memory until the first explicit save creates the file.
    path = questionnaire_path(date_text)
    text = (
        read_text(path)
        if path.exists()
        else read_text(QUESTION_TEMPLATE).replace("YYYY-MM-DD", date_text)
    )
    result = {title: "" for title in QUESTION_TITLES}
    current = None
    buffer = []
    for line in text.splitlines():
        if line.startswith("## "):
            if current:
                result[current] = "\n".join(buffer).strip()
            title = line[3:].strip()
            current = title if title in result else None
            buffer = []
        elif current:
            buffer.append(line)
    if current:
        result[current] = "\n".join(buffer).strip()
    return result


def save_questionnaire(date_text, form):
    lines = [f"# 数据岗位每日 PDCA 问卷 {date_text}", ""]
    for index, title in enumerate(QUESTION_TITLES):
        value = (form.get(f"q{index}", [""])[0] or "").strip()
        lines.extend([f"## {title}", value if value else "-", ""])
    write_text(questionnaire_path(date_text), "\n".join(lines))


def append_todo(date_text, form):
    path = todo_path(date_text)
    fieldnames = ["date", "source", "title", "priority", "status", "owner", "due_date", "notes"]
    rows = read_csv_rows(path)
    rows.append({
        "date": date_text,
        "source": "workbench",
        "title": (form.get("title", [""])[0] or "").strip(),
        "priority": form.get("priority", ["MEDIUM"])[0],
        "status": form.get("status", ["pending"])[0],
        "owner": (form.get("owner", ["frank"])[0] or "frank").strip(),
        "due_date": (form.get("due_date", [date_text])[0] or date_text).strip(),
        "notes": (form.get("notes", [""])[0] or "").strip(),
    })
    rows = [row for row in rows if row.get("title")]
    write_csv_rows(path, fieldnames, rows)


def append_logistics(date_text, form):
    path = logistics_path(date_text)
    fieldnames = ["tracking_number", "carrier", "customer", "salesperson", "ship_date", "expected_status", "current_status", "note"]
    rows = read_csv_rows(path)
    rows.append({
        "tracking_number": (form.get("tracking_number", [""])[0] or "").strip(),
        "carrier": form.get("carrier", ["UPS"])[0],
        "customer": (form.get("customer", [""])[0] or "").strip(),
        "salesperson": (form.get("salesperson", [""])[0] or "").strip(),
        "ship_date": (form.get("ship_date", [date_text])[0] or date_text).strip(),
        "expected_status": (form.get("expected_status", [""])[0] or "").strip(),
        "current_status": (form.get("current_status", [""])[0] or "").strip(),
        "note": (form.get("note", [""])[0] or "").strip(),
    })
    rows = [row for row in rows if row.get("tracking_number")]
    write_csv_rows(path, fieldnames, rows)


def run_pdca(date_text, push=False, start_date=None):
    if os.name == "nt" and RUN_SCRIPT.is_file():
        command = [
            "powershell.exe",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(RUN_SCRIPT),
            "-Date",
            date_text,
        ]
        if start_date:
            command.extend(["-StartDate", start_date])
        if push:
            command.append("-Push")
    else:
        # Linux/Docker 不依赖 PowerShell，直接调用同一底层 Python 流水线。
        script = WORKSPACE / "scripts" / "data_role_pdca_daily.py"
        command = [
            sys.executable,
            str(script),
            "--date", date_text,
            "--workspace", str(WORKSPACE),
        ]
        if start_date:
            command.extend(["--start-date", start_date])

        sources = read_json(WORKSPACE / "config" / "data_sources.json")
        sales_json = Path(str(sources.get("sales_json") or ""))
        if sales_json.is_file():
            command.extend(["--sales-json", str(sales_json)])
        else:
            suffix = f"{start_date}_to_{date_text}" if start_date and start_date != date_text else date_text
            cached = WORKSPACE.parents[1] / "data_raw" / f"dealer_sales_month_to_date_{suffix}.json"
            if cached.is_file():
                command.extend(["--sales-json", str(cached)])

        sales_xlsx = Path(str(sources.get("sales_xlsx") or ""))
        if "--sales-json" not in command and sources.get("allow_excel_demo") and sales_xlsx.is_file():
            command.extend(["--sales-xlsx", str(sales_xlsx)])
            if sources.get("sales_sheet"):
                command.extend(["--sales-sheet", str(sources["sales_sheet"])])

        logistics_csv = Path(str(sources.get("logistics_csv") or ""))
        if logistics_csv.is_file():
            command.extend(["--logistics-csv", str(logistics_csv)])
        if push:
            command.append("--push")
    try:
        completed = subprocess.run(
            command,
            cwd=str(WORKSPACE),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=180,
        )
    except subprocess.TimeoutExpired as exc:
        stdout = exc.stdout.decode("utf-8", errors="replace") if isinstance(exc.stdout, bytes) else (exc.stdout or "")
        stderr = exc.stderr.decode("utf-8", errors="replace") if isinstance(exc.stderr, bytes) else (exc.stderr or "")
        return 124, stdout, stderr or "PDCA 生成超过 180 秒，已自动终止。请稍后重试或使用已缓存看板。"
    return completed.returncode, completed.stdout, completed.stderr


def status_card(label, ok, detail):
    state = "ok" if ok else "warn"
    text = "正常" if ok else "待处理"
    return f"""
    <div class="card {state}">
      <div class="label">{esc(label)}</div>
      <div class="state">{text}</div>
      <p>{esc(detail)}</p>
    </div>
    """


def metric_card(label, value, detail, ok=True, href=None):
    state = "ok" if ok else "warn"
    content = f"""
      <div class="label">{esc(label)}</div>
      <div class="state">{esc(value)}</div>
      <p>{esc(detail)}</p>
    """
    if href:
        return f"""
    <a class="card {state} entry-card" href="{esc(href)}">
{content}
    </a>
    """
    return f"""
    <div class="card {state}">
{content}
    </div>
    """


def is_port_listening(port: int) -> bool:
    """检查本机指定端口是否已在监听。"""
    import socket
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(0.5)
        return s.connect_ex(("127.0.0.1", port)) == 0


def ensure_customer_server() -> str:
    """确保客户管理服务在 8787 端口运行。返回错误信息或空字符串。"""
    global _customer_proc
    if is_port_listening(CUSTOMER_MGMT_PORT):
        return ""
    server_script = CUSTOMER_MGMT_ROOT / "server.py"
    if not server_script.exists():
        return f"客户管理服务脚本不存在：{server_script}"
    try:
        python = sys.executable
        _customer_proc = subprocess.Popen(
            [python, str(server_script)],
            cwd=str(CUSTOMER_MGMT_ROOT),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        # 等待最多 3 秒让端口就绪
        import time
        for _ in range(12):
            time.sleep(0.25)
            if is_port_listening(CUSTOMER_MGMT_PORT):
                return ""
            if _customer_proc.poll() is not None:
                return f"客户管理服务启动后退出，退出码：{_customer_proc.returncode}"
        return "客户管理服务启动超时，请手动运行 server.py"
    except Exception as exc:
        return f"客户管理服务启动失败：{exc}"


def render_customer_mgmt_frame(date_text):
    """把客户管理 8787 以 iframe 方式嵌入工作台，顶部保留返回导航。"""
    back_href = route_url("/", date_text)
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>客户管理 · 经销商PDCA工作台</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{display:flex;flex-direction:column;height:100vh;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif}}
.frame-bar{{display:flex;align-items:center;gap:12px;padding:0 16px;height:42px;background:#1e1e2e;color:#fff;flex-shrink:0;border-bottom:1px solid #313147}}
.frame-bar a.back{{display:flex;align-items:center;gap:6px;color:#cdd6f4;text-decoration:none;font-size:13px;padding:4px 10px;border-radius:6px;border:1px solid #45475a;transition:background .15s}}
.frame-bar a.back:hover{{background:#313147;color:#fff}}
.frame-bar .sep{{width:1px;height:20px;background:#45475a}}
.frame-bar .title{{font-size:13px;font-weight:600;color:#cdd6f4}}
.frame-bar .sub{{font-size:11px;color:#6c7086;margin-left:4px}}
.frame-bar .ext{{margin-left:auto;font-size:12px;color:#6c7086}}
.frame-bar a.ext-link{{color:#89b4fa;text-decoration:none;font-size:12px}}
.frame-bar a.ext-link:hover{{text-decoration:underline}}
iframe{{flex:1;border:none;width:100%}}
</style>
</head>
<body>
<div class="frame-bar">
  <a class="back" href="{esc(back_href)}">← 返回工作台</a>
  <div class="sep"></div>
  <span class="title">客户管理</span>
  <span class="sub">经销商客户台账 · 拜访记录 · 漏斗与回款</span>
  <a class="ext-link ext" href="http://127.0.0.1:{CUSTOMER_MGMT_PORT}?v=20260529-2" target="_blank">在新标签页打开 ↗</a>
</div>
<iframe src="http://127.0.0.1:{CUSTOMER_MGMT_PORT}?v=20260529-2" allowfullscreen></iframe>
</body>
</html>"""


def customer_mgmt_card():
    running = is_port_listening(CUSTOMER_MGMT_PORT)
    state_text = "进入客户管理" if running else "点击启动并进入"
    detail = "经销商客户台账、拜访记录、漏斗与回款管理" if running else "服务未启动，点击后自动启动"
    return f"""
    <a class="card ok entry-card" href="/customer-mgmt">
      <div class="entry-top">
        <div>
          <div class="label">客户管理</div>
          <div class="state">{esc(state_text)}</div>
        </div>
        <span class="entry-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" width="28" height="28" fill="none">
            <circle cx="9" cy="7" r="3" stroke="currentColor" stroke-width="2"/>
            <path d="M3 19c0-3.314 2.686-6 6-6s6 2.686 6 6" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
            <path d="M16 11l2 2 4-4" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
          </svg>
        </span>
      </div>
      <p>{esc(detail)}</p>
    </a>
    """


def dashboard_card(date_text, exists):
    state_text = "打开看板" if exists else "先运行生成"
    detail = "点击进入今日数据看板" if exists else "运行今日 PDCA 后生成看板"
    return f"""
    <a class="card ok entry-card" href="{esc(route_url('/dashboard', date_text))}">
      <div class="entry-top">
        <div>
          <div class="label">数据看板</div>
          <div class="state">{esc(state_text)}</div>
        </div>
        <span class="entry-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" width="28" height="28" fill="none">
            <rect x="3" y="4" width="18" height="16" rx="3" stroke="currentColor" stroke-width="2"/>
            <path d="M7 15v-3M12 15V8M17 15v-5" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
          </svg>
        </span>
      </div>
      <p>{esc(detail)}</p>
    </a>
    """


def walkin_cockpit_card():
    """经销商海外客流分析台入口（modules/walkin_cockpit）。"""
    return """
    <a class="card ok entry-card entry-card-wide" href="/walkin-cockpit/">
      <div class="entry-top">
        <div>
          <div class="label">海外客流</div>
          <div class="state">打开分析台</div>
        </div>
        <span class="entry-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" width="28" height="28" fill="none">
            <path d="M4 10.5 12 4l8 6.5V20a1 1 0 0 1-1 1h-5v-6H10v6H5a1 1 0 0 1-1-1v-9.5Z" stroke="currentColor" stroke-width="2" stroke-linejoin="round"/>
          </svg>
        </span>
      </div>
      <p>海外经销客流：代理商终销、大区与团队表现（墨绿主题）</p>
    </a>
    """


def online_cockpit_card():
    """经销商线上经营入口（已并入客流分析）。"""
    return """
    <a class="card ok entry-card entry-card-wide" href="/walkin-cockpit/#oi-merged">
      <div class="entry-top">
        <div>
          <div class="label">线上经营</div>
          <div class="state">打开驾驶舱</div>
        </div>
        <span class="entry-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" width="28" height="28" fill="none">
            <rect x="3" y="5" width="18" height="14" rx="2" stroke="currentColor" stroke-width="2"/>
            <path d="M8 15h8M8 11h5" stroke="currentColor" stroke-width="2" stroke-linecap="round"/>
            <circle cx="17" cy="8" r="2" fill="currentColor"/>
          </svg>
        </span>
      </div>
      <p>经销商线上经营：OKR、渠道线索、区域汇总（已并入客流分析台）</p>
    </a>
    """


def button(label, href, style=""):
    target = ' target="_blank" rel="noopener"' if str(href).startswith(("http://", "https://")) else ""
    return f'<a class="button {style}" href="{esc(href)}"{target}>{esc(label)}</a>'


def header_subtitle():
    cached = _VPS_CACHE.get("current_user")
    if cached:
        user = cached.get("payload", {})
        name = user.get("employee_name") or user.get("name") or user.get("login")
        if name:
            return f"IM 登录：{name} · 数据看板 · 今日待办 · Hermes 智能体"
    return "IM 登录识别中 · 数据看板 · 今日待办 · Hermes 智能体"


def warm_identity_cache():
    try:
        fetch_vps_identity()
    except Exception as exc:
        _VPS_CACHE["identity_warmup_error"] = {"time": datetime.now().timestamp(), "payload": {"error": str(exc)}}


def is_previewable(path):
    return Path(path).suffix.lower() in {".md", ".txt", ".json", ".csv", ".html", ".htm"}


def page(title, body, date_text, message=""):
    msg_html = f'<div class="message">{esc(message)}</div>' if message else ""
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(title)}</title>
  <style>
    :root {{
      --bg: #eef4fb;
      --surface: #ffffff;
      --surface-soft: #f6f9ff;
      --ink: #17223a;
      --ink-soft: #56647a;
      --ink-muted: #8da0bb;
      --line: #d7e3f2;
      --line-soft: #e7eef8;
      --accent: #2f6fed;
      --accent-soft: #eaf2ff;
      --accent-ink: #174fbf;
      --success: #0f8a4b;
      --warn: #b46a00;
      --shadow-sm: 0 1px 2px rgba(31,34,48,.05);
      --shadow-md: 0 10px 28px rgba(31,80,150,.10);
      --shadow-lg: 0 18px 46px rgba(31,80,150,.16);
    }}
    * {{ box-sizing:border-box; }}
    body {{
      margin:0; background:linear-gradient(180deg,#f6faff 0%, var(--bg) 240px); color:var(--ink);
      font-family:"Inter","SF Pro Text","PingFang SC","Microsoft YaHei",system-ui,sans-serif;
      font-size:15px; line-height:1.6; -webkit-font-smoothing:antialiased;
    }}
    h1,h2,h3 {{ font-family:"Inter","SF Pro Display","PingFang SC","Microsoft YaHei",system-ui,sans-serif; letter-spacing:0; }}
    h2 {{ font-size:18px; margin:0 0 6px; font-weight:600; }}
    h3 {{ font-size:15px; margin:0 0 6px; font-weight:600; color:var(--ink); }}
    p {{ color:var(--ink-soft); margin:6px 0; }}
    a {{ color:var(--accent-ink); }}
    input, textarea, select {{
      background:var(--surface); border:1px solid var(--line); border-radius:10px;
      padding:11px 13px; font-size:14px; color:var(--ink); width:auto;
      font-family:inherit; transition:border-color .15s, box-shadow .15s;
    }}
    input:focus, textarea:focus, select:focus {{
      outline:none; border-color:var(--accent); box-shadow:0 0 0 3px var(--accent-soft);
    }}
    textarea {{ width:100%; min-height:96px; resize:vertical; line-height:1.6; }}
    header {{
      background:rgba(255,255,255,.9); border-bottom:1px solid var(--line);
      padding:18px 36px;
      backdrop-filter:blur(10px);
    }}
    header h1 {{ margin:0 0 4px; font-size:20px; font-weight:700; color:var(--ink); }}
    header p {{ color:var(--ink-muted); margin:0; font-size:13px; }}
    main {{ max-width:1440px; margin:0 auto; padding:24px 28px 56px; display:flex; flex-direction:column; gap:18px; }}
    section {{
      background:var(--surface); border:1px solid var(--line); border-radius:18px;
      box-shadow:var(--shadow-sm); padding:22px 24px;
    }}
    section + section {{ margin-top:0; }}
    .grid {{ display:grid; grid-template-columns:repeat(auto-fit, minmax(220px, 1fr)); gap:16px; }}
    .grid-secondary {{ margin-top:16px; grid-template-columns:repeat(auto-fit, minmax(320px, 1fr)); }}
    .entry-card-wide {{ min-height:88px; }}
    .card {{
      background:var(--surface); border:1px solid var(--line); border-radius:18px;
      padding:18px; box-shadow:var(--shadow-sm); transition:transform .15s, box-shadow .15s, border-color .15s;
    }}
    .card.ok {{ border-top:3px solid var(--success); }}
    .card.warn {{ border-top:3px solid var(--warn); }}
    .entry-card {{ color:inherit; display:block; text-decoration:none; }}
    .entry-card:hover {{ transform:translateY(-2px); box-shadow:var(--shadow-md); border-color:#dcd5c6; }}
    .entry-top {{ display:flex; align-items:flex-start; justify-content:space-between; gap:12px; }}
    .entry-icon {{ align-items:center; background:var(--accent-soft); border-radius:12px; color:var(--accent-ink); display:inline-flex; height:42px; justify-content:center; width:42px; }}
    .label {{ color:var(--ink-muted); font-size:12px; letter-spacing:.04em; text-transform:uppercase; }}
    .state {{ font-size:24px; font-weight:600; margin:6px 0; color:var(--ink); }}
    .actions {{ display:flex; gap:10px; flex-wrap:wrap; margin:14px 0 0; }}
    .button, button {{
      border:0; border-radius:12px; background:var(--accent); color:white;
      padding:10px 18px; font-size:14px; font-weight:600;
      text-decoration:none; cursor:pointer; display:inline-block; line-height:1.2;
      transition:background .15s, transform .05s;
      font-family:inherit;
    }}
    .button:hover, button:hover {{ background:var(--accent-ink); }}
    .button:active, button:active {{ transform:translateY(1px); }}
    button:disabled {{ background:#d8c7bb; color:#fff; cursor:not-allowed; transform:none; opacity:.75; }}
    .button.secondary {{ background:var(--ink); }}
    .button.secondary:hover {{ background:#000; }}
    .button.light {{ background:#fff; color:var(--accent-ink); border:1px solid #b9d1ff; }}
    .button.light:hover {{ background:var(--accent-soft); color:var(--accent-ink); }}
    .button.danger {{ background:#c0413e; }}
    .thinking-inline {{ align-items:center; color:var(--ink-muted); display:none; gap:8px; font-size:13px; font-weight:600; }}
    .thinking-inline.on {{ display:inline-flex; }}
    .spinner {{ animation:spin .8s linear infinite; border:2px solid #eadbd0; border-top-color:var(--accent); border-radius:50%; display:inline-block; height:16px; width:16px; }}
    @keyframes spin {{ to {{ transform:rotate(360deg); }} }}
    table {{ width:100%; border-collapse:collapse; margin-top:10px; font-size:14px; }}
    th, td {{ border-bottom:1px solid var(--line-soft); padding:11px 8px; text-align:left; vertical-align:top; }}
    th {{ color:var(--ink-muted); font-size:12px; font-weight:600; letter-spacing:.04em; text-transform:uppercase; }}
    .message {{ background:var(--accent-soft); color:var(--accent-ink); border:1px solid #ecd6c7; padding:11px 16px; border-radius:10px; font-size:14px; }}
    .two {{ display:grid; grid-template-columns:1fr 1fr; gap:14px; }}
    .agent-section {{ background:var(--surface-soft); border-color:var(--line); }}
    .agent-grid {{ display:grid; grid-template-columns:repeat(auto-fit, minmax(260px, 1fr)); gap:18px; margin-top:18px; }}
    .agent-card {{
      background:var(--surface); border:1px solid var(--line); border-top:none;
      border-radius:14px; padding:24px 20px; text-align:center; min-height:240px;
      display:flex; flex-direction:column; align-items:center;
      transition:transform .15s, box-shadow .15s, border-color .15s;
    }}
    .agent-card:hover {{ transform:translateY(-2px); box-shadow:var(--shadow-md); border-color:#dcd5c6; }}
    .agent-card h3 {{ margin:12px 0 4px; font-size:16px; }}
    .agent-card p {{ color:var(--ink-soft); line-height:1.6; min-height:48px; font-size:14px; }}
    .agent-card .actions {{ justify-content:center; margin-top:auto; }}
    .agent-avatar {{
      align-items:center; background:var(--accent-soft); color:var(--accent-ink);
      border-radius:999px; display:inline-flex; font-size:30px; height:64px; width:64px;
      justify-content:center;
    }}
    .agent-meta {{ color:var(--ink-muted); font-size:12px; margin:4px 0; letter-spacing:.02em; }}
    .agent-button {{ background:var(--accent); color:#fff; min-width:160px; }}
    .agent-button:hover {{ background:var(--accent-ink); }}
    .editor-layout {{ display:grid; grid-template-columns:220px 1fr; gap:20px; }}
    .file-nav {{ background:var(--surface-soft); border:1px solid var(--line); border-radius:12px; padding:8px; }}
    .file-nav h3 {{ color:var(--ink-muted); font-size:11px; letter-spacing:.08em; text-transform:uppercase; padding:8px 12px 6px; margin:0; }}
    .file-nav a {{ color:var(--ink-soft); display:block; padding:9px 12px; text-decoration:none; border-radius:8px; font-size:14px; }}
    .file-nav a:hover {{ background:#efe9da; color:var(--ink); }}
    .file-nav a.active {{ background:var(--accent-soft); color:var(--accent-ink); font-weight:600; }}
    .page-toolbar {{ align-items:flex-start; display:flex; justify-content:space-between; gap:12px; margin-bottom:18px; }}
    .skill-chip {{ background:var(--surface-soft); border:1px solid var(--line); color:var(--ink-soft); border-radius:999px; display:inline-block; margin:4px 6px 4px 0; padding:5px 11px; font-size:13px; }}
    .drop-zone {{ border:1px dashed #cfc7b3; border-radius:12px; color:var(--ink-muted); margin-top:12px; padding:18px; background:var(--surface-soft); }}
    .hermes-result {{ background:#1c1e2a; color:#e8e6df; border-radius:12px; margin-top:14px; max-height:280px; overflow:auto; padding:16px; white-space:pre-wrap; font-family:"JetBrains Mono","SF Mono",ui-monospace,monospace; font-size:13px; line-height:1.6; }}
    .output-grid {{ display:grid; grid-template-columns:repeat(auto-fit, minmax(220px, 1fr)); gap:14px; margin-top:14px; }}
    .output-card {{ background:var(--surface-soft); border:1px solid var(--line); border-radius:12px; padding:16px; display:flex; flex-direction:column; gap:10px; min-height:150px; }}
    .output-card h3 {{ font-size:16px; margin:0; }}
    .output-card p {{ flex:1; font-size:13px; margin:0; }}
    .output-card .output-meta {{ color:var(--ink-muted); flex:0; font-size:12px; }}
    .output-icon {{ align-items:center; background:var(--accent-soft); border-radius:12px; color:var(--accent-ink); display:inline-flex; font-size:22px; height:42px; justify-content:center; width:42px; }}
    .output-card.missing {{ opacity:.55; }}
    details.output-paths {{ background:var(--surface-soft); border:1px solid var(--line); border-radius:12px; margin-top:14px; padding:12px 14px; }}
    details.output-paths summary {{ color:var(--ink-muted); cursor:pointer; font-weight:600; }}
    details.output-paths code {{ color:var(--ink-soft); word-break:break-all; }}
    .result-banner {{ background:var(--surface); border:1px solid var(--line); border-radius:14px; box-shadow:var(--shadow-md); padding:18px; margin-top:14px; }}
    .result-banner.ok {{ border-top:3px solid var(--success); }}
    .result-banner.warn {{ border-top:3px solid var(--warn); }}
    .result-banner h3 {{ margin:0 0 8px; }}
    .result-preview {{ background:var(--surface-soft); border:1px solid var(--line); border-radius:12px; color:var(--ink); margin-top:14px; max-height:360px; overflow:auto; padding:16px; white-space:pre-wrap; }}
    .result-file {{ color:var(--ink-muted); font-size:13px; margin-top:8px; word-break:break-all; }}
    .result-modal {{ align-items:center; background:rgba(31,34,48,.38); bottom:0; display:flex; justify-content:center; left:0; padding:28px; position:fixed; right:0; top:0; z-index:50; }}
    .result-dialog {{ background:var(--surface); border:1px solid var(--line); border-top:3px solid var(--success); border-radius:18px; box-shadow:var(--shadow-lg); max-height:86vh; max-width:900px; overflow:auto; padding:22px; width:min(900px, 100%); }}
    .result-dialog.warn {{ border-top-color:var(--warn); }}
    .result-dialog-head {{ align-items:flex-start; display:flex; gap:14px; justify-content:space-between; }}
    .result-close {{ align-items:center; border:1px solid var(--line); border-radius:999px; color:var(--ink-muted); display:inline-flex; height:34px; justify-content:center; text-decoration:none; width:34px; }}
    .result-close:hover {{ background:var(--surface-soft); color:var(--ink); }}
    .delivery-list {{ display:flex; flex-direction:column; gap:10px; margin-top:14px; }}
    .delivery-card {{ background:var(--surface-soft); border:1px solid var(--line); border-radius:12px; padding:0; overflow:hidden; }}
    .delivery-card summary {{ align-items:center; cursor:pointer; display:flex; justify-content:space-between; gap:12px; padding:14px 16px; }}
    .delivery-title {{ color:var(--ink); font-weight:600; }}
    .delivery-badge {{ border-radius:999px; font-size:12px; font-weight:700; padding:4px 10px; white-space:nowrap; }}
    .delivery-card.done .delivery-badge {{ background:#e1f1e8; color:var(--success); }}
    .delivery-card.progress .delivery-badge {{ background:#fff1d8; color:var(--warn); }}
    .delivery-card.pending .delivery-badge {{ background:#ece8df; color:var(--ink-soft); }}
    .delivery-card.risk .delivery-badge {{ background:#f7dddd; color:#b23a35; }}
    .delivery-card.risk {{ border-color:#efc7c3; }}
    .delivery-body {{ border-top:1px solid var(--line); padding:14px 16px; }}
    .delivery-body p {{ margin:4px 0; }}
    .progress-form {{ align-items:end; background:var(--surface); border:1px solid var(--line); border-radius:12px; display:grid; gap:10px; grid-template-columns:120px 140px 1fr auto; margin-top:12px; padding:12px; }}
    .progress-form label {{ color:var(--ink-muted); font-size:12px; font-weight:600; }}
    .progress-form input, .progress-form select {{ margin-top:4px; width:100%; }}
    .workbench-home {{ display:flex; flex-direction:column; gap:20px; }}
    .period-tabs {{ display:grid; grid-template-columns:repeat(5, minmax(0, 1fr)); gap:14px; }}
    .period-tab {{
      align-items:center; background:#fff; border:1px solid var(--line); border-radius:18px;
      box-shadow:var(--shadow-sm); color:#233451; display:flex; font-size:17px; font-weight:800;
      height:56px; justify-content:center; text-decoration:none;
    }}
    .period-tab.active {{ background:var(--accent); border-color:var(--accent); color:#fff; box-shadow:0 14px 26px rgba(47,111,237,.22); }}
    .home-top {{ display:grid; grid-template-columns:repeat(3, minmax(0, 1fr)); gap:16px; }}
    .home-metric {{
      background:rgba(255,255,255,.94); border:1px solid #cfe0f5; border-radius:22px;
      box-shadow:var(--shadow-sm); min-height:118px; padding:18px;
      display:flex; flex-direction:column; justify-content:space-between;
    }}
    .home-metric-row {{ align-items:center; display:flex; justify-content:space-between; gap:12px; }}
    .home-metric-kicker {{ color:var(--ink-muted); font-size:12px; font-weight:700; text-transform:uppercase; }}
    .home-metric-value {{ color:var(--ink); font-size:34px; font-weight:800; line-height:1; margin:10px 0 6px; }}
    .home-chip {{ background:#edf5ff; border:1px solid #cddfff; border-radius:999px; color:var(--accent-ink); display:inline-flex; font-size:12px; font-weight:800; line-height:1; padding:7px 11px; white-space:nowrap; }}
    .home-chip.ok {{ background:#e7f4ed; border-color:#cfe8d9; color:var(--success); }}
    .home-chip.warn {{ background:#fff4dc; border-color:#ffd36e; color:var(--warn); }}
    .home-board {{ display:grid; grid-template-columns:320px minmax(0, 1fr) 300px; gap:18px; align-items:start; }}
    .home-stack {{ display:flex; flex-direction:column; gap:18px; }}
    .home-panel {{
      background:rgba(255,255,255,.94); border:1px solid #cfe0f5; border-radius:22px;
      box-shadow:var(--shadow-sm); overflow:hidden;
    }}
    .home-panel-head {{
      align-items:center; border-bottom:1px solid var(--line-soft); display:flex;
      justify-content:space-between; gap:12px; min-height:54px; padding:14px 16px;
    }}
    .home-panel-head h2 {{ color:#15223a; font-size:18px; font-weight:800; margin:0; }}
    .home-panel-body {{ padding:14px 16px 16px; }}
    .home-todo-list {{ display:flex; flex-direction:column; }}
    .home-todo {{
      display:grid; grid-template-columns:24px 1fr auto; gap:10px; padding:11px 0;
      border-bottom:1px solid var(--line-soft); align-items:start;
    }}
    .home-todo:last-child {{ border-bottom:0; }}
    .home-check {{ border:1px solid #d7cfc0; border-radius:7px; height:22px; width:22px; }}
    .home-todo b {{ display:block; font-size:14px; line-height:1.35; }}
    .home-todo small {{ color:var(--ink-muted); display:block; font-size:12px; margin-top:3px; }}
    .home-module-grid {{ display:grid; grid-template-columns:repeat(4, minmax(0, 1fr)); gap:12px; }}
    .home-module {{
      background:linear-gradient(180deg,#fff 0%, #f8fbff 100%); border:1px solid #cfe0f5; border-radius:18px;
      color:inherit; display:flex; flex-direction:column; min-height:144px; padding:14px; text-decoration:none;
    }}
    .home-module:hover {{ border-color:#aecaef; box-shadow:var(--shadow-md); transform:translateY(-1px); }}
    .home-module b {{ color:var(--ink); font-size:15px; margin-bottom:4px; }}
    .home-module strong {{ color:var(--ink); display:block; font-size:30px; line-height:1; margin:14px 0 8px; }}
    .home-module span {{ color:var(--ink-soft); font-size:13px; line-height:1.45; }}
    .home-progress {{ background:#e1ebfa; border-radius:999px; height:8px; margin-top:auto; overflow:hidden; }}
    .home-progress i {{ background:var(--accent); border-radius:inherit; display:block; height:100%; }}
    .home-status-grid {{ border:1px solid #cfe0f5; border-radius:18px; display:grid; grid-template-columns:repeat(3, 1fr); margin-top:14px; overflow:hidden; }}
    .home-status-cell {{ background:var(--surface); border-right:1px solid var(--line); padding:14px; }}
    .home-status-cell:last-child {{ border-right:0; }}
    .home-status-cell span {{ color:var(--ink-muted); font-size:12px; }}
    .home-status-cell strong {{ display:block; font-size:24px; margin-top:4px; }}
    .home-alert {{ display:grid; grid-template-columns:30px 1fr; gap:10px; padding:12px 0; border-bottom:1px solid var(--line-soft); }}
    .home-alert:last-child {{ border-bottom:0; }}
    .home-alert-icon {{ align-items:center; background:#edf5ff; border-radius:10px; color:var(--accent-ink); display:flex; font-weight:800; height:30px; justify-content:center; width:30px; }}
    .home-alert b {{ display:block; font-size:13px; line-height:1.35; }}
    .home-alert p {{ font-size:12px; line-height:1.45; margin:3px 0 0; }}
    .home-important {{ display:grid; grid-template-columns:1fr 1fr; gap:18px; }}
    .home-note-list {{ margin:10px 0 0; padding-left:20px; color:var(--ink-soft); }}
    .home-note-list li {{ margin:6px 0; }}
    .home-mini-grid {{ display:grid; grid-template-columns:repeat(3, minmax(0, 1fr)); gap:12px; }}
    .home-mini {{
      background:linear-gradient(180deg,#fff 0%, #f8fbff 100%); border:1px solid #cfe0f5; border-radius:18px;
      min-height:104px; padding:14px;
    }}
    .home-section-banner {{
      align-items:center; background:linear-gradient(100deg,#16386d 0%, #2f6fed 100%);
      border-radius:20px; color:#fff; display:flex; justify-content:space-between; gap:18px;
      min-height:80px; padding:18px 22px;
    }}
    .home-section-banner strong {{ align-items:center; background:rgba(255,255,255,.16); border-radius:14px; display:inline-flex; font-size:18px; height:42px; justify-content:center; margin-right:14px; width:42px; }}
    .home-section-banner h2 {{ align-items:center; color:#fff; display:inline-flex; font-size:24px; margin:0; }}
    .home-section-banner p {{ color:#dce9ff; margin:0; max-width:760px; }}
    .home-mini span {{ color:var(--ink-muted); font-size:12px; }}
    .home-mini strong {{ display:block; font-size:26px; margin:10px 0 4px; }}
    .home-quick-actions {{ display:flex; flex-wrap:wrap; gap:10px; margin-top:14px; }}
    @media (max-width:1200px) {{
      .home-board {{ grid-template-columns:300px 1fr; }}
      .home-stack.right {{ grid-column:1 / -1; display:grid; grid-template-columns:1fr 1fr; }}
      .home-module-grid {{ grid-template-columns:repeat(2, minmax(0, 1fr)); }}
      .home-important {{ grid-template-columns:1fr; }}
    }}
    @media (max-width:800px) {{
      .two, .editor-layout, .period-tabs, .home-top, .home-board, .home-stack.right, .home-module-grid, .home-status-grid, .home-mini-grid {{ grid-template-columns:1fr; }}
      main {{ padding:18px 14px 36px; }}
      header {{ padding:16px 18px; }}
      .home-status-cell {{ border-right:0; border-bottom:1px solid var(--line); }}
      .home-status-cell:last-child {{ border-bottom:0; }}
    }}
  </style>
</head>
<body>
<header>
  <h1>经销商PDCA工作台</h1>
  <p>{esc(header_subtitle())}</p>
</header>
<main>
{msg_html}
{body}
</main>
</body>
</html>"""


def route_url(route_path, date_text, **params):
    payload = {"date": date_text}
    payload.update(params)
    return f"{route_path}?{urlencode(payload)}"


def render_hermes_panel(date_text, result=""):
    result_html = ""
    if result:
        if isinstance(result, dict):
            state = "ok" if result.get("ok") else "warn"
            path = result.get("path")
            actions = ""
            preview = ""
            if path:
                primary_href = route_url("/view-path", date_text, path=path) if is_previewable(path) else route_url("/open-path", date_text, path=path)
                extra_links = "".join(
                    button(link.get("label", "打开官网"), link.get("url", ""), "secondary")
                    for link in result.get("links", [])
                )
                actions = (
                    f'<div class="actions">'
                    f'{button("查看结果", primary_href)}'
                    f'{button("用本机软件打开", route_url("/open-path", date_text, path=path), "light")}'
                    f'{button("打开所在目录", route_url("/open-path", date_text, path=str(Path(path).parent)), "light")}'
                    f'{extra_links}'
                    f'</div>'
                )
                preview = (
                    f'<div class="result-preview">{esc(result.get("content", ""))}</div>'
                    f'<div class="result-file">已生成：{esc(result.get("filename") or Path(path).name)}</div>'
                )
            else:
                preview = f'<div class="result-preview">{esc(result.get("content", ""))}</div>'
            result_html = f"""
            <div class="result-banner {state}">
              <h3>{"执行完成" if result.get("ok") else "执行失败"}</h3>
              <p>{esc("报告已生成，可直接在下方查看或点按钮打开。" if path and result.get("ok") else "请查看下方反馈。")}</p>
              {actions}
              {preview}
            </div>
            """
        else:
            result_html = f'<div class="hermes-result">{esc(result)}</div>'
    return f"""
    <section style="margin-top:16px">
      <h2>向 Hermes 派任务</h2>
      <p>这是日常唯一任务入口。数据类任务交给 Hermes 数据 Agent；物流单号会进入物流核查流程，避免生成空报告。</p>
      <form method="post" action="{esc(route_url('/hermes-chat', date_text))}" id="hermesTaskForm">
        <textarea name="query" placeholder="例如：从 VPS 拉 5 月经销商业绩，按团队汇总并生成 Excel 表格；或：查这些物流单号并判断是否异常"></textarea>
        <div class="actions">
          <button type="submit" id="hermesSubmitBtn">交给 Hermes 拆解执行</button>
          <span class="thinking-inline" id="hermesThinking"><span class="spinner"></span>thinking...</span>
        </div>
      </form>
      <script>
      (function() {{
        const form = document.getElementById('hermesTaskForm');
        const btn = document.getElementById('hermesSubmitBtn');
        const thinking = document.getElementById('hermesThinking');
        if (!form || !btn || !thinking) return;
        form.addEventListener('submit', function() {{
          btn.disabled = true;
          btn.textContent = '执行中...';
          thinking.classList.add('on');
        }});
      }})();
      </script>
      {result_html}
    </section>
    """


def render_tracking_cards(result):
    rows = result.get("tracking_results") if isinstance(result, dict) else None
    if not rows:
        return ""
    cards = []
    for row in rows:
        status = row.get("status", "")
        ok = "delivered" in status.lower() or "签收" in status
        color = "#178a4b" if ok else "#b26b00"
        bg = "#eefaf3" if ok else "#fff7e6"
        cards.append(f"""
        <div style="border:1px solid #d8e4d8;background:{bg};border-radius:14px;padding:16px;margin:12px 0">
          <div style="display:flex;align-items:flex-start;justify-content:space-between;gap:12px">
            <div>
              <div style="font-size:12px;color:#7b8496">Tracking Code</div>
              <div style="font-size:20px;font-weight:800;color:#111827">{esc(row.get("tracking_number", ""))}</div>
            </div>
            <div style="border-radius:999px;background:{color};color:white;padding:6px 12px;font-weight:700;font-size:13px">
              {esc(status)}
            </div>
          </div>
          <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px;margin-top:14px">
            <div><b>承运商</b><br>{esc(row.get("carrier", ""))}</div>
            <div><b>更新时间</b><br>{esc(row.get("last_update", "") or "未识别")}</div>
            <div><b>始发地</b><br>{esc(row.get("origin", "") or "未识别")}</div>
            <div><b>目的地</b><br>{esc(row.get("destination", "") or "未识别")}</div>
          </div>
          <div style="margin-top:14px;height:8px;background:#d8d8d8;border-radius:999px;overflow:hidden">
            <div style="height:100%;width:{'100' if ok else '65'}%;background:{color};border-radius:999px"></div>
          </div>
        </div>
        """)
    return "".join(cards)


def render_hermes_result_modal(date_text, result):
    if not result:
        return ""
    if not isinstance(result, dict):
        content = esc(result)
        return f"""
        <div class="result-modal" role="dialog" aria-modal="true">
          <div class="result-dialog">
            <div class="result-dialog-head">
              <div><h2>执行结果</h2><p>Hermes 已返回反馈。</p></div>
              <a class="result-close" href="{esc(route_url('/', date_text))}">×</a>
            </div>
            <div class="result-preview">{content}</div>
          </div>
        </div>
        """
    state = "ok" if result.get("ok") else "warn"
    path = result.get("path")
    actions = ""
    if path:
        primary_href = route_url("/view-path", date_text, path=path) if is_previewable(path) else route_url("/open-path", date_text, path=path)
        extra_links = "".join(
            button(link.get("label", "打开官网"), link.get("url", ""), "secondary")
            for link in result.get("links", [])
        )
        actions = (
            f'<div class="actions">'
            f'{button("查看结果", primary_href)}'
            f'{button("用本机软件打开", route_url("/open-path", date_text, path=path), "light")}'
            f'{button("打开所在目录", route_url("/open-path", date_text, path=str(Path(path).parent)), "light")}'
            f'{extra_links}'
            f'</div>'
        )
    title = "执行完成" if result.get("ok") else "执行失败"
    desc = "报告已生成，可直接查看或打开。" if path and result.get("ok") else "请查看下方反馈。"
    filename = result.get("filename") or (Path(path).name if path else "")
    file_html = f'<div class="result-file">已生成：{esc(filename)}</div>' if filename else ""
    tracking_html = render_tracking_cards(result)
    return f"""
    <div class="result-modal" role="dialog" aria-modal="true">
      <div class="result-dialog {state}">
        <div class="result-dialog-head">
          <div>
            <h2>{esc(title)}</h2>
            <p>{esc(desc)}</p>
          </div>
          <a class="result-close" href="{esc(route_url('/', date_text))}">×</a>
        </div>
        {actions}
        {tracking_html}
        <div class="result-preview">{esc(result.get("content", ""))}</div>
        {file_html}
      </div>
    </div>
    """


def render_agent_cards(date_text):
    cards = []
    avatars = ["🤖", "📦", "🔎"]
    for index, agent in enumerate(AGENT_CARDS):
        agent_key = agent["key"]
        ensure_agent_soul(agent)
        skill_count = len(list_agent_skills(agent))
        cards.append(f"""
        <div class="card agent-card">
          <div class="agent-avatar">{esc(avatars[index % len(avatars)])}</div>
          <h3>{esc(agent["title"])}</h3>
          <div class="agent-meta">{esc(agent["key"])}</div>
          <p>{esc(agent["desc"])}</p>
          <div class="actions">
            {button("+ 编辑 Agent", route_url("/agent-edit", date_text, agent=agent_key), "agent-button")}
          </div>
          <div class="agent-meta">Core files 5 · Skills {skill_count}</div>
        </div>
        """)
    return f"""
    <section class="agent-section" style="margin-top:16px">
      <h2>子 Agent 能力与维护</h2>
      <p>这些是真实存在的本地 Hermes profile 或项目 agent 定义。日常不要直接点子 Agent 派活，而是在上方告诉 Hermes，由 Hermes 选择谁来做。</p>
      <div class="agent-grid">{''.join(cards)}</div>
    </section>
    """


def render_agent_soul(date_text, agent_key, message=""):
    agent = agent_by_key(agent_key)
    if not agent:
        return page("未知 Agent", "<section><h2>未知 Agent</h2></section>", date_text, message)
    path = ensure_agent_soul(agent)
    content = read_text(path)
    title = agent["title"]
    body = f"""
    <section>
      <div class="page-toolbar">
        <div>
          <h2>编辑 SOUL.md：{esc(title)}</h2>
          <p>{esc(path)}</p>
        </div>
        {button("← 返回首页", route_url("/", date_text), "light")}
      </div>
      <form method="post" action="{esc(route_url('/agent-soul', date_text, agent=agent_key))}">
        <textarea name="content" style="min-height:420px">{esc(content)}</textarea>
        <div class="actions">
          <button type="submit">保存 SOUL.md</button>
          {button("返回首页", route_url("/", date_text), "light")}
        </div>
      </form>
    </section>
    """
    return page("编辑 SOUL.md", body, date_text, message)


def render_agent_edit(date_text, agent_key, active_file="SOUL.md", message=""):
    agent = agent_by_key(agent_key)
    if not agent:
        return page("未知 Agent", "<section><h2>未知 Agent</h2></section>", date_text, message)
    if active_file not in AGENT_CORE_FILES:
        active_file = "SOUL.md"
    path = ensure_agent_core_file(agent, active_file)
    content = read_text(path)
    nav = "".join(
        f'<a class="{"active" if name == active_file else ""}" href="{esc(route_url("/agent-edit", date_text, agent=agent_key, file=name))}">{esc(name)}</a>'
        for name in AGENT_CORE_FILES
    )
    skills = list_agent_skills(agent)
    skill_html = "".join(f'<span class="skill-chip">{esc(name)}</span>' for name in skills) or "暂无已安装 Skill"
    body = f"""
    <section>
      <div class="page-toolbar">
        <div>
          <h2>{esc(agent["title"])}</h2>
          <p>这里维护该 Agent 的核心文件和 Skill。保存后下一次 Hermes 调用会读取最新内容，不需要重启。</p>
        </div>
        {button("← 返回首页", route_url("/", date_text), "light")}
      </div>
      <div class="editor-layout">
        <div class="file-nav">
          <h3>Core Files</h3>
          {nav}
        </div>
        <div>
          <form method="post" action="{esc(route_url('/agent-core-file', date_text, agent=agent_key, file=active_file))}">
            <h3>{esc(active_file)}</h3>
            <textarea name="content" style="min-height:430px">{esc(content)}</textarea>
            <div class="actions">
              <button type="submit">保存</button>
              {button("返回首页", route_url("/", date_text), "light")}
            </div>
          </form>
        </div>
      </div>
    </section>
    <section style="margin-top:16px">
      <h2>Skill 热插拔</h2>
      <p>拖入或选择一个 `SKILL.md`，会立即安装到该 Agent 的 skills 目录。下一次 Hermes 执行任务时即可使用。</p>
      <form class="drop-zone" method="post" action="{esc(route_url('/agent-skill', date_text, agent=agent_key))}" enctype="multipart/form-data">
        <input type="file" name="skill" accept=".md,.txt" required style="width:100%">
        <div class="actions"><button type="submit">安装 Skill</button></div>
      </form>
      <p>{skill_html}</p>
    </section>
    """
    return page("编辑 Agent", body, date_text, message)


def output_result_card(title, icon, desc, exists, href, meta=""):
    if exists:
        action = button("打开", href)
        state = ""
    else:
        action = '<span class="button light">运行后生成</span>'
        state = " missing"
    return f"""
    <div class="output-card{state}">
      <span class="output-icon">{esc(icon)}</span>
      <h3>{esc(title)}</h3>
      <p>{esc(desc)}</p>
      {f'<p class="output-meta">{esc(meta)}</p>' if meta else ''}
      <div class="actions">{action}</div>
    </div>
    """


def render_output_panel(date_text, out, dashboard, workbook, report, pdca):
    latest_workbook = latest_output_file(date_text, "workbook")
    latest_report = latest_output_file(date_text, "report")
    latest_pdca = latest_output_file(date_text, "pdca")
    cards = "".join([
        output_result_card("Excel 表格", "📄", "每次打开最新生成的数据汇总 Excel，可直接发给业务使用。", bool(latest_workbook), route_url("/open", date_text, target="workbook"), f"最新：{file_time_label(latest_workbook)}"),
        output_result_card("数据报告", "🧾", "每次查看最新的数据来源、口径、团队汇总和风险说明。", bool(latest_report), route_url("/open", date_text, target="report"), f"最新：{file_time_label(latest_report)}"),
        output_result_card("PDCA 日结", "✅", "从「经销商-日报推送」群日报和 VPS 待办生成今日完成、进度、上级交办和明日计划。", True, route_url("/pdca-vps", date_text)),
    ])
    return f"""
    <section>
      <h2>今日输出</h2>
      <p>这些是 Hermes/PDCA 已经产出的结果，直接点开使用。路径只作为技术详情保留。</p>
      <div class="output-grid">{cards}</div>
      <details class="output-paths">
        <summary>查看文件路径</summary>
        <p>输出目录：<code>{esc(out)}</code></p>
        <p>最新 Excel：<code>{esc(latest_workbook or "运行后生成")}</code></p>
        <p>最新数据报告：<code>{esc(latest_report or "运行后生成")}</code></p>
        <p>旧版本地 PDCA 日结：<code>{esc(latest_pdca or "运行后生成")}</code></p>
      </details>
    </section>
    """


def render_pdca_vps(date_text, message=""):
    yesterday_text = previous_date_text(date_text)
    if should_use_local_pdca_cache(date_text):
        daily = local_daily_report_cache(date_text, "本地汇报缓存")
        yesterday = local_daily_report_cache(yesterday_text, "本地汇报缓存")
        month_okr = local_month_okr_cache(date_text, "本地汇报缓存")
        all_todos = {"ok": True, "rows": local_todo_payload_rows(date_text), "count": len(local_todo_payload_rows(date_text)), "error": ""}
    else:
        daily = fetch_vps_daily_report(date_text)
        yesterday = fetch_vps_daily_report(yesterday_text)
        month_okr = fetch_vps_month_okr(date_text)
        all_todos = fetch_vps_all_todos()
    today_plan_rows = report_payload_items(yesterday, "tomorrow") if yesterday["ok"] else []
    yesterday_done_rows = report_payload_items(yesterday, "today") if yesterday["ok"] else []
    today_done_rows = report_payload_items(daily, "today") if daily["ok"] else []
    delivery_checks = build_delivery_checks(
        today_plan_rows,
        today_done_rows,
        all_todos["rows"] if all_todos["ok"] else [],
        date_text,
    )
    user = daily.get("identity", {}).get("user", {})
    if daily["ok"]:
        report_rows = daily_report_table_rows(daily["reports"]) or '<tr><td colspan="3">未查询到「经销商-日报推送」群的今日日报，建议补交或确认 IM 日报是否同步。</td></tr>'
        report_status = "已同步" if daily["reports"] else "未查询到今日日报"
    else:
        report_rows = f'<tr><td colspan="3">VPS 日报拉取失败：{esc(daily["error"])}</td></tr>'
        report_status = "拉取失败"
    if yesterday["ok"]:
        today_rows = pdca_todo_rows(today_plan_rows) or '<tr><td colspan="4">「经销商-日报推送」群昨日日报未写入今日计划。</td></tr>'
        done_rows = pdca_todo_rows(yesterday_done_rows) or '<tr><td colspan="4">「经销商-日报推送」群昨日日报未写入完成事项。</td></tr>'
    else:
        today_rows = f'<tr><td colspan="4">「经销商-日报推送」群昨日日报拉取失败：{esc(yesterday["error"])}</td></tr>'
        done_rows = f'<tr><td colspan="4">「经销商-日报推送」群昨日日报拉取失败：{esc(yesterday["error"])}</td></tr>'
    if month_okr["ok"]:
        okr_table = okr_rows(month_okr["rows"]) or '<tr><td colspan="3">VPS 暂无本月 OKR/月待办数据。</td></tr>'
    else:
        okr_table = f'<tr><td colspan="3">VPS OKR 拉取失败：{esc(month_okr["error"])}</td></tr>'
    source_note = pdca_vps_source_note(daily, yesterday, month_okr, all_todos)
    cache_banner = ""
    if daily.get("from_cache") or yesterday.get("from_cache") or month_okr.get("from_cache"):
        cache_banner = f'<p class="message" style="border-color:#efc7c3;background:#fff6f4;color:#8a3b2f;">{esc(source_note)}</p>'
    else:
        cache_banner = f'<p class="message">{esc(source_note)}</p>'
    body = f"""
    <section>
      <div class="page-toolbar">
        <div>
          <h2>PDCA 日结（VPS）</h2>
          <p>来源：「经销商-日报推送」IM 群日报 + VPS OKR。今日计划优先取昨日群日报里的“明日计划”，月待办取本月 OKR。</p>
          {cache_banner}
        </div>
        {button("返回首页", route_url("/", date_text), "light")}
      </div>
      <div class="grid">
        {metric_card("群日报状态", report_status, f"经销商-日报推送 / {date_text}", daily["ok"] and bool(daily["reports"]))}
        {metric_card("今日预计待办", f"{len(today_plan_rows)} 项", f"来自 {yesterday_text} 群日报的明日计划", bool(today_plan_rows))}
        {metric_card("交付检查", f"{sum(1 for item in delivery_checks if item['level'] == 'done')} / {len(delivery_checks)}", "已交付 / 今日计划", bool(delivery_checks) and all(item["level"] != "risk" for item in delivery_checks))}
        {metric_card("本月 OKR/月待办", f"{month_okr.get('count', 0)} 项", "来自 VPS OKR employee-okr-list", month_okr["ok"] and month_okr.get("count", 0) > 0)}
      </div>
    </section>
    {render_delivery_agent(delivery_checks, daily["ok"] and bool(daily["reports"]))}
    <section>
      <h2>今日群日报记录（经销商-日报推送）</h2>
      <table><tr><th>提交时间</th><th>状态</th><th>内容摘要</th></tr>{report_rows}</table>
    </section>
    <section>
      <h2>今日预计待办（来自昨日「经销商-日报推送」群日报）</h2>
      <table><tr><th>事项</th><th>状态</th><th>截止</th><th>进度</th></tr>{today_rows}</table>
    </section>
    <section>
      <h2>昨天完成与进度（来自「经销商-日报推送」群日报）</h2>
      <table><tr><th>事项</th><th>状态</th><th>截止</th><th>进度</th></tr>{done_rows}</table>
    </section>
    <section>
      <h2>本月 OKR / 月待办</h2>
      <table><tr><th>目标</th><th>得分</th><th>KR 数</th></tr>{okr_table}</table>
    </section>
    """
    return page("PDCA 日结（VPS）", body, date_text, message)


def fetch_home_vps_summary():
    results = {}

    def load_im():
        results["im"] = fetch_vps_im_unread(with_latest=False)

    def load_todos():
        results["todos"] = fetch_vps_today_todos()

    threads = [threading.Thread(target=load_im), threading.Thread(target=load_todos)]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=0.8)
    return (
        results.get("im") or {"ok": False, "channels": [], "channel_count": 0, "unread_count": 0, "error": "IM 请求超时"},
        results.get("todos") or {"ok": False, "rows": [], "count": 0, "error": "待办请求超时"},
    )


def home_module_card(title, value, detail, href, progress=50):
    return f"""
    <a class="home-module" href="{esc(href)}">
      <b>{esc(title)}</b>
      <span>{esc(detail)}</span>
      <strong>{esc(value)}</strong>
      <div class="home-progress"><i style="width:{max(0, min(100, int(progress)))}%"></i></div>
    </a>
    """


def home_todo_cards(rows, date_text):
    if not rows:
        return """
        <div class="home-todo">
          <span class="home-check"></span>
          <div><b>暂无今日待办</b><small>VPS 没有返回待处理事项</small></div>
          <span class="home-chip ok">正常</span>
        </div>
        """
    cards = []
    for row in rows[:5]:
        priority = first_text(row, "priority", "priority_name") or "普通"
        title = first_text(row, "title", "name") or "未命名事项"
        status = first_text(row, "status_name", "status.name", "stage_name") or "待处理"
        deadline = first_text(row, "deadline", "due_date", "date_deadline") or date_text
        chip_class = "warn" if any(word in priority for word in ("高", "紧急", "High", "high")) else ""
        cards.append(f"""
        <div class="home-todo">
          <span class="home-check"></span>
          <div><b>{esc(title)}</b><small>{esc(status)} · 截止 {esc(deadline)}</small></div>
          <span class="home-chip {chip_class}">{esc(priority)}</span>
        </div>
        """)
    return "".join(cards)


def render_home(date_text, message="", hermes_result=None):
    out = output_dir(date_text)
    report = out / "data_summary_report.md"
    dashboard = out / "dashboard.html"
    workbook = out / f"{date_text}_data_summary.xlsx"
    pdca = out / "pdca_daily_check.md"
    latest_workbook = latest_output_file(date_text, "workbook")
    latest_report = latest_output_file(date_text, "report")
    latest_pdca = latest_output_file(date_text, "pdca")
    im_unread, today_todos = fetch_home_vps_summary()
    unread_im_count = im_unread["unread_count"]
    unread_channel_count = im_unread["channel_count"]
    today_todo_count = today_todos["count"]
    if today_todos["ok"]:
        todo_rows = todo_table_rows(today_todos["rows"][:8]) or '<tr><td colspan="4">VPS 暂无今日待办</td></tr>'
        todo_cards = home_todo_cards(today_todos["rows"], date_text)
    else:
        todo_rows = f'<tr><td colspan="4">VPS 待办拉取失败：{esc(today_todos["error"])}</td></tr>'
        todo_cards = f"""
        <div class="home-todo">
          <span class="home-check"></span>
          <div><b>VPS 待办拉取失败</b><small>{esc(today_todos["error"][:80])}</small></div>
          <span class="home-chip warn">异常</span>
        </div>
        """
    dashboard_state = "已生成" if dashboard.exists() else "待生成"
    output_count = sum(1 for item in (latest_workbook, latest_report, latest_pdca) if item)
    issue_count = (0 if im_unread["ok"] else 1) + (0 if today_todos["ok"] else 1) + (0 if dashboard.exists() else 1)
    health_text = "正常" if issue_count == 0 else f"{issue_count} 项待处理"
    body = f"""
    <div class="workbench-home">
      <div class="period-tabs" aria-label="周期切换">
        <a class="period-tab active" href="{esc(route_url('/', date_text))}">日</a>
        <a class="period-tab" href="{esc(route_url('/', date_text))}">周</a>
        <a class="period-tab" href="{esc(route_url('/', date_text))}">月</a>
        <a class="period-tab" href="{esc(route_url('/', date_text))}">季度</a>
        <a class="period-tab" href="{esc(route_url('/', date_text))}">年度</a>
      </div>

      <div class="home-top">
        <div class="home-metric">
          <div class="home-metric-row"><span class="home-metric-kicker">负责人</span><span class="home-chip">Sam</span></div>
          <div><div class="home-metric-value">{today_todo_count}</div><p>今日待办事项</p></div>
        </div>
        <div class="home-metric">
          <div class="home-metric-row"><span class="home-metric-kicker">业务入口</span><span class="home-chip ok">4 个模块</span></div>
          <div><div class="home-metric-value">{output_count}/3</div><p>今日输出物完成度</p></div>
        </div>
        <div class="home-metric">
          <div class="home-metric-row"><span class="home-metric-kicker">事务状态</span><span class="home-chip {'ok' if issue_count == 0 else 'warn'}">{esc(health_text)}</span></div>
          <div><div class="home-metric-value">{unread_im_count}</div><p>IM 未读消息 · {unread_channel_count} 个会话</p></div>
        </div>
      </div>

      <div class="home-section-banner">
        <h2><strong>01</strong>行动闭环</h2>
        <p>把 AI 提醒、OKR 拆解和日常管理动作沉淀为个人任务中心，员工可自行标记状态、维护备注和完成进度。</p>
      </div>

      <div class="home-board">
        <div class="home-stack">
          <section class="home-panel">
            <div class="home-panel-head">
              <h2>今日待办</h2>
              <a class="button light" href="{esc(route_url('/todos', date_text))}">More</a>
            </div>
            <div class="home-panel-body home-todo-list">{todo_cards}</div>
          </section>

          <section class="home-panel">
            <div class="home-panel-head"><h2>待分析消息</h2><span class="home-chip {'ok' if unread_im_count == 0 else 'warn'}">{unread_im_count} 条</span></div>
            <div class="home-panel-body">
              <p>{esc("IM 暂无未读会话" if unread_im_count == 0 else f"来自 {unread_channel_count} 个会话，需要判断是否转任务或转 Hermes。")}</p>
              <div class="home-quick-actions">{button("打开 IM 未读", route_url("/im-unread", date_text), "light")}</div>
            </div>
          </section>
        </div>

        <section class="home-panel">
          <div class="home-panel-head">
            <h2>业务进度</h2>
            <span class="home-chip">{esc(date_text)}</span>
          </div>
          <div class="home-panel-body">
            <div class="home-module-grid">
              {home_module_card("数据看板", dashboard_state, "日报、业务指标与风险摘要", route_url("/dashboard", date_text), 80 if dashboard.exists() else 30)}
              {home_module_card("客户管理", "CRM", "经销商客户台账、拜访与回款", "/customer-mgmt", 72)}
              {home_module_card("客流分析", "Sell Out", "海外客流与线上经营（代理商终销、OKR、渠道线索）", "/walkin-cockpit/", 72)}
            </div>
            <div class="home-status-grid">
              <div class="home-status-cell"><span>数据报告</span><strong>{'1' if latest_report else '0'}</strong></div>
              <div class="home-status-cell"><span>Excel 表格</span><strong>{'1' if latest_workbook else '0'}</strong></div>
              <div class="home-status-cell"><span>PDCA 日结</span><strong>{'1' if latest_pdca else '0'}</strong></div>
            </div>
          </div>
        </section>

        <div class="home-stack right">
          <section class="home-panel">
            <div class="home-panel-head"><h2>异常情况</h2><span class="home-chip {'ok' if issue_count == 0 else 'warn'}">{issue_count}</span></div>
            <div class="home-panel-body">
              <div class="home-alert">
                <div class="home-alert-icon">!</div>
                <div><b>{esc("首页服务正常" if issue_count == 0 else "存在待处理事项")}</b><p>{esc("VPS、看板与输出入口均可继续使用。" if issue_count == 0 else "优先检查 VPS 拉取状态和今日看板是否生成。")}</p></div>
              </div>
              <div class="home-alert">
                <div class="home-alert-icon">i</div>
                <div><b>当前端口 8767</b><p>首页与客流分析台（含线上经营）由 pdca_workbench.py 托管。</p></div>
              </div>
            </div>
          </section>

          <section class="home-panel">
            <div class="home-panel-head"><h2>服务健康</h2><span class="home-chip ok">在线</span></div>
            <div class="home-panel-body">
              <div class="home-alert"><div class="home-alert-icon">D</div><div><b>数据看板</b><p>{esc("今日 dashboard.html 已生成" if dashboard.exists() else "今日 dashboard.html 尚未生成")}</p></div></div>
              <div class="home-alert"><div class="home-alert-icon">V</div><div><b>VPS 摘要</b><p>{esc("待办和 IM 摘要已返回" if im_unread["ok"] and today_todos["ok"] else "部分 VPS 摘要拉取失败")}</p></div></div>
            </div>
          </section>
        </div>
      </div>

      <div class="home-section-banner">
        <h2><strong>02</strong>行政事务</h2>
        <p>把合同、审批、资料、报销和跨部门协同集中处理，确保行政节点不拖慢销售动作和 OKR 执行。</p>
      </div>

      <div class="home-important">
        <section class="home-panel">
          <div class="home-panel-head"><h2>重要事项</h2><span class="home-chip">业务</span></div>
          <div class="home-panel-body">
            <ul class="home-note-list">
              <li>先确认今日看板与 Excel 是否生成，缺失时运行 PDCA 日跑。</li>
              <li>今日待办需要和 IM 消息闭环，避免只看数据不落行动。</li>
              <li>点击 Sell Out 或「客流分析」进入海外客流与线上经营合一的分析台。</li>
            </ul>
            <div class="home-quick-actions">
              {button("运行今日 PDCA", route_url("/run", date_text))}
              {button("打开今日输出", route_url("/open", date_text, target="report"), "light")}
            </div>
          </div>
        </section>

        <section class="home-panel">
          <div class="home-panel-head"><h2>处理建议</h2><span class="home-chip">事务</span></div>
          <div class="home-panel-body">
            <ul class="home-note-list">
              <li>有未读 IM 时，先判断是否需要派给 Hermes 生成报告或动作清单。</li>
              <li>代理商名单和客流指标只从 JSON 数据包读取，更新 Excel 后需重跑构建脚本。</li>
              <li>改完首页样式后重启服务，并在浏览器 Ctrl+F5 强制刷新。</li>
            </ul>
          </div>
        </section>
      </div>

      <section class="home-panel">
        <div class="home-panel-head"><h2>任务中心</h2><a class="button light" href="{esc(route_url('/pdca-vps', date_text))}">PDCA 日结</a></div>
        <div class="home-panel-body home-mini-grid">
          <div class="home-mini"><span>总任务数</span><strong>{today_todo_count}</strong><p>来自 VPS 今日待办</p></div>
          <div class="home-mini"><span>已完成</span><strong>{max(0, output_count)}</strong><p>今日已生成输出物</p></div>
          <div class="home-mini"><span>未完成</span><strong>{max(0, today_todo_count - output_count)}</strong><p>待继续跟进事项</p></div>
        </div>
      </section>

      <section class="home-panel">
        <div class="home-panel-head"><h2>会议中心</h2><span class="home-chip">复盘</span></div>
        <div class="home-panel-body home-mini-grid">
          <div class="home-mini"><span>待确认</span><strong>{unread_channel_count}</strong><p>可从 IM 会话转入</p></div>
          <div class="home-mini"><span>业务复盘</span><strong>1</strong><p>客流分析台</p></div>
          <div class="home-mini"><span>输出闭环</span><strong>{output_count}</strong><p>报告、Excel、PDCA</p></div>
        </div>
      </section>

      <section class="home-panel">
        <div class="home-panel-head"><h2>今日待办明细（VPS）</h2><a class="button light" href="{esc(route_url('/todos', date_text))}">查看全部</a></div>
        <div class="home-panel-body"><table><tr><th>优先级</th><th>事项</th><th>状态</th><th>截止</th></tr>{todo_rows}</table></div>
      </section>

      {render_hermes_panel(date_text)}
      {render_agent_cards(date_text)}
      {render_output_panel(date_text, out, dashboard, workbook, report, pdca)}
      {render_hermes_result_modal(date_text, hermes_result)}
    </div>
    """
    return page("数据岗位 PDCA 工作台", body, date_text, message)


def render_questionnaire(date_text, message=""):
    values = parse_questionnaire(date_text)
    fields = []
    for index, title in enumerate(QUESTION_TITLES):
        content = values.get(title, "")
        if content == "-":
            content = ""
        fields.append(f"<label><h3>{esc(title)}</h3><textarea name='q{index}'>{esc(content)}</textarea></label>")
    body = f"""
    <section>
      <div class="page-toolbar">
        <div><h2>填写每日问卷</h2></div>
        {button("← 返回首页", route_url("/", date_text), "light")}
      </div>
      <form method="post" action="{esc(route_url('/questionnaire', date_text))}">
        {''.join(fields)}
        <div class="actions"><button type="submit">保存问卷</button>{button("返回首页", route_url("/", date_text), "light")}</div>
      </form>
    </section>
    """
    return page("填写每日问卷", body, date_text, message)


def render_todos(date_text, message=""):
    today_todos = fetch_vps_today_todos()
    if today_todos["ok"]:
        table = todo_table_rows(today_todos["rows"]) or '<tr><td colspan="4">VPS 暂无今日待办</td></tr>'
    else:
        table = f'<tr><td colspan="4">VPS 待办拉取失败：{esc(today_todos["error"])}</td></tr>'
    body = f"""
    <section>
      <div class="page-toolbar">
        <div>
          <h2>今日待办（VPS）</h2>
          <p>来源：vertu odoo project todo list --for-me --due-within-days 0</p>
        </div>
        {button("← 返回首页", route_url("/", date_text), "light")}
      </div>
      <table><tr><th>优先级</th><th>事项</th><th>状态</th><th>截止</th></tr>{table}</table>
      <div class="actions">{button("返回首页", route_url("/", date_text), "light")}</div>
    </section>
    """
    return page("今日待办", body, date_text, message)


def render_im_unread(date_text, message=""):
    im_unread = fetch_vps_im_unread()
    if im_unread["ok"]:
        table = im_table_rows(im_unread["channels"], date_text) or '<tr><td colspan="4">VPS 暂无未读 IM</td></tr>'
        summary = f"合计 {im_unread['unread_count']} 条未读，分布在 {im_unread['channel_count']} 个会话。"
    else:
        table = f'<tr><td colspan="4">VPS IM 拉取失败：{esc(im_unread["error"])}</td></tr>'
        summary = "VPS IM 拉取失败。"
    body = f"""
    <section>
      <div class="page-toolbar">
        <div>
          <h2>IM 未读消息（VPS）</h2>
          <p>{esc(summary)}</p>
        </div>
        {button("← 返回首页", route_url("/", date_text), "light")}
      </div>
      <table><tr><th>会话</th><th>未读数</th><th>最新消息</th><th>最后活跃</th></tr>{table}</table>
      <div class="actions">{button("返回首页", route_url("/", date_text), "light")}</div>
    </section>
    """
    return page("IM 未读消息", body, date_text, message)


def render_logistics(date_text, message=""):
    rows = read_csv_rows(logistics_path(date_text))
    table = "".join(
        f"<tr><td>{esc(row.get('tracking_number'))}</td><td>{esc(row.get('carrier'))}</td><td>{esc(row.get('customer'))}</td><td>{esc(row.get('salesperson'))}</td><td>{esc(row.get('current_status'))}</td></tr>"
        for row in rows
    ) or '<tr><td colspan="5">暂无物流单号</td></tr>'
    body = f"""
    <section>
      <div class="page-toolbar">
        <div><h2>录入物流单号</h2></div>
        {button("← 返回首页", route_url("/", date_text), "light")}
      </div>
      <form method="post" action="{esc(route_url('/logistics', date_text))}">
        <div class="two">
          <label>物流单号<input name="tracking_number" required style="width:100%"></label>
          <label>承运商<select name="carrier" style="width:100%"><option>UPS</option><option>FedEx</option><option>DHL</option><option>SF</option></select></label>
          <label>客户<input name="customer" style="width:100%"></label>
          <label>销售<input name="salesperson" style="width:100%"></label>
          <label>发货日期<input type="date" name="ship_date" value="{esc(date_text)}" style="width:100%"></label>
          <label>当前状态<input name="current_status" placeholder="不知道可留空" style="width:100%"></label>
          <label>预期状态<input name="expected_status" style="width:100%"></label>
          <label>备注<input name="note" style="width:100%"></label>
        </div>
        <div class="actions"><button type="submit">保存物流单号</button>{button("返回首页", route_url("/", date_text), "light")}</div>
      </form>
    </section>
    <section style="margin-top:16px">
      <h2>当前物流</h2>
      <table><tr><th>单号</th><th>承运商</th><th>客户</th><th>销售</th><th>当前状态</th></tr>{table}</table>
    </section>
    """
    return page("录入物流单号", body, date_text, message)


def open_target(date_text, target):
    path = latest_output_file(date_text, target)
    if path and path.exists():
        os.startfile(path)
        return f"已打开：{path}"
    return "文件还不存在，请先运行今日 PDCA。"


def open_path(path_text):
    path = Path(path_text)
    try:
        resolved = path.resolve()
        if not resolved.exists():
            return "文件还不存在。"
        os.startfile(resolved)
        return f"已打开：{resolved}"
    except OSError as exc:
        return f"打开失败：{exc}"


def render_view_path(date_text, path_text, back_url=""):
    path = Path(path_text)
    back_href = back_url or route_url("/", date_text)
    back_label = "← 返回"
    if not path.exists():
        return page("查看结果", f"""
        <section>
          <div class="page-toolbar">
            <div><h2>结果文件不存在</h2><p>可能是文件路径已变化，请重新执行 Hermes 任务。</p></div>
            {button(back_label, back_href, "light")}
          </div>
        </section>
        """, date_text)
    if path.suffix.lower() in {".html", ".htm"}:
        return read_text(path)
    content = read_text(path)
    return page("查看结果", f"""
    <section>
      <div class="page-toolbar">
        <div>
          <h2>Hermes 执行结果</h2>
          <p>{esc(path.name)}</p>
        </div>
        <div class="actions">
          {button("用本机软件打开", route_url("/open-path", date_text, path=str(path)))}
          {button(back_label, back_href, "light")}
        </div>
      </div>
      <div class="result-preview">{esc(content)}</div>
    </section>
    """, date_text)


def open_im_channel(channel_id):
    if not channel_id:
        return "缺少 IM 会话 ID。"
    url = im_channel_url(channel_id)
    os.startfile(url)
    return f"已打开 IM 会话：{url}"


class WorkbenchHandler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        return

    def date_from_query(self, query):
        return parse_qs(query).get("date", [today_text()])[0] or today_text()

    def read_form(self):
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length).decode("utf-8", errors="replace")
        return parse_qs(raw)

    def read_multipart(self):
        import cgi  # 独立 HTTP 服务模式专用；Python 3.13 移除该模块，故延迟到用到时才导入
        environ = {
            "REQUEST_METHOD": "POST",
            "CONTENT_TYPE": self.headers.get("Content-Type", ""),
            "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
        }
        return cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ=environ)

    def send_html(self, content):
        encoded = content.encode("utf-8")
        try:
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
            self.send_header("Pragma", "no-cache")
            self.send_header("Content-Length", str(len(encoded)))
            self.end_headers()
            self.wfile.write(encoded)
        except (BrokenPipeError, ConnectionAbortedError, ConnectionResetError, OSError):
            return

    def send_json(self, payload, status=200):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        try:
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except (BrokenPipeError, ConnectionAbortedError, ConnectionResetError, OSError):
            return

    def send_file(self, path):
        content_type = "text/html; charset=utf-8"
        self.send_bytes(path.read_bytes(), content_type)

    def send_bytes(self, content, content_type):
        try:
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
            self.send_header("Pragma", "no-cache")
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
        except (BrokenPipeError, ConnectionAbortedError, ConnectionResetError, OSError):
            return

    def serve_cockpit_module(self, mount, parsed_path, resolve_asset, date_text=""):
        """静态驾驶舱模块：/walkin-cockpit/、/online-cockpit/ 等。"""
        if parsed_path == mount:
            self.send_response(302)
            self.send_header("Location", f"{mount}/")
            self.end_headers()
            return
        rel = parsed_path[len(mount) + 1 :]
        asset = resolve_asset(rel)
        if not asset:
            self.send_response(404)
            self.end_headers()
            return
        guessed, _ = mimetypes.guess_type(str(asset))
        content_type = guessed or "application/octet-stream"
        if asset.suffix.lower() in {".js", ".mjs"}:
            content_type = "application/javascript; charset=utf-8"
        elif asset.suffix.lower() == ".json":
            content_type = "application/json; charset=utf-8"
        elif asset.suffix.lower() in {".html", ".htm"}:
            title = "经销商海外客流分析台" if "walkin" in mount else "经销商线上经营"
            html = skin_cockpit_html(asset.read_text(encoding="utf-8"), date_text, title)
            self.send_html(html)
            return
        self.send_bytes(asset.read_bytes(), content_type)

    def serve_walkin_cockpit(self, parsed_path, date_text=""):
        self.serve_cockpit_module("/walkin-cockpit", parsed_path, resolve_walkin_asset, date_text)

    def serve_online_cockpit(self, parsed_path, date_text=""):
        self.serve_cockpit_module("/online-cockpit", parsed_path, resolve_online_asset, date_text)

    def serve_meeting_center(self, parsed_path, date_text=""):
        if parsed_path == "/meeting-center":
            self.send_response(302)
            self.send_header("Location", f"/meeting-center/?date={date_text}")
            self.end_headers()
            return
        rel = parsed_path[len("/meeting-center/") :]
        if not rel:
            rel = "index.html"
        asset = resolve_meeting_center_asset(rel)
        if not asset:
            self.send_response(404)
            self.end_headers()
            return
        if asset.suffix.lower() in {".html", ".htm"}:
            html = skin_cockpit_html(asset.read_text(encoding="utf-8"), date_text, "会议中心")
            self.send_html(html)
            return
        guessed, _ = mimetypes.guess_type(str(asset))
        self.send_bytes(asset.read_bytes(), guessed or "application/octet-stream")

    def send_walkin_api(self, query, date_text):
        """Walk-in 数据：VPS > Excel 参考 JSON > mock。"""
        month = (query.get("month", [""])[0] or "").strip()
        if not re.fullmatch(r"\d{4}-\d{2}", month):
            month = (date_text or today_text())[:7]
        if build_walkin_api_payload is None:
            self.send_json({"error": "workbench_data module missing"}, status=500)
            return
        try:
            payload = build_walkin_api_payload(month, date_text or today_text())
            self.send_json(payload)
        except Exception as exc:
            self.send_json({"error": str(exc)}, status=500)

    def send_online_channel_api(self, date_text):
        """线上 OKR 表：大区东南亚/欧洲，真实销售来自 vertu 经销商业绩 JSON。"""
        if build_online_channel_payload is None:
            self.send_json({"error": "workbench_data module missing"}, status=500)
            return
        try:
            payload = build_online_channel_payload(date_text or today_text())
            self.send_json(payload)
        except Exception as exc:
            self.send_json({"error": str(exc)}, status=500)

    def send_error_page(self, exc):
        try:
            self.send_html(page("工作台异常", f"""
            <section>
              <h2>工作台遇到问题</h2>
              <p>这次请求没有完成，但服务没有崩。请刷新后重试。</p>
              <div class="message">{esc(exc)}</div>
              <div class="actions">{button("返回首页", route_url("/", today_text()), "light")}</div>
            </section>
            """, today_text()))
        except (BrokenPipeError, ConnectionAbortedError, ConnectionResetError, OSError):
            return
        except Exception:
            try:
                self.send_response(500)
                self.end_headers()
            except (BrokenPipeError, ConnectionAbortedError, ConnectionResetError, OSError):
                return

    def redirect(self, path, date_text, message=""):
        params = {"date": date_text}
        if message:
            params["message"] = message
        self.send_response(303)
        self.send_header("Location", f"{path}?{urlencode(params)}")
        self.end_headers()

    def do_GET(self):
        try:
            self._do_GET()
        except Exception as exc:
            self.send_error_page(exc)

    def _do_GET(self):
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        date_text = self.date_from_query(parsed.query)
        message = query.get("message", [""])[0]
        if parsed.path == "/":
            serve_home_dashboard_index(self)
        elif parsed.path == "/home-classic":
            self.send_html(render_home(date_text, message))
        elif parsed.path.startswith("/api/dashboard/") or parsed.path.startswith("/api/todos/") or parsed.path.startswith("/api/hermes-agent/") or parsed.path.startswith("/api/customer-center/") or parsed.path.startswith("/api/hr/") or parsed.path == "/api/exceptions" or parsed.path.startswith("/api/important-matters") or parsed.path.startswith("/api/task-center/") or parsed.path.startswith("/api/meeting-center/"):
            data = dispatch_home_dashboard_api(parsed.path, query)
            if data is None:
                self.send_response(404)
                self.end_headers()
            else:
                self.send_json(data)
        elif parsed.path == "/questionnaire":
            self.send_html(render_questionnaire(date_text, message))
        elif parsed.path == "/todos":
            self.send_html(render_todos(date_text, message))
        elif parsed.path == "/im-unread":
            self.send_html(render_im_unread(date_text, message))
        elif parsed.path == "/pdca-vps":
            self.send_html(render_pdca_vps(date_text, message))
        elif parsed.path == "/meeting-center" or parsed.path.startswith("/meeting-center/"):
            self.serve_meeting_center(parsed.path, date_text)
        elif parsed.path == "/agent-soul":
            agent_key = query.get("agent", [""])[0]
            self.send_html(render_agent_soul(date_text, agent_key, message))
        elif parsed.path == "/agent-edit":
            agent_key = query.get("agent", [""])[0]
            active_file = query.get("file", ["SOUL.md"])[0]
            self.send_html(render_agent_edit(date_text, agent_key, active_file, message))
        elif parsed.path == "/logistics":
            self.send_html(render_logistics(date_text, message))
        elif parsed.path == "/dashboard":
            start_date = query.get("start", [""])[0]
            end_date = query.get("end", [""])[0]
            if start_date and end_date:
                date_text = end_date
                code, stdout, stderr = run_pdca(date_text, push=False, start_date=start_date)
                if code != 0:
                    self.redirect("/", date_text, f"区间看板生成失败：{stderr or stdout}"[:300])
                    return
            dashboard = output_dir(date_text) / "dashboard.html"
            if not dashboard.exists():
                code, stdout, stderr = run_pdca(date_text, push=False)
                if code != 0:
                    self.redirect("/", date_text, f"当天看板生成失败：{stderr or stdout}"[:300])
                    return
            if dashboard.exists():
                serve_dashboard_html(self, dashboard, date_text)
            else:
                self.redirect("/", date_text, "这个日期还没有看板，请先运行当天 PDCA。")
        elif parsed.path == "/dashboard-theme.css":
            if DASHBOARD_THEME_CSS.is_file():
                self.send_bytes(DASHBOARD_THEME_CSS.read_bytes(), "text/css; charset=utf-8")
            else:
                self.send_response(404)
                self.end_headers()
        elif parsed.path == "/workbench-cockpit-shell.css":
            if COCKPIT_SHELL_CSS.is_file():
                self.send_bytes(COCKPIT_SHELL_CSS.read_bytes(), "text/css; charset=utf-8")
            else:
                self.send_response(404)
                self.end_headers()
        elif parsed.path == "/open":
            target = query.get("target", [""])[0]
            self.redirect("/", date_text, open_target(date_text, target))
        elif parsed.path == "/open-path":
            path_text = query.get("path", [""])[0]
            self.redirect("/", date_text, open_path(path_text))
        elif parsed.path == "/view-path":
            path_text = query.get("path", [""])[0]
            back_url = query.get("from", [""])[0]
            self.send_html(render_view_path(date_text, path_text, back_url))
        elif parsed.path == "/customer-mgmt":
            err = ensure_customer_server()
            if err:
                self.redirect("/", date_text, err)
            else:
                self.send_html(render_customer_mgmt_frame(date_text))
        elif parsed.path == "/open-im-channel":
            channel_id = query.get("channel_id", [""])[0]
            self.redirect("/im-unread", date_text, open_im_channel(channel_id))
        elif parsed.path == "/walkin-cockpit" or parsed.path.startswith("/walkin-cockpit/"):
            self.serve_walkin_cockpit(parsed.path, date_text)
        elif parsed.path == "/online-cockpit" or parsed.path.startswith("/online-cockpit/"):
            self.serve_online_cockpit(parsed.path, date_text)
        elif parsed.path == "/api/walkin":
            self.send_walkin_api(query, date_text)
        elif parsed.path == "/api/online-channel":
            self.send_online_channel_api(date_text)
        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        try:
            self._do_POST()
        except Exception as exc:
            self.send_error_page(exc)

    def _do_POST(self):
        parsed = urlparse(self.path)
        date_text = self.date_from_query(parsed.query)
        if parsed.path == "/api/agent/process-suggestion":
            length = int(self.headers.get("Content-Length", 0) or 0)
            raw = self.rfile.read(length) if length else b"{}"
            try:
                payload = json.loads(raw.decode("utf-8"))
            except json.JSONDecodeError:
                payload = {}
            self.send_json({"ok": True, "message": "建议已记录", "id": payload.get("id")})
            return
        if parsed.path == "/api/meeting-center/dispatch":
            length = int(self.headers.get("Content-Length", 0) or 0)
            raw = self.rfile.read(length) if length else b"{}"
            try:
                payload = json.loads(raw.decode("utf-8"))
            except json.JSONDecodeError:
                self.send_json({"ok": False, "error": "请求体不是 JSON"}, status=400)
                return
            result = api_meeting_center_dispatch(payload, (payload.get("date") or date_text))
            self.send_json(result)
            return
        if parsed.path == "/agent-skill":
            query = parse_qs(parsed.query)
            agent_key = query.get("agent", [""])[0]
            try:
                fields = self.read_multipart()
                uploaded = fields["skill"] if "skill" in fields else None
                if uploaded is None or not getattr(uploaded, "filename", ""):
                    raise ValueError("没有收到 skill 文件。")
                target = install_skill_to_agent(agent_key, uploaded.filename, uploaded.file.read())
                self.send_response(303)
                self.send_header("Location", f"/agent-edit?{urlencode({'date': date_text, 'agent': agent_key, 'message': f'Skill 已安装：{target}'})}")
                self.end_headers()
            except Exception as exc:
                self.send_response(303)
                self.send_header("Location", f"/agent-edit?{urlencode({'date': date_text, 'agent': agent_key, 'message': f'Skill 安装失败：{exc}'})}")
                self.end_headers()
            return
        form = self.read_form()
        if parsed.path == "/run":
            code, stdout, stderr = run_pdca(date_text, push=False)
            message = "运行成功，结果已刷新。" if code == 0 else f"运行失败：{stderr or stdout}"
            self.redirect("/", date_text, message[:300])
        elif parsed.path == "/hermes-chat":
            query_text = (form.get("query", [""])[0] or "").strip()
            result = run_hermes_chat(query_text)
            self.send_html(render_home(date_text, hermes_result=result))
        elif parsed.path == "/pdca-task":
            task_date = (form.get("date", [date_text])[0] or date_text)
            try:
                message = save_pdca_task_update(form)
            except Exception as exc:
                message = f"保存失败：{exc}"
            self.redirect("/pdca-vps", task_date, message[:300])
        elif parsed.path == "/agent-soul":
            agent_key = parse_qs(parsed.query).get("agent", [""])[0]
            agent = agent_by_key(agent_key)
            if not agent:
                self.redirect("/", date_text, "未知 Agent。")
                return
            write_text(ensure_agent_soul(agent), form.get("content", [""])[0])
            self.send_response(303)
            self.send_header("Location", f"/agent-soul?{urlencode({'date': date_text, 'agent': agent_key, 'message': 'SOUL.md 已保存。'})}")
            self.end_headers()
        elif parsed.path == "/agent-core-file":
            query = parse_qs(parsed.query)
            agent_key = query.get("agent", [""])[0]
            active_file = query.get("file", ["SOUL.md"])[0]
            agent = agent_by_key(agent_key)
            if not agent or active_file not in AGENT_CORE_FILES:
                self.redirect("/", date_text, "未知 Agent 或文件。")
                return
            write_text(ensure_agent_core_file(agent, active_file), form.get("content", [""])[0])
            self.send_response(303)
            self.send_header("Location", f"/agent-edit?{urlencode({'date': date_text, 'agent': agent_key, 'file': active_file, 'message': f'{active_file} 已保存。'})}")
            self.end_headers()
        elif parsed.path == "/questionnaire":
            save_questionnaire(date_text, form)
            self.redirect("/questionnaire", date_text, "问卷已保存。")
        elif parsed.path == "/todos":
            append_todo(date_text, form)
            self.redirect("/todos", date_text, "代办已保存。")
        elif parsed.path == "/logistics":
            append_logistics(date_text, form)
            self.redirect("/logistics", date_text, "物流单号已保存。")
        else:
            self.send_response(404)
            self.end_headers()


def main():
    server = ThreadingHTTPServer((HOST, PORT), WorkbenchHandler)
    url = f"http://{HOST}:{PORT}/"
    print(f"数据岗位 PDCA 工作台已启动：{url}")
    threading.Thread(target=warm_identity_cache, daemon=True).start()
    threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass


if __name__ == "__main__":
    main()
