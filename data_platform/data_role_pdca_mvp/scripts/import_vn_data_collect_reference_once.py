# -*- coding: utf-8 -*-
"""
一次性从「Data collecet(5).xlsx」导入越南门店参考数据（区域/门店数为真实汇总）。

日常驾驶舱只读 modules/walkin_cockpit/data/vn_data_collect_reference.json，不读 xlsx。
"""
import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

WORKSPACE = Path(__file__).resolve().parents[1]
OUT = WORKSPACE / "modules" / "walkin_cockpit" / "data" / "vn_data_collect_reference.json"
DEFAULT_XLSX = Path(r"c:\Users\frank\Desktop\Data collecet(5).xlsx")

SHEET_STORE = {
    "ĐỒNG KHỞI": {"storeName": "Dong Khoi", "storeNameZh": "同启店"},
    "CARAVELLE": {"storeName": "Caravelle", "storeNameZh": "Caravelle"},
    "MAJESTIC": {"storeName": "MaJestic", "storeNameZh": "Majestic"},
    "REX": {"storeName": "Rex", "storeNameZh": "Rex"},
}
REGION_LABEL = "越南区"


def _num(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if not v:
        return None
    s = str(v).strip()[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def aggregate_sheet(ws, sheet_name):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return None
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {name: i for i, name in enumerate(header) if name}
    idx_date = col.get("Date", 0)
    idx_perf = col.get("Monthly Performance", 3)
    idx_served = col.get("Number of Customers Served", 4)
    idx_contact = col.get("Number of Customers add contact", 5)
    idx_store = col.get("belong which store", 2)

    by_month = defaultdict(
        lambda: {
            "days": 0,
            "performance": 0.0,
            "customersServed": 0,
            "contactsAdded": 0,
            "dates": [],
        }
    )
    store_label = SHEET_STORE.get(sheet_name, {}).get("storeName") or sheet_name

    for row in rows[1:]:
        if not row or not row[idx_date]:
            continue
        d = _parse_date(row[idx_date])
        if not d:
            continue
        ym = f"{d.year:04d}-{d.month:02d}"
        bucket = by_month[ym]
        bucket["days"] += 1
        bucket["dates"].append(d.isoformat())
        bucket["performance"] += _num(row[idx_perf] if idx_perf < len(row) else 0)
        bucket["customersServed"] += int(_num(row[idx_served] if idx_served < len(row) else 0))
        bucket["contactsAdded"] += int(_num(row[idx_contact] if idx_contact < len(row) else 0))
        if idx_store < len(row) and row[idx_store]:
            store_label = str(row[idx_store]).strip()

    meta = SHEET_STORE.get(sheet_name, {"storeName": store_label, "storeNameZh": store_label})
    return {
        "sheet": sheet_name,
        "storeName": store_label,
        "storeNameZh": meta.get("storeNameZh", store_label),
        "region": REGION_LABEL,
        "city": "胡志明",
        "byMonth": dict(by_month),
    }


def build_payload(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    stores = []
    month_keys = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        item = aggregate_sheet(ws, sheet_name)
        if item:
            stores.append(item)
            month_keys.update(item["byMonth"].keys())
    wb.close()

    months = {}
    for ym in sorted(month_keys):
        regions = [
            {
                "rg": REGION_LABEL,
                "storeCount": len(stores),
                "storeNames": [s["storeName"] for s in stores],
                "city": "胡志明",
            }
        ]
        store_details = []
        for s in stores:
            m = s["byMonth"].get(ym, {})
            store_details.append(
                {
                    "storeName": s["storeName"],
                    "storeNameZh": s["storeNameZh"],
                    "sheet": s["sheet"],
                    "daysInMonth": m.get("days", 0),
                    "customersServed": m.get("customersServed", 0),
                    "contactsAdded": m.get("contactsAdded", 0),
                    "performanceRaw": round(m.get("performance", 0), 4),
                }
            )
        months[ym] = {"regions": regions, "stores": store_details}

    return {
        "note": "参考 Data collecet(5).xlsx 越南 walk-in 四店日登记；驾驶舱只读 JSON。区域/门店数为真实，汇总表其余列为 mock。",
        "sourceFile": xlsx_path.name,
        "regionDefault": REGION_LABEL,
        "stores": [
            {
                "sheet": s["sheet"],
                "storeName": s["storeName"],
                "storeNameZh": s["storeNameZh"],
                "region": s["region"],
                "city": s["city"],
            }
            for s in stores
        ],
        "months": months,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()
    if not args.xlsx.is_file():
        raise SystemExit(f"找不到文件: {args.xlsx}")
    payload = build_payload(args.xlsx)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print("wrote", OUT)
    for ym, m in payload["months"].items():
        r = m["regions"][0]
        print(ym, r["rg"], "stores", r["storeCount"], r["storeNames"])


if __name__ == "__main__":
    main()
