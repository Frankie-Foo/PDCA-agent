# -*- coding: utf-8 -*-
"""
【可选一次性】从「代理商终销 Distribution Sell out.xlsx」导入经销商清单。
日常 build_walkin_bundle.py 只读 dealer_distribution_reference.json。
"""
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl

DEFAULT_XLSX = Path(r"c:\Users\frank\Downloads\代理商终销Distribution Sell out.xlsx")
OUT = Path(__file__).resolve().parents[1] / "modules" / "walkin_cockpit" / "data" / "dealer_distribution_reference.json"

REGION_MAP = {
    "中国": "东区", "香港": "南区", "澳门": "南区", "台湾": "东区", "新加坡": "南区",
    "马来西亚": "南区", "泰国": "越南区", "越南": "越南区", "印尼": "越南区",
    "印度尼西亚": "越南区", "菲律宾": "越南区", "柬埔寨": "越南区", "老挝": "越南区",
    "巴基斯坦": "西区", "哈萨克斯坦": "西区", "乌兹别克斯坦": "西区", "土库曼斯坦": "西区",
    "俄罗斯": "北区", "土耳其": "西区", "阿联酋": "南区", "沙特": "西区", "卡塔尔": "西区",
    "科威特": "西区", "伊拉克": "西区", "约旦": "西区", "黎巴嫩": "西区",
    "英国": "北区", "法国": "北区", "德国": "北区", "意大利": "北区",
    "美国": "北区", "加拿大": "北区", "墨西哥": "北区",
}


def _num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return 0.0


def _header_map(row):
    mapping = {}
    for idx, cell in enumerate(row):
        if cell is None:
            continue
        text = str(cell).replace("\n", " ").strip().lower()
        if "team" in text or text == "团队":
            mapping["team"] = idx
        elif "country" in text or "国家" in text:
            mapping["country"] = idx
        elif "store name" in text or "门店名称" in text:
            mapping["store_name"] = idx
        elif "customer types" in text or "客户类型" in text:
            mapping["ctype"] = idx
        elif "dealer name" in text or "代理商名称" in text:
            mapping["dealer"] = idx
        elif "sales" in text and "sell" not in text and "销售" in str(cell):
            mapping["sales_person"] = idx
        elif "amount" in text or "金额" in text:
            mapping.setdefault("amount_cols", []).append(idx)
        elif ("sell out" in text or "sellout" in text) and ("quantity" in text or "数量" in text):
            mapping.setdefault("qty_cols", []).append(idx)
    return mapping


def _region(country):
    c = (country or "").strip()
    if not c:
        return "东区"
    for key, region in REGION_MAP.items():
        if key in c:
            return region
    if any(x in c for x in ("欧", "美", "英", "德", "法", "俄")):
        return "北区"
    if any(x in c for x in ("亚", "新", "马", "泰", "越", "菲", "印尼")):
        return "越南区"
    return "西区"


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    dealers = []
    seen = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        colmap = None
        team_from_sheet = sheet_name.replace("组", "").strip()
        for i, row in enumerate(rows):
            if not row:
                continue
            cells = list(row)
            if any(c and ("Team" in str(c) or "团队" in str(c)) for c in cells[:3]):
                header_idx = i
                colmap = _header_map(cells)
                continue
            if header_idx is None or not colmap or "dealer" not in colmap:
                continue
            team = str(cells[colmap.get("team", 0)] or "").strip() if colmap.get("team") is not None else ""
            if not team:
                team = team_from_sheet
            country = str(cells[colmap["country"]] or "").strip() if "country" in colmap else ""
            dealer = str(cells[colmap["dealer"]] or "").strip()
            if not dealer or len(dealer) < 2:
                continue
            if dealer in ("Dealer Name", "代理商名称") or "开店" in dealer:
                continue
            key = (team, country, dealer)
            if key in seen:
                continue
            seen.add(key)
            ctype = ""
            if "ctype" in colmap and cells[colmap["ctype"]] is not None:
                ctype = str(cells[colmap["ctype"]]).strip().upper()[:1]
            amount = 0.0
            for ci in colmap.get("amount_cols") or []:
                amount += _num(cells[ci] if ci < len(cells) else None)
            qty = 0.0
            for ci in colmap.get("qty_cols") or []:
                qty += _num(cells[ci] if ci < len(cells) else None)
            store_name = ""
            if "store_name" in colmap and cells[colmap["store_name"]]:
                store_name = str(cells[colmap["store_name"]]).strip()
            if store_name in ("没店", "没有", "无"):
                store_name = ""
            sales_person = ""
            if "sales_person" in colmap and cells[colmap["sales_person"]]:
                sales_person = str(cells[colmap["sales_person"]]).strip()
            dealers.append({
                "team": team,
                "country": country,
                "region": _region(country),
                "dealerName": dealer,
                "storeName": store_name,
                "customerType": ctype or "S",
                "sellOutAmount": round(amount, 2),
                "sellOutQty": round(qty, 2),
                "salesPerson": sales_person,
                "sheet": sheet_name,
            })
    wb.close()
    return dealers


def main():
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    args = parser.parse_args()
    path = Path(args.xlsx)
    if not path.exists():
        raise SystemExit(f"找不到: {path}")
    dealers = parse_workbook(path)
    for i, d in enumerate(dealers, start=1):
        d["id"] = f"d{i:03d}"
    payload = {
        "note": "代理商终销经销商清单（参考 Distribution Sell out 表）。驾驶舱不读 xlsx。",
        "sourceFile": str(path.name),
        "importedAt": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "dealerCount": len(dealers),
        "dealers": dealers,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print("written", OUT, "dealers", len(dealers))


if __name__ == "__main__":
    main()
